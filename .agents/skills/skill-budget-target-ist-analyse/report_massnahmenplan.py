"""
Budget-Maßnahmenplan EKEK/1
Erzeugt Markdown mit Aufgabenbereich- und Firmen-Tabelle.
Maßnahmen-Spalte bleibt leer (Agent füllt nach User-Vorgabe).

Usage:
    python .agents/skills/skill-budget-massnahmenplan/report_massnahmenplan.py [--year YYYY]
"""

import argparse
import csv
import datetime
import os
import re
import sqlite3
import subprocess
import sys
from collections import defaultdict
from copy import copy
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", ".."))
BUDGET_DB = os.path.join(REPO_ROOT, "scripts", "budget", "budget_db.py")
DB_PATH = os.path.join(REPO_ROOT, "userdata", "budget.db")
OUT_DIR = os.path.join(REPO_ROOT, "userdata", "budget")
VORGABEN_DIR = os.path.join(OUT_DIR, "vorgaben")
TARGET_CSV = os.path.join(VORGABEN_DIR, "target.csv")
STATUS_CSV = os.path.join(SCRIPT_DIR, "status_mapping.csv")
PRAEMISSEN_MD = os.path.join(VORGABEN_DIR, "praemissen.md")
TARGET_IMAGE = os.path.join(VORGABEN_DIR, "target.png")
EA_MATRIX_TEMPLATE = os.path.join(VORGABEN_DIR, "budget_vorlage_ea_matrix.xlsx")
CONFIG_XLSX = os.path.join(REPO_ROOT, "userdata", "budget", "planning", "beauftragungsplanung_config.xlsx")
PLANNING_SCRIPTS_DIR = os.path.join(REPO_ROOT, "scripts", "budget")
if PLANNING_SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, PLANNING_SCRIPTS_DIR)

COMPANY_ALIAS_FALLBACKS = {
    "4SOFT": "4SOFT GMBH MUENCHEN",
    "THIESEN": "THIESEN HARDWARE SOFTW. GMBH WARTENBERG",
    "FES": "FES GMBH ZWICKAU/00",
    "SUMITOMO": "SUMITOMO ELECTRIC BORDNETZE SE WOLFSBURG",
    "BERTRANDT": "BERTRANDT INGENIEURBUERO GMBH TAPPENBECK",
    "EDAG": "EDAG ENGINEERING GMBH WOLFSBURG",
    "VOLKSWAGEN": "VOLKSWAGEN GROUP SERVICES GMBH WOLFSBURG",
}

try:
    from beauftragungsplanung_core import (  # noqa: E402
        CURRENT_QUARTER_TODO_STATUS,
        COMPANY_AREA_MAP,
        SYSTEMS_AREA,
        EDAG_COMPANY,
        BERTRANDT_COMPANY,
        status_label_for_raw_status as _core_status_label,
    )
except ImportError:  # pragma: no cover - fallback for missing planning core in local workspace
    CURRENT_QUARTER_TODO_STATUS = "02_Freigabe Kostenstelle"
    COMPANY_AREA_MAP: dict[str, list[str]] = {}
    SYSTEMS_AREA = "Systemschaltpläne und Bibl. (EDAG, Bertrandt)"
    EDAG_COMPANY = "EDAG ENGINEERING GMBH WOLFSBURG"
    BERTRANDT_COMPANY = "BERTRANDT INGENIEURBUERO GMBH TAPPENBECK"
    _core_status_label = None  # type: ignore[assignment]


def _quarter_for_date(target_date: str | None, *, title: str = "", bm_text: str = "") -> str | None:
    text = str(target_date or "").strip()
    match = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if not match:
        return None
    month = int(match.group(2))
    if month <= 3:
        return "Q1"
    if month <= 6:
        return "Q2"
    if month <= 9:
        return "Q3"
    return "Q4"

PDF_CSS = """
body {
  font-family: Arial, sans-serif;
  font-size: 8pt;
  line-height: 1.25;
}
h1, h2, h3 { margin: 10pt 0 4pt; }
p, li { margin: 2pt 0; }
hr {
  border: 0;
  border-top: 1px solid #999;
  margin: 8pt 0;
}
table {
  width: auto;
  border-collapse: collapse;
  table-layout: auto;
  font-size: 6.5pt;
  margin: 0 0 6pt 0;
}
th, td {
  border: 0.4pt solid #999;
  padding: 1pt 6pt;
  vertical-align: top;
  white-space: nowrap;
}
"""

SOURCE_NOTICE_PREFIX = "[SOURCE_NOTICE]"
SOURCE_TABLES = {"btl", "btl_opt"}

# ── CSV laden: 2025 + Target ──────────────────────────────────────────

def load_status_mapping() -> dict[str, str]:
    """Liest status_mapping.csv → {status: benennung}."""
    mapping: dict[str, str] = {}
    with open(STATUS_CSV, encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            mapping[row["Status"]] = row["Benennung"]
    return mapping


def load_targets() -> tuple[
    dict[str, int],
    dict[str, int],
    list[str],
    dict[str, int],
    dict[str, dict[str, int | str]],
    dict[str, Any] | None,
]:
    """Liest target.csv inkl. optionaler Quartalsaufteilung."""
    ref_2025: dict[str, int] = {}
    target_2026: dict[str, int] = {}
    start_q: dict[str, int] = {}
    area_order: list[str] = []
    company_target_overrides: dict[str, dict[str, int | str]] = {}
    quarter_split: dict[str, Any] | None = None
    with open(TARGET_CSV, encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=";")
        headers: list[str] = []
        mode: str | None = None
        for raw_row in reader:
            row = [cell.strip() for cell in raw_row]
            if not any(row):
                continue
            first = row[0]
            if first == "Aufgabenbereich":
                headers = row
                mode = "areas"
                continue
            if first == "Jahr/Quartal":
                headers = row
                mode = "quarters"
                continue
            if not headers or mode is None:
                continue

            record = {
                headers[idx]: row[idx].strip() if idx < len(row) else ""
                for idx in range(len(headers))
            }
            if mode == "areas":
                area = record["Aufgabenbereich"].strip()
                ref_2025[area] = int(record["2025"].strip())
                target_2026[area] = int(record["Target_2026"].strip())
                start_q[area] = int(record.get("Start_Q", "1").strip() or "1")
                area_order.append(area)
                for col_name, raw_value in record.items():
                    if col_name in {"Aufgabenbereich", "2025", "Target_2026", "Start_Q"}:
                        continue
                    if not col_name or not col_name.endswith("_Target_2026"):
                        continue
                    value = (raw_value or "").strip()
                    if not value:
                        continue
                    company_key = col_name.removesuffix("_Target_2026").upper()
                    company_target_overrides[company_key] = {
                        "area": area,
                        "target_te": int(value),
                        "start_q": start_q[area],
                    }
                continue

            if mode == "quarters" and first == "Summe":
                year_header = next((header for header in headers if re.fullmatch(r"\d{4}", header)), None)
                annual_te = _parse_te_value(record.get(year_header or ""))
                quarter_values = {quarter: _parse_te_value(record.get(quarter, "")) for quarter in QUARTERS}
                if year_header is None or annual_te is None or any(value is None for value in quarter_values.values()):
                    raise ValueError(f"Ungültige Quartalsaufteilung in {TARGET_CSV}.")
                quarter_sum_te = sum(float(quarter_values[quarter]) for quarter in QUARTERS)
                if abs(float(annual_te) - quarter_sum_te) > 1e-6:
                    raise ValueError(
                        f"Quartalsaufteilung in {TARGET_CSV} ist inkonsistent: Jahr {annual_te} T€ != Summe Quartale {quarter_sum_te} T€."
                    )
                quarter_split = {
                    "year": int(year_header),
                    "annual_te": float(annual_te),
                    "quarters_te": {quarter: float(quarter_values[quarter]) for quarter in QUARTERS},
                }
    return ref_2025, target_2026, area_order, start_q, company_target_overrides, quarter_split

# ── Prämissen laden ───────────────────────────────────────────────────

def load_praemissen() -> str:
    """Liest praemissen.md als String."""
    with open(PRAEMISSEN_MD, encoding="utf-8") as f:
        return f.read().strip()


def _normalize_source_table(source_table: str) -> str:
    normalized = str(source_table or "btl").strip().lower()
    if normalized not in SOURCE_TABLES:
        raise ValueError(f"Ungueltige Source-Tabelle: {source_table}")
    return normalized


def _source_notice_line(source_table: str) -> str | None:
    if _normalize_source_table(source_table) != "btl_opt":
        return None
    return (
        f"{SOURCE_NOTICE_PREFIX} OPTIMIERUNGSVORSCHLAG AUS btl_opt - "
        "keine Ist-Analyse aus btl"
    )


def _is_source_notice(line: str) -> bool:
    return str(line).startswith(SOURCE_NOTICE_PREFIX)


def _display_meta_line(line: str) -> str:
    if _is_source_notice(line):
        return str(line)[len(SOURCE_NOTICE_PREFIX):].strip()
    return line


def _render_meta_line_markdown(line: str) -> str:
    text = _display_meta_line(line)
    if _is_source_notice(line):
        return f'<span style="color:#c00000"><strong>{text}</strong></span>'
    return text


def _meta_line_font(line: str) -> Font:
    if _is_source_notice(line):
        return Font(size=10, bold=True, color="FFC00000")
    return Font(size=10, italic=True)


def _budget_source_sql(source_table: str) -> str:
    table = _normalize_source_table(source_table)
    return f"SELECT concept, dev_order, ea, title, planned_value, company, status, bm_text, target_date FROM {table}"


MONTH_COLUMNS = (
    "pct_jan",
    "pct_feb",
    "pct_mar",
    "pct_apr",
    "pct_may",
    "pct_jun",
    "pct_jul",
    "pct_aug",
    "pct_sep",
    "pct_oct",
    "pct_nov",
    "pct_dec",
)
QUARTERS = ("Q1", "Q2", "Q3", "Q4")
SPECIAL_STATUS_HEADERS = {"EL DIFF", "IO/nIO"}


def _normalize_ea_key(value: Any) -> str:
    digits = re.sub(r"[^0-9]", "", str(value or ""))
    if not digits:
        return ""
    return digits.lstrip("0") or "0"


def _parse_number_text(text: str) -> float | None:
    cleaned = text.strip()
    if not cleaned:
        return None
    if re.fullmatch(r"-?\d{1,3}(?:\.\d{3})+(?:,\d+)?", cleaned):
        return float(cleaned.replace(".", "").replace(",", "."))
    if re.fullmatch(r"-?\d+(?:,\d+)?", cleaned):
        return float(cleaned.replace(",", "."))
    if re.fullmatch(r"-?\d+\.\d+", cleaned):
        return float(cleaned)
    return None


def _parse_te_value(raw: Any) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if not text or text in {"-", "–"}:
        return None
    text = text.replace("T€", "").replace("€", "").replace(" ", "")
    if text.startswith("="):
        parts = [part for part in text[1:].split("+") if part]
        if parts:
            values = [_parse_te_value(part) for part in parts]
            if any(value is None for value in values):
                return None
            return sum(value for value in values if value is not None)
    return _parse_number_text(text)


def _parse_hour_value(raw: Any) -> float | None:
    if raw is None:
        return None
    text = str(raw).strip().lower()
    if not text or text in {"-", "–"}:
        return None
    text = text.removesuffix("h").replace(" ", "")
    if re.fullmatch(r"\d{1,3}(?:\.\d{3})+", text):
        return float(text.replace(".", ""))
    return _parse_number_text(text)


def _parse_el_target(raw: Any) -> dict[str, Any]:
    text = str(raw or "").strip()
    if not text or text in {"-", "–"}:
        return {"mode": "none", "value": None, "display": "-"}
    if text.lower().endswith("h"):
        hours = _parse_hour_value(text)
        if hours is None:
            return {"mode": "manual", "value": None, "display": text}
        return {"mode": "hours", "value": hours, "display": text}
    te_value = _parse_te_value(raw)
    if te_value is None:
        return {"mode": "manual", "value": None, "display": text}
    if te_value >= 1000 and "€" not in text.upper():
        return {"mode": "manual", "value": None, "display": text}
    return {"mode": "te", "value": te_value, "display": text}


def _resolve_company_list(raw: str, known_companies: list[str]) -> list[str]:
    if not raw:
        return []
    resolved: list[str] = []
    for token in re.split(r"[,;]+", raw):
        token = token.strip()
        if not token:
            continue
        upper = token.upper()
        match = next((c for c in known_companies if upper in c.upper()), None)
        if match and match not in resolved:
            resolved.append(match)
    return resolved


def _load_sondervorgaben(config_path: str, known_companies: list[str]) -> list[dict[str, Any]]:
    from pathlib import Path
    from planning_config_io import read_config

    if not os.path.isfile(config_path):
        return []
    raw_rules = read_config(Path(config_path)).sondervorgaben
    rules: list[dict[str, Any]] = []
    for rule in raw_rules:
        ea_keys = [_normalize_ea_key(ea) for ea in rule.get("ea_keys", []) if _normalize_ea_key(ea)]
        rules.append({
            "topic": rule["topic"],
            "ea_display": rule.get("ea_display", ", ".join(ea_keys)),
            "ea_keys": ea_keys,
            "fl_target_te": rule.get("fl_target_te"),
            "el_target": _parse_el_target(rule.get("el_target")),
            "remark": rule.get("remark", ""),
            "notes": rule.get("notes", ""),
            "cadence_type": rule.get("cadence_type", "annual_exact"),
            "allowed_companies": _resolve_company_list(rule.get("allowed_companies_raw", ""), known_companies),
            "priority_companies": _resolve_company_list(rule.get("priority_companies_raw", ""), known_companies),
        })
    return rules


def _annual_target_for_cadence(base_target: float | None, cadence_type: str) -> float | None:
    if base_target is None:
        return None
    return base_target


def _period_target_for_cadence(base_target: float | None, cadence_type: str, num_q: int) -> float | None:
    annual_target = _annual_target_for_cadence(base_target, cadence_type)
    if annual_target is None:
        return None
    if cadence_type == "first_half_exact":
        if num_q < 2:
            return 0.0
        return base_target
    if cadence_type == "quarterly_tranche_exact":
        return (base_target * num_q / 4) if base_target is not None and num_q >= 1 else 0.0
    if cadence_type == "semiannual_tranche_exact":
        if num_q < 2:
            return 0.0
        return base_target / 2 if base_target is not None else None
    return annual_target * num_q / 4


def _values_match(actual: float, target: float, *, tolerance: float = 0.1) -> bool:
    return abs(actual - target) <= tolerance


def _status_text(value: bool | None) -> str:
    if value is None:
        return "-"
    return "OK" if value else "X"


def _io_text(value: bool | None) -> str:
    if value is None:
        return "-"
    return "IO" if value else "nIO"


def _diff_band_status(actual: float, target: float | None, *, warn_ratio: float = 0.10, ok_ratio: float = 0.025) -> str:
    if target is None:
        return "-"
    if abs(target) <= 0.1:
        return "OK" if abs(actual - target) <= 0.1 else "X"
    ratio = abs(actual - target) / abs(target)
    if ratio < ok_ratio:
        return "OK"
    if ratio <= warn_ratio:
        return "WARN"
    return "X"


_BAND_SEVERITY = {"-": 0, "OK": 1, "WARN": 2, "X": 3}


def _worst_band(*bands: str) -> str:
    applicable = [b for b in bands if b != "-"]
    if not applicable:
        return "-"
    return max(applicable, key=lambda b: _BAND_SEVERITY.get(b, 0))


def _band_to_io(band: str) -> str:
    if band == "-":
        return "-"
    return "IO" if band == "OK" else "nIO"


def _evaluate_hint_status(hint: str, ctx: dict) -> str:
    h = hint.lower()

    # Rule 1: "EL und FL müssen exakt passen!"
    if "el und fl müssen exakt passen" in h or "el und fl müssen exakt passen" in h.replace("ü", "ue"):
        return _band_to_io(ctx["year_band"])

    # Rule 2: "FL muss exakt passen!" (without "el und fl")
    if "fl muss exakt passen" in h or "fl muss exakt passen" in h.replace("ü", "ue"):
        return _band_to_io(ctx["year_band"])

    # Rule 3: "EL und FL dürfen nur Quartalsweise beauftragen..."
    if "quartalsweise" in h and "beauftrag" in h:
        worst = _worst_band(ctx["quarter_band"], ctx["el_period_band"])
        return _band_to_io(worst)

    # Rule 4a: "Werte für 1. HJ müssen auch wirklich im 1.HJ beauftragt werden"
    # Entire annual budget must be ordered in Q1-2 → compare period actual vs annual target
    if "1. hj" in h and "beauftragt" in h:
        band = _diff_band_status(ctx["fl_period_te"], ctx.get("fl_target_annual"))
        return _band_to_io(band)

    # Rule 4b: "EL und FL müssen hier pro Halbjahr beauftragt werden"
    if "halbjahr" in h and "beauftragt" in h:
        return _band_to_io(ctx["quarter_band"])

    # Rule 5: "EL kann niedriger sein, darf aber auf keinen Fall höher sein."
    if "kann niedriger" in h:
        el_target = ctx.get("el_annual_target")
        el_actual = ctx.get("el_actual_annual", 0.0)
        if el_target is None or el_target == 0:
            return "-"
        if el_actual <= el_target:
            return "IO"
        ratio = (el_actual - el_target) / abs(el_target)
        if ratio <= 0.01:
            return "IO"
        if ratio <= 0.10:
            return "WARN"
        return "nIO"

    # Rule 6: "Folgende Firmen können auf ... buchen: ... in genau dieser Priorität."
    if "folgende firmen" in h and ("priorität" in h or "priorit" in h):
        checks = [ctx.get("allowed_company_ok"), ctx.get("priority_order_ok")]
        applicable = [c for c in checks if c is not None]
        if not applicable:
            return "-"
        return "IO" if all(applicable) else "nIO"

    # Rule 7: "Nur für ..."
    if h.startswith("nur für ") or h.startswith("nur f\u00fcr ") or h.startswith("nur fuer "):
        val = ctx.get("allowed_company_ok")
        return _io_text(val)

    # Rule 8: "Achtung: 4soft kann nur auf PMT und Digi-budget gebucht werden!"
    if "4soft kann nur" in h:
        val = ctx.get("four_soft_ok")
        return _io_text(val)

    # Rule 9: "Hier nur FL buchen."
    if "hier nur fl buchen" in h:
        el_actual = ctx.get("el_actual_annual", 0.0)
        return "IO" if abs(el_actual) <= 0.1 else "nIO"

    # Rule 10: "Nur EL. Keine FL."
    if "nur el" in h and "keine fl" in h:
        fl_actual = ctx.get("fl_actual_sum", 0.0)
        return "IO" if abs(fl_actual) <= 0.1 else "nIO"

    # Rule 11: "FL kann am Stück beauftragt werden..." (explicit exemption)
    if "am stück beauftragt" in h or "am stueck beauftragt" in h or "am st\u00fcck" in h:
        return "IO"

    # Rule 3b: "pro Quartal beauftragt werden" / "pro por Quartal" (typo variant)
    if ("pro quartal" in h or "pro por quartal" in h) and "beauftragt" in h:
        worst = _worst_band(ctx["quarter_band"], ctx["el_period_band"])
        return _band_to_io(worst)

    # Fallback: unmatched hint
    return "-"


def _special_rule_matches_budget_row(rule: dict[str, Any], row: dict[str, Any]) -> bool:
    topic = str(rule.get("topic", "")).upper()
    if "DIGI-BUDGET" in topic:
        title = str(row.get("title", "")).upper()
        bm_text = str(row.get("bm_text", "")).upper()
        return "VW-EK" in title and "DIGITALISIERUNG" in bm_text
    return True


def _load_el_aggregates(year: int, num_q: int) -> dict[str, dict[str, float]]:
    period_columns = MONTH_COLUMNS[: num_q * 3]
    year_expr = " + ".join(f"COALESCE({column}, 0)" for column in MONTH_COLUMNS)
    period_expr = " + ".join(f"COALESCE({column}, 0)" for column in period_columns) or "0"
    sql = f"""
        SELECT
            ea_number,
            ROUND(SUM((({year_expr}) / 1200.0) * year_work_hours), 2) AS year_hours,
            ROUND(SUM((({period_expr}) / 1200.0) * year_work_hours), 2) AS period_hours,
            ROUND(SUM((({year_expr}) / 1200.0) * year_work_hours * hourly_rate) / 1000.0, 2) AS year_te,
            ROUND(SUM((({period_expr}) / 1200.0) * year_work_hours * hourly_rate) / 1000.0, 2) AS period_te
        FROM el_planning
        GROUP BY ea_number
    """
    try:
        rows = run_sql(sql, year)
    except BaseException:
        return {}
    aggregates: dict[str, dict[str, float]] = {}
    for row in rows:
        key = _normalize_ea_key(row.get("ea_number"))
        if not key:
            continue
        aggregates[key] = {
            "year_hours": float(row.get("year_hours") or 0),
            "period_hours": float(row.get("period_hours") or 0),
            "year_te": float(row.get("year_te") or 0),
            "period_te": float(row.get("period_te") or 0),
        }
    return aggregates


def build_special_rule_section(
    *,
    year: int,
    num_q: int,
    q_label: str,
    rows: list[dict],
    status_map: dict[str, str],
) -> Any:
    ea_budget_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    known_companies: set[str] = set()
    for row in rows:
        ea_key = _normalize_ea_key(row.get("dev_order") or row.get("ea_number") or row.get("ea"))
        if not ea_key:
            continue
        pv = int(float(row.get("planned_value", "0") or 0))
        company = row.get("company", "")
        title = row.get("title", "")
        bm_text = row.get("bm_text", "")
        status_label = status_map.get(row.get("status", ""), "")
        if not status_label and _core_status_label is not None:
            status_label = _core_status_label(row.get("status", ""))
        reporting_company = _reporting_company(company)
        ea_budget_rows[ea_key].append(
            {
                "planned_value": pv,
                "reporting_company": reporting_company,
                "status_label": status_label,
                "title": title,
                "bm_text": bm_text,
            }
        )
        known_companies.add(reporting_company)

    special_rules = _load_sondervorgaben(CONFIG_XLSX, sorted(known_companies))
    if not special_rules:
        return None

    el_aggregates = _load_el_aggregates(year, num_q)
    pmt_or_digi_eas = {
        ea_key
        for rule in special_rules
        if "PMT" in rule["topic"].upper() or "DIGI-BUDGET" in rule["topic"].upper()
        for ea_key in rule["ea_keys"]
    }
    rows_out: list[TableRow] = []

    for rule in special_rules:
        totals = {
            "count": 0,
            "sum_te": 0.0,
            "bestellt_te": 0.0,
            "durchlauf_te": 0.0,
            "konzept_te": 0.0,
            "storniert_te": 0.0,
            "period_te": 0.0,
            "el_year_hours": 0.0,
            "el_period_hours": 0.0,
            "el_year_te": 0.0,
            "el_period_te": 0.0,
        }
        company_totals: dict[str, int] = defaultdict(int)
        for ea_key in rule["ea_keys"]:
            for budget_row in ea_budget_rows.get(ea_key, []):
                if not _special_rule_matches_budget_row(rule, budget_row):
                    continue
                pv = int(budget_row["planned_value"])
                totals["count"] += 1
                totals["sum_te"] += pv / 1000
                company_totals[str(budget_row["reporting_company"])] += pv
                status_label = str(budget_row["status_label"])
                if status_label == "bestellt":
                    totals["bestellt_te"] += pv / 1000
                elif status_label == "im Durchlauf":
                    totals["durchlauf_te"] += pv / 1000
                elif status_label == "Konzept":
                    totals["konzept_te"] += pv / 1000
                elif status_label == "storniert":
                    totals["storniert_te"] += pv / 1000
            el_row = el_aggregates.get(ea_key)
            if el_row:
                totals["el_year_hours"] += el_row["year_hours"]
                totals["el_period_hours"] += el_row["period_hours"]
                totals["el_year_te"] += el_row["year_te"]
                totals["el_period_te"] += el_row["period_te"]
        totals["period_te"] = totals["bestellt_te"] + totals["durchlauf_te"]

        fl_target_annual = _annual_target_for_cadence(rule["fl_target_te"], rule["cadence_type"])
        fl_target_period = _period_target_for_cadence(rule["fl_target_te"], rule["cadence_type"], num_q)
        year_band = _diff_band_status(totals["sum_te"], fl_target_annual)
        quarter_band = _diff_band_status(totals["period_te"], fl_target_period)
        year_ok = None if fl_target_annual is None else year_band == "OK"
        quarter_ok = None if fl_target_period is None else quarter_band == "OK"

        el_diff_display = "-"
        el_diff_band = "-"
        el_period_band = "-"
        el_annual_target: float | None = None
        el_actual_annual: float = 0.0
        el_mode = rule["el_target"]["mode"]
        if el_mode in {"hours", "te"}:
            el_actual_annual = totals["el_year_hours"] if el_mode == "hours" else totals["el_year_te"]
            actual_period = totals["el_period_hours"] if el_mode == "hours" else totals["el_period_te"]
            el_base_target = rule["el_target"]["value"]
            if el_base_target is not None:
                el_annual_target = _annual_target_for_cadence(el_base_target, rule["cadence_type"])
                el_period_target = _period_target_for_cadence(el_base_target, rule["cadence_type"], num_q)
                if el_annual_target is not None:
                    if el_mode == "hours":
                        el_diff_display = delta_hours_fmt(el_actual_annual - el_annual_target)
                    else:
                        el_diff_display = delta_fmt(el_actual_annual - el_annual_target)
                if el_annual_target is not None:
                    el_diff_band = _diff_band_status(el_actual_annual, el_annual_target)
                el_period_band = _diff_band_status(actual_period, el_period_target)

        active_companies = {
            company
            for company, value in company_totals.items()
            if abs(value) > 0
        }
        allowed_company_ok: bool | None = None
        if rule["allowed_companies"]:
            allowed_company_ok = active_companies.issubset(set(rule["allowed_companies"]))
        priority_order_ok: bool | None = None
        if rule["priority_companies"]:
            rank_map = {company: idx for idx, company in enumerate(rule["priority_companies"], start=1)}
            active_ranks = sorted(rank_map[company] for company in active_companies if company in rank_map)
            priority_order_ok = active_ranks == list(range(1, len(active_ranks) + 1))
        four_soft_ok: bool | None = None
        four_soft = COMPANY_ALIAS_FALLBACKS["4SOFT"]
        if four_soft in active_companies:
            four_soft_ok = set(rule["ea_keys"]).issubset(pmt_or_digi_eas)

        ctx = {
            "year_band": year_band,
            "quarter_band": quarter_band,
            "el_diff_band": el_diff_band,
            "el_period_band": el_period_band,
            "el_annual_target": el_annual_target,
            "el_actual_annual": el_actual_annual,
            "el_mode": el_mode,
            "allowed_company_ok": allowed_company_ok,
            "priority_order_ok": priority_order_ok,
            "four_soft_ok": four_soft_ok,
            "fl_actual_sum": totals["sum_te"],
            "fl_target_annual": fl_target_annual,
            "fl_period_te": totals["period_te"],
        }

        base_values = [
            rule["topic"],
            rule["ea_display"] or "-",
            fmt(fl_target_annual) if fl_target_annual is not None else "-",
            fmt(totals["sum_te"]),
            delta_fmt(totals["sum_te"] - fl_target_annual) if fl_target_annual is not None else "-",
            fmt(fl_target_period) if fl_target_period is not None else "-",
            fmt(totals["bestellt_te"]),
            fmt(totals["durchlauf_te"]),
            fmt(totals["konzept_te"]),
            el_diff_display,
            delta_fmt(totals["period_te"] - fl_target_period) if fl_target_period is not None else "-",
            rule["remark"],
        ]

        hints = [h.strip() for h in rule["notes"].split("\n") if h.strip()]
        if not hints:
            rows_out.append(
                TableRow(
                    base_values + ["-", "-"],
                    cell_statuses={
                        "DIFF Ges.": year_band,
                        f"DIFF {q_label}": quarter_band,
                        "EL DIFF": el_diff_band,
                        "IO/nIO": "-",
                    },
                )
            )
        else:
            for hint_idx, hint_text in enumerate(hints):
                hint_status = _evaluate_hint_status(hint_text, ctx)
                if hint_idx == 0:
                    rows_out.append(
                        TableRow(
                            base_values + [hint_text, hint_status],
                            cell_statuses={
                                "DIFF Ges.": year_band,
                                f"DIFF {q_label}": quarter_band,
                                "EL DIFF": el_diff_band,
                                "IO/nIO": hint_status,
                            },
                        )
                    )
                else:
                    rows_out.append(
                        TableRow(
                            [""] * 12 + [hint_text, hint_status],
                            cell_statuses={"IO/nIO": hint_status},
                        )
                    )

    return TableSection(
        title="Sondervereinbarungen",
        headers=[
            "Thema",
            "EA",
            "Soll",
            "Ist",
            "DIFF Ges.",
            f"Soll {q_label}",
            "bestellt",
            "im Durchlauf",
            "01 Erstellung",
            "EL DIFF",
            f"DIFF {q_label}",
            "Bemerkung",
            "Hinweise",
            "IO/nIO",
        ],
        rows=rows_out,
        align_right={1, 2, 3, 4, 5, 6, 7, 8, 10},
        body_row_height=15,
    )


def write_pdf_beside_markdown(md_file: str) -> str | None:
    """Best-effort Markdown->PDF Export im A4-Querformat."""
    pdf_file = os.path.splitext(md_file)[0] + ".pdf"
    try:
        from markdown_pdf import MarkdownPdf, Section
    except ImportError as exc:
        print(f"WARNUNG: PDF-Export übersprungen: markdown-pdf nicht installiert ({exc})", file=sys.stderr)
        return None

    try:
        with open(md_file, encoding="utf-8") as f:
            md_text = f.read()
        pdf = MarkdownPdf(toc_level=2, optimize=True)
        pdf.add_section(Section(md_text, paper_size="A4-L"), user_css=PDF_CSS)
        pdf.save(pdf_file)
        return pdf_file
    except Exception as exc:  # PDF ist Zusatzartefakt; Markdown/XLSX bleiben gültig.
        print(f"WARNUNG: PDF-Export fehlgeschlagen: {exc}", file=sys.stderr)
        return None


def open_excel_file(xlsx_file: str) -> None:
    """Öffnet die erzeugte XLSX-Datei auf Windows in der Standardanwendung (Excel)."""
    if os.name != "nt":
        return
    try:
        os.startfile(xlsx_file)
    except OSError as exc:
        print(f"WARNUNG: Excel-Datei konnte nicht automatisch geöffnet werden: {exc}", file=sys.stderr)

# ── AUDI fest (aus Target-CSV abgeleitet) ─────────────────────────────
AUDI_KEY = "AUDI - Bibliotheksarbeiten (Audi)"
THIESEN_SUPPORT_KEY = "Bordnetz Support, RollOut (Thiesen)"
THIESEN_SPEC_KEY = "Spez. und Test VOBES2025 (Thiesen)"

# ── Firma→Aufgabenbereich Mapping ─────────────────────────────────────

def _parse_gewerk_numbers(bm_text: str) -> set[int]:
    """Extrahiert alle Gewerk-Nummern (#1, #2, …) aus dem BM-Text."""
    numbers: set[int] = set()
    for m in re.finditer(r'Gewerk\s+#([\d,#\s]+)', bm_text, re.IGNORECASE):
        for n in re.findall(r'\d+', m.group(1)):
            numbers.add(int(n))
    return numbers


def classify_bm(company: str, title: str, bm_text: str = "") -> str:
    """Ordnet eine BM anhand Firma+Titel+BM-Text einem Aufgabenbereich zu."""
    c = company.upper()
    t = title.upper() if title else ""

    if "EDAG" in c or "BERTRANDT" in c:
        return "Systemschaltpläne und Bibl. (EDAG, Bertrandt)"
    if "GROUP SERVICES" in c or "T-SYSTEMS" in c:
        return "CATIA-Bibl. (GroupServices)"
    if "FES " in c or "FEV " in c:
        return "Projektbüro / Prüfbüro (FES, B&W)"
    if "MSG " in c:
        return "BordnetzGPT"
    if "PEC " in c:
        return "SYS-Flow (PEC)"
    if "SUMITOMO" in c or "SEBN" in c:
        return "Pilot und Anwendertest VOBES2025 (SEBN)"
    if "VOITAS" in c:
        return "RuleChecker (4soft, ex Voitas)"

    # 4SOFT: Split nach BM-Titel
    if "4SOFT" in c:
        if any(kw in t for kw in ["TE-PMT", "BORDNETZ", "KONZEPTENTW", "INTEGRATION"]):
            return "SW-Entwicklung VOBES2025 (4soft)"
        return "Vorentwicklung (4soft)"

    # THIESEN: Split nach Gewerk-Nummern im BM-Text
    #   nur Gewerk #2           → Spezifikation (Stückpreis-Abruf)
    #   mehrere Gewerke (#1,#2,#3,#5 + #4 FPs) → Bordnetz Support/RollOut
    if "THIESEN" in c:
        gewerke = _parse_gewerk_numbers(bm_text)
        if gewerke and gewerke != {2}:
            return "Bordnetz Support, RollOut (Thiesen)"
        return "Spez. und Test VOBES2025 (Thiesen)"

    # Fallback: KST/Werk/Sonstige → ignorieren (planned_value meist 0)
    return "_SONSTIGE"


def run_sql(sql: str, year: int) -> list[dict]:
    """Führt SQL nach budget_db.py-Sync direkt auf SQLite aus."""
    sql_oneline = " ".join(sql.split())
    cmd = [
        sys.executable, BUDGET_DB, "query", sql_oneline,
        "--no-file", "--sync", "--year", str(year),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, cwd=REPO_ROOT)
    if result.returncode != 0:
        raise RuntimeError(f"budget_db.py query fehlgeschlagen:\n{result.stderr}")
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(sql_oneline).fetchall()
    return [dict(row) for row in rows]


def fmt(v: float | int) -> str:
    number = float(v)
    if number.is_integer():
        return f"{int(number):,}".replace(",", ".") + " T€"
    text = f"{number:,.1f}"
    return text.replace(",", "X").replace(".", ",").replace("X", ".") + " T€"


def delta_fmt(d: float | int) -> str:
    sign = "+" if d > 0 else ""
    return sign + fmt(d)


def fmt_hours(v: float | int) -> str:
    number = float(v)
    if number.is_integer():
        return f"{int(number):,}".replace(",", ".") + " h"
    text = f"{number:,.1f}"
    return text.replace(",", "X").replace(".", ",").replace("X", ".") + " h"


def delta_hours_fmt(d: float | int) -> str:
    sign = "+" if d > 0 else ""
    return sign + fmt_hours(d)


@dataclass
class TableRow:
    values: list[str]
    style: str = "body"
    cell_statuses: dict[str, str] = field(default_factory=dict)


@dataclass
class TextSection:
    title: str | None = None
    lines: list[str] = field(default_factory=list)
    level: int = 2
    separator_before: bool = True
    sheet_name: str = "Übersicht"


@dataclass
class TableSection:
    title: str
    headers: list[str]
    rows: list[TableRow]
    align_right: set[int] = field(default_factory=set)
    intro_lines: list[str] = field(default_factory=list)
    level: int = 2
    separator_before: bool = True
    sheet_name: str = "Übersicht"
    body_row_height: float | None = None


@dataclass
class ReportDocument:
    title: str
    meta_lines: list[str]
    sections: list[TextSection | TableSection]
    max_columns: int
    overview_quarter_block: dict[str, Any] | None = None
    ea_matrix: dict[str, Any] | None = None


def _cell_has_text(value) -> bool:
    return value is not None and str(value).strip() != ""


def _copy_cell_display(src, dst, *, copy_value: bool = True) -> None:
    if copy_value:
        dst.value = src.value
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.comment = copy(src.comment) if src.comment else None


def _find_previous_xlsx(out_dir: str, current_file: str) -> str | None:
    current_name = os.path.basename(current_file)
    candidates: list[str] = []
    for name in os.listdir(out_dir):
        if name.startswith("~$"):
            continue
        if name == current_name or not name.endswith("_budget_massnahmenplan_ekek1.xlsx"):
            continue
        path = os.path.join(out_dir, name)
        if os.path.isfile(path):
            candidates.append(path)
    candidates.sort(key=os.path.getmtime, reverse=True)
    return candidates[0] if candidates else None


def _find_table_header_row(ws, first_header: str) -> int | None:
    needle = _strip_markdown(first_header)
    for row_idx in range(1, ws.max_row + 1):
        if _strip_markdown(str(ws.cell(row_idx, 1).value or "")) == needle:
            return row_idx
    return None


def _write_overview_quarter_block(ws, overview_quarter_block: dict[str, Any] | None) -> None:
    if not overview_quarter_block:
        return

    row_labels = ("IST 2025 (Referenz)", "Ist (BPLUS)", "Target", "Delta Ist vs. Target")
    row_map: dict[str, int] = {}
    for row_idx in range(1, ws.max_row + 1):
        label = _strip_markdown(str(ws.cell(row_idx, 1).value or ""))
        if label in row_labels:
            row_map[label] = row_idx
    if len(row_map) != len(row_labels):
        return

    header_row = row_map["IST 2025 (Referenz)"] - 1
    header_template = ws.cell(header_row, 2)
    ws.cell(header_row, 2).value = "Jahr"

    for col_idx, quarter in enumerate(overview_quarter_block.get("headers", QUARTERS), start=3):
        cell = ws.cell(header_row, col_idx)
        _copy_cell_display(header_template, cell, copy_value=False)
        cell.value = quarter
        ws.column_dimensions[get_column_letter(col_idx)].width = max(ws.column_dimensions[get_column_letter(col_idx)].width or 0, 12)

    delta_year_cell = ws.cell(row_map["Delta Ist vs. Target"], 2)
    if isinstance(delta_year_cell.value, (int, float)):
        _apply_signed_delta_style(delta_year_cell, delta_year_cell.value)

    for label in row_labels:
        template_cell = ws.cell(row_map[label], 2)
        for col_idx, value in enumerate(overview_quarter_block["rows"].get(label, []), start=3):
            cell = ws.cell(row_map[label], col_idx)
            _copy_cell_display(template_cell, cell, copy_value=False)
            cell.value = value
            cell.number_format = _te_number_format(value)
            if label == "Delta Ist vs. Target":
                _apply_signed_delta_style(cell, value)


def _collect_table_key_rows(ws, first_header: str) -> tuple[int | None, dict[str, int]]:
    header_row = _find_table_header_row(ws, first_header)
    if header_row is None:
        return None, {}

    rows: dict[str, int] = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        first_value = _strip_markdown(str(ws.cell(row_idx, 1).value or ""))
        if not first_value:
            if rows:
                break
            continue
        if rows and first_value in {"Firma", "Konzept", "Korrektur Überplanung", "Prämissen", "Gesamtübersicht", "Target - Ist - Vergleich"}:
            break
        nonempty = sum(1 for col_idx in range(1, ws.max_column + 1) if _cell_has_text(ws.cell(row_idx, col_idx).value))
        if rows and nonempty == 1 and first_value not in {"Summe", "Summe inkl. Audi", "Gesamt"}:
            break
        rows[first_value] = row_idx
    return header_row, rows


def _parse_correction_section(first_value: str) -> tuple[str, str] | None:
    text = _strip_markdown(first_value)
    match = re.match(r"^(?P<firm>.+?)\s+—\s+(?P<section>Stornierte Vorgänge|Quartals-Korrektur|Jahres-Korrektur)", text)
    if not match:
        return None
    section = match.group("section")
    kind = {
        "Stornierte Vorgänge": "storno",
        "Quartals-Korrektur": "quartal",
        "Jahres-Korrektur": "jahr",
    }[section]
    return match.group("firm").strip(), kind


def _collect_correction_rows(ws) -> dict[tuple[str, str, str], int]:
    rows: dict[tuple[str, str, str], int] = {}
    current_section: tuple[str, str] | None = None
    for row_idx in range(1, ws.max_row + 1):
        first_value = _strip_markdown(str(ws.cell(row_idx, 1).value or ""))
        if not first_value:
            continue
        parsed = _parse_correction_section(first_value)
        if parsed is not None:
            current_section = parsed
            continue
        if first_value == "Firma":
            continue
        concept_value = _strip_markdown(str(ws.cell(row_idx, 3).value or ""))
        if current_section is None or not re.fullmatch(r"\d+", concept_value):
            continue
        rows[(current_section[0], current_section[1], concept_value)] = row_idx
    return rows


def _copy_note_cell(source_ws, target_ws, source_row: int, source_col: int, target_row: int, target_col: int) -> None:
    source_cell = source_ws.cell(source_row, source_col)
    if not _cell_has_text(source_cell.value):
        return
    target_cell = target_ws.cell(target_row, target_col)
    _copy_cell_display(source_cell, target_cell, copy_value=True)
    source_height = source_ws.row_dimensions[source_row].height
    if source_height is not None:
        target_ws.row_dimensions[target_row].height = source_height


def _column_width_to_pixels(width: float | None) -> int:
    width = 8.43 if width is None else width
    return int(width * 7 + 5)


def _row_height_to_pixels(height: float | None) -> int:
    height = 15 if height is None else height
    return int(height * 4 / 3)


def _embed_target_image(ws, image_path: str) -> None:
    if not os.path.isfile(image_path):
        return
    img = XLImage(image_path)
    start_col = 8   # H
    end_col = 12    # L
    start_row = 2
    end_row = 25
    target_width = sum(
        _column_width_to_pixels(ws.column_dimensions[get_column_letter(col_idx)].width)
        for col_idx in range(start_col, end_col + 1)
    )
    target_height = sum(
        _row_height_to_pixels(ws.row_dimensions[row_idx].height)
        for row_idx in range(start_row, end_row + 1)
    )
    img.width = target_width
    img.height = target_height
    ws.add_image(img, "H2")


def _inherit_notes_from_workbook(workbook: Workbook, source_xlsx: str | None) -> None:
    if not source_xlsx or not os.path.isfile(source_xlsx):
        return

    source_wb = load_workbook(source_xlsx)
    source_overview = source_wb["Übersicht"] if "Übersicht" in source_wb.sheetnames else None
    target_overview = workbook["Übersicht"] if "Übersicht" in workbook.sheetnames else None
    if source_overview and target_overview:
        source_f_width = source_overview.column_dimensions["F"].width
        if source_f_width is not None:
            target_overview.column_dimensions["F"].width = source_f_width
        source_area_header, source_area_rows = _collect_table_key_rows(source_overview, "Aufgabenbereich")
        target_area_header, target_area_rows = _collect_table_key_rows(target_overview, "Aufgabenbereich")
        for key, target_row in target_area_rows.items():
            source_row = source_area_rows.get(key)
            if source_row is not None:
                _copy_note_cell(source_overview, target_overview, source_row, 6, target_row, 6)

        source_firm_header, source_firm_rows = _collect_table_key_rows(source_overview, "Firma")
        target_firm_header, target_firm_rows = _collect_table_key_rows(target_overview, "Firma")
        if target_firm_header is not None:
            target_header_cell = target_overview.cell(target_firm_header, 13)
            if source_firm_header is not None and _cell_has_text(source_overview.cell(source_firm_header, 13).value):
                _copy_cell_display(source_overview.cell(source_firm_header, 13), target_header_cell, copy_value=True)
            else:
                template_header = target_overview.cell(target_firm_header, 12)
                _copy_cell_display(template_header, target_header_cell, copy_value=False)
                target_header_cell.value = "Notizen"
            source_width = source_overview.column_dimensions["M"].width
            target_overview.column_dimensions["M"].width = source_width if source_width is not None else 65
        for key, target_row in target_firm_rows.items():
            source_row = source_firm_rows.get(key)
            if source_row is not None:
                _copy_note_cell(source_overview, target_overview, source_row, 13, target_row, 13)

    source_correction = source_wb["Korrektur"] if "Korrektur" in source_wb.sheetnames else None
    target_correction = workbook["Korrektur"] if "Korrektur" in workbook.sheetnames else None
    if source_correction and target_correction:
        source_j_width = source_correction.column_dimensions["J"].width
        if source_j_width is None:
            source_j_width = source_correction.column_dimensions["G"].width
        if source_j_width is None:
            source_j_width = source_correction.column_dimensions["F"].width
        if source_j_width is not None:
            target_correction.column_dimensions["J"].width = source_j_width
        source_rows = _collect_correction_rows(source_correction)
        target_rows = _collect_correction_rows(target_correction)
        for key, target_row in target_rows.items():
            source_row = source_rows.get(key)
            if source_row is not None:
                source_col = 10
                if not _cell_has_text(source_correction.cell(source_row, source_col).value):
                    source_col = 7
                if not _cell_has_text(source_correction.cell(source_row, source_col).value):
                    source_col = 6
                _copy_note_cell(source_correction, target_correction, source_row, source_col, target_row, 10)


def _heading(level: int, title: str) -> str:
    return f"{'#' * level} {title}"


def _strip_markdown(line: str) -> str:
    text = line.strip()
    if text == "---":
        return ""
    text = re.sub(r"^#{1,6}\s*", "", text)
    text = text.replace("**", "").replace("`", "")
    if text.startswith("_") and text.endswith("_") and len(text) > 1:
        text = text[1:-1]
    return text


def _markdown_heading_level(line: str) -> int | None:
    match = re.match(r"^(#{1,6})\s+.+$", line.strip())
    return len(match.group(1)) if match else None


def _markdown_table_cell(value: str, style: str) -> str:
    if not value:
        return value
    if style == "summary":
        return f"**{value}**"
    if style == "note":
        return f"_{value}_"
    return value


def _parse_te_number(value: str) -> float | int | None:
    text = _strip_markdown(str(value))
    match = re.fullmatch(r"([+-]?)(\d{1,3}(?:\.\d{3})*|\d+)(?:,(\d+))?\s*T€", text)
    if not match:
        return None
    sign, digits, decimals = match.groups()
    normalized = digits.replace(".", "")
    number = float(f"{normalized}.{decimals}") if decimals else int(normalized)
    return -number if sign == "-" else number


def _te_number_format(value: float | int) -> str:
    return '#,##0.0 "T€";[Red]-#,##0.0 "T€"' if float(value) % 1 else '#,##0 "T€";[Red]-#,##0 "T€"'


def _round_te_display(value: float | int) -> int:
    return int(round(float(value)))


def _apply_signed_delta_style(cell, value: float | int) -> None:
    cell.number_format = '#,##0 "T€";-#,##0 "T€"'
    if value > 0:
        cell.font = copy(cell.font)
        cell.font = Font(
            name=cell.font.name,
            sz=cell.font.sz,
            b=cell.font.b,
            i=cell.font.i,
            underline=cell.font.underline,
            strike=cell.font.strike,
            color="FFFF0000",
        )
        cell.fill = PatternFill("solid", fgColor="FFFFC7CE")
    elif value < 0:
        cell.font = copy(cell.font)
        cell.font = Font(
            name=cell.font.name,
            sz=cell.font.sz,
            b=cell.font.b,
            i=cell.font.i,
            underline=cell.font.underline,
            strike=cell.font.strike,
            color="FF000000",
        )
        cell.fill = PatternFill("solid", fgColor="FFC6EFCE")


def _quarter_split_period_target_eur(quarter_split: dict[str, Any] | None, num_q: int) -> int | None:
    if not quarter_split:
        return None
    total_te = sum(float(quarter_split["quarters_te"][quarter]) for quarter in QUARTERS[:num_q])
    return round(total_te * 1000)


def _quarter_index(quarter: str) -> int:
    return QUARTERS.index(str(quarter).upper()) + 1


def _scale_distribution_to_target(values: dict[str, int], target_total: int | None) -> dict[str, int]:
    if target_total is None:
        return dict(values)
    positive = {key: value for key, value in values.items() if value > 0}
    if target_total <= 0 or not positive:
        return {key: 0 for key in values}

    current_total = sum(positive.values())
    scaled = {key: 0 for key in values}
    remainders: list[tuple[float, str]] = []
    assigned = 0
    for key, value in positive.items():
        raw_value = value * target_total / current_total
        base_value = int(raw_value)
        scaled[key] = base_value
        assigned += base_value
        remainders.append((raw_value - base_value, key))

    remainders.sort(reverse=True)
    for _, key in remainders[: max(target_total - assigned, 0)]:
        scaled[key] += 1
    return scaled


def _bplus_vorgang_url(value: str) -> str | None:
    text = _strip_markdown(str(value))
    return f"https://bplus-ng-mig.r02.vwgroup.com/ek/view?id={text}" if re.fullmatch(r"\d+", text) else None


def _exclude_from_budget(benennung: str) -> bool:
    return False


def _hide_from_firm_overview(company: str) -> bool:
    return "VOITAS" in company.upper()


def _reporting_company(company: str) -> str:
    if "VOITAS" in company.upper():
        return "4SOFT GMBH MUENCHEN"
    return company


def _new_firm_totals() -> dict[str, Any]:
    return {
        "sum": 0,
        "count": 0,
        "bestellt": 0,
        "durchlauf": 0,
        "konzept": 0,
        "storniert": 0,
        "konzept_quarters": defaultdict(int),
    }


def _row_ea_number(row: dict[str, Any]) -> str:
    return str(row.get("dev_order") or row.get("ea") or "").strip()


def _load_ea_title_by_number(year: int) -> dict[str, str]:
    if not os.path.isfile(DB_PATH):
        return {}
    titles: dict[str, str] = {}
    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            for row in conn.execute(
                """
                SELECT dev_order, ea
                FROM btl
                WHERE substr(COALESCE(target_date, ''), 1, 4) = ?
                  AND COALESCE(dev_order, '') <> ''
                  AND COALESCE(ea, '') <> ''
                ORDER BY dev_order
                """,
                (str(year),),
            ).fetchall():
                ea_number = str(row["dev_order"] or "").strip()
                ea_title = str(row["ea"] or "").strip()
                if ea_number and ea_title and ea_number != ea_title and not re.fullmatch(r"[0-9]+", ea_title):
                    titles.setdefault(ea_number, ea_title)
            for row in conn.execute(
                """
                SELECT ea_number, title
                FROM devorder
                WHERE COALESCE(ea_number, '') <> ''
                  AND COALESCE(title, '') <> ''
                ORDER BY ea_number
                """
            ).fetchall():
                ea_number = str(row["ea_number"] or "").strip()
                title = str(row["title"] or "").strip()
                if ea_number and title:
                    titles.setdefault(ea_number, title)
    except sqlite3.OperationalError:
        return titles
    return titles


def _display_ea_title(ea_value: str, ea_number: str, ea_title_by_number: dict[str, str] | None = None) -> str:
    ea_text = str(ea_value or "").strip()
    ea_number_text = str(ea_number or "").strip()
    if ea_title_by_number and ea_number_text and (ea_text == ea_number_text or re.fullmatch(r"[0-9]+", ea_text)):
        resolved = str(ea_title_by_number.get(ea_number_text) or "").strip()
        if resolved:
            return resolved
    return ea_text


def _budget_entry_from_row(
    row: dict[str, Any],
    status_map: dict[str, str],
    ea_title_by_number: dict[str, str] | None = None,
) -> dict[str, Any] | None:
    pv = int(float(row.get("planned_value", "0")))
    company = str(row.get("company", "") or "")
    ea_number = _row_ea_number(row)
    ea = _display_ea_title(str(row.get("ea", "") or ""), ea_number, ea_title_by_number)
    title = str(row.get("title", "") or "")
    status = str(row.get("status", "") or "")
    bm_text = str(row.get("bm_text", "") or "")
    status_label = status_map.get(status, "")
    if not status_label and _core_status_label is not None:
        status_label = _core_status_label(status)
    if _exclude_from_budget(status_label):
        return None
    quarter = _quarter_for_date(row.get("target_date"), title=title, bm_text=bm_text)
    area = classify_bm(company, title, bm_text)
    reporting_company = _reporting_company(company)
    return {
        "concept": row.get("concept", ""),
        "ea": ea,
        "ea_number": ea_number,
        "title": title,
        "value": pv,
        "status_label": status_label,
        "status_raw": status,
        "source_company": company,
        "reporting_company": reporting_company,
        "quarter": quarter,
        "area": area,
    }


def _accumulate_firm_totals(totals: dict[str, Any], status_label: str, value_eur: int, quarter: str | None) -> None:
    totals["sum"] += value_eur
    totals["count"] += 1
    if status_label == "bestellt":
        totals["bestellt"] += value_eur
        return
    if status_label == "im Durchlauf":
        totals["durchlauf"] += value_eur
        return
    if status_label == "Konzept":
        totals["konzept"] += value_eur
        if quarter in QUARTERS:
            totals["konzept_quarters"][quarter] += value_eur
        return
    if status_label == "storniert":
        totals["storniert"] += value_eur


def _firm_period_value_eur(totals: dict[str, Any], num_q: int) -> int:
    base_value = int(totals["bestellt"]) + int(totals["durchlauf"])
    if num_q <= 2:
        return base_value
    if num_q == 3:
        return base_value + int(totals["konzept_quarters"].get("Q3", 0))
    return base_value + int(totals["konzept_quarters"].get("Q4", 0))


def _load_reporting_company_targets(year: int) -> tuple[dict[str, int], dict[str, dict[str, int]]]:
    if not os.path.isfile(DB_PATH):
        return {}, {}
    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                """
                SELECT company, quarter, SUM(target_value) AS target_value
                FROM plan_company_targets
                WHERE year = ?
                GROUP BY company, quarter
                """,
                (year,),
            ).fetchall()
    except sqlite3.OperationalError:
        return {}, {}

    annual_targets: dict[str, int] = defaultdict(int)
    quarter_targets: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for row in rows:
        quarter = str(row["quarter"] or "").strip().upper()
        if quarter not in QUARTERS:
            continue
        reporting_company = _reporting_company(str(row["company"] or ""))
        target_value = int(row["target_value"] or 0)
        annual_targets[reporting_company] += target_value
        quarter_targets[reporting_company][quarter] += target_value
    return dict(annual_targets), {firm: dict(values) for firm, values in quarter_targets.items()}


def _cumulative_company_targets(
    targets_by_quarter: dict[str, dict[str, int]],
    num_q: int,
) -> dict[str, int]:
    quarters = QUARTERS[:num_q]
    return {
        firm: sum(values.get(quarter, 0) for quarter in quarters)
        for firm, values in targets_by_quarter.items()
    }


_STATUS_PRIORITY = {"bestellt": 1, "im Durchlauf": 2, "Konzept": 3, "storniert": 4}
_STATUS_FILL_MAP = {
    "Konzept": PatternFill("solid", fgColor="BDD7EE"),
    "im Durchlauf": PatternFill("solid", fgColor="FFC000"),
    "bestellt": PatternFill("solid", fgColor="63BE7B"),
    "storniert": PatternFill("solid", fgColor="FFC7CE"),
}

_SUBGROUP_AREAS: dict[str, str] = {
    THIESEN_SPEC_KEY: "Spez",
    THIESEN_SUPPORT_KEY: "Support",
}
_CATCH_ALL_HINTS = ("WEITERE", "SONSTIGE", "OTHER")
_MARKER_OVERRIDES: dict[str, str] = {"VWGS": "VOLKSWAGEN GROUP SERVICES"}
_MATRIX_START_COL = 10


def _matrix_marker_for_label(label: str) -> str:
    marker = "" if any(h in label.upper() for h in _CATCH_ALL_HINTS) else label.upper()
    parts = marker.split()
    if len(parts) > 1:
        marker = parts[0]
    return _MARKER_OVERRIDES.get(marker, marker)


def _detect_matrix_template_groups() -> dict[str, tuple[int, str]]:
    """Scan row 1 of the EA-Matrix template to keep template order and known aliases."""
    if not os.path.isfile(EA_MATRIX_TEMPLATE):
        return {}
    wb = load_workbook(EA_MATRIX_TEMPLATE, read_only=True, data_only=True)
    ws = wb.active
    groups: dict[str, tuple[int, str]] = {}
    for c in range(_MATRIX_START_COL, (ws.max_column or 45) + 1, len(QUARTERS)):
        header = ws.cell(1, c).value
        if not header or not str(header).strip():
            continue
        key = str(header).strip()
        groups[key] = (c, _matrix_marker_for_label(key))
    wb.close()
    return groups


EA_MATRIX_CATEGORIES = ("Sondervereinbarungen", "Serienbetreuung", "Fahrzeugprojekte")


def _find_catch_all(groups: dict[str, tuple[int, str]]) -> str | None:
    """Return the key of the catch-all group (empty marker)."""
    return next((k for k, (_c, m) in groups.items() if not m), None)


def _company_matrix_label(company: str) -> str:
    upper = company.upper()
    for alias, fallback in COMPANY_ALIAS_FALLBACKS.items():
        if alias in upper or fallback.upper() in upper:
            return "VWGS" if alias == "VOLKSWAGEN" else alias
    return company.strip()


def _company_matrix_labels(company: str) -> list[str]:
    if "THIESEN" in company.upper():
        return ["THIESEN Spez.", "THIESEN Support"]
    label = _company_matrix_label(company)
    return [label] if label else ["Weitere"]


def _matrix_groups_from_labels(labels: list[str]) -> dict[str, tuple[int, str]]:
    template_groups = _detect_matrix_template_groups()
    template_order = {label: idx for idx, label in enumerate(template_groups)}

    def sort_key(label: str) -> tuple[int, str]:
        if label in template_order:
            return template_order[label], label
        if any(h in label.upper() for h in _CATCH_ALL_HINTS):
            return 9999, label
        return 1000, label

    ordered = sorted(dict.fromkeys(labels), key=sort_key)
    return {
        label: (_MATRIX_START_COL + idx * len(QUARTERS), template_groups.get(label, (0, _matrix_marker_for_label(label)))[1])
        for idx, label in enumerate(ordered)
    }


def _build_matrix_groups(
    entries: list[dict[str, Any]],
    quarter_targets: dict[str, dict[str, int]],
) -> dict[str, tuple[int, str]]:
    labels: list[str] = []
    for entry in entries:
        if entry.get("quarter") not in QUARTERS:
            continue
        company = str(entry.get("reporting_company") or entry.get("source_company") or "").strip()
        labels.extend(_company_matrix_labels(company))
    for company, values in quarter_targets.items():
        if any(int(value or 0) for value in values.values()):
            labels.extend(_company_matrix_labels(str(company)))
    return _matrix_groups_from_labels(labels)


def _subgroup_target_map(groups: dict[str, tuple[int, str]]) -> dict[str, str]:
    return {
        group: area_key
        for area_key, hint in _SUBGROUP_AREAS.items()
        for group in groups
        if hint.upper() in group.upper()
    }


def _matrix_group_for_entry(entry: dict[str, Any], groups: dict[str, tuple[int, str]]) -> str:
    company_text = str(entry.get("reporting_company") or entry.get("source_company") or "")
    company = company_text.upper()
    area = str(entry.get("area") or "")
    matches = [k for k, (_c, m) in groups.items() if m and m in company]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        for area_key, hint in _SUBGROUP_AREAS.items():
            if area == area_key:
                return next((k for k in matches if hint.upper() in k.upper()), matches[0])
        return next((k for k in matches if not any(h.upper() in k.upper() for h in _SUBGROUP_AREAS.values())), matches[0])
    label = _company_matrix_label(company_text)
    if label in groups:
        return label
    return _find_catch_all(groups) or (list(groups)[-1] if groups else "Weitere")


def _parse_iso_date(value: Any) -> datetime.date | Any:
    if isinstance(value, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", value.strip()):
        return datetime.date.fromisoformat(value.strip())
    return value


def _load_ea_metadata() -> dict[str, dict[str, Any]]:
    if not os.path.isfile(DB_PATH):
        return {}
    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                """
                SELECT ea_number, title, project_family, sop, hierarchy
                FROM devorder
                WHERE COALESCE(ea_number, '') <> ''
                """
            ).fetchall()
    except sqlite3.OperationalError:
        return {}
    return {
        _normalize_ea_key(row["ea_number"]): {
            "ea": str(row["ea_number"] or ""),
            "title": str(row["title"] or ""),
            "family": str(row["project_family"] or ""),
            "sop": _parse_iso_date(row["sop"]),
            "hierarchy": str(row["hierarchy"] or ""),
        }
        for row in rows
        if _normalize_ea_key(row["ea_number"])
    }


def _ea_matrix_category(ea_key: str, meta: dict[str, Any], special_order: dict[str, int]) -> str:
    if ea_key in special_order:
        return "Sondervereinbarungen"
    title = str(meta.get("title") or "").upper()
    hierarchy = str(meta.get("hierarchy") or "").upper()
    if "SERIENBETREUUNG" in hierarchy or title.startswith("SB "):
        return "Serienbetreuung"
    return "Fahrzeugprojekte"


def _ea_matrix_sort_key(row: dict[str, Any]) -> tuple[int, int, str, str]:
    return (
        EA_MATRIX_CATEGORIES.index(row["category"]),
        int(row.get("special_order", 9999)),
        str(row.get("family") or ""),
        str(row.get("ea") or ""),
    )


def _target_quarter_values_for_area(targets: dict[str, int], start_q: dict[str, int], area: str) -> dict[str, int]:
    target_eur = int(targets.get(area, 0)) * 1000
    first_q = int(start_q.get(area, 1))
    active = [q for idx, q in enumerate(QUARTERS, start=1) if idx >= first_q]
    if not active:
        return {q: 0 for q in QUARTERS}
    base = round(target_eur / len(active))
    values = {q: (base if q in active else 0) for q in QUARTERS}
    values[active[-1]] += target_eur - sum(values.values())
    return values


def _load_ek_totals(year: int) -> dict[str, tuple[int, int]]:
    """EK Plan / EK Genehmigt per EA from btl_all (all OEs), in EUR."""
    if not os.path.isfile(DB_PATH):
        return {}
    try:
        with sqlite3.connect(DB_PATH) as conn:
            rows = conn.execute(
                "SELECT dev_order,"
                " SUM(planned_value) AS ek_plan,"
                " SUM(CASE WHEN status LIKE '%Bestellt%' THEN planned_value ELSE 0 END) AS ek_gen"
                " FROM btl_all"
                " WHERE substr(COALESCE(target_date,''),1,4)=? AND COALESCE(dev_order,'') <> ''"
                " GROUP BY dev_order",
                (str(year),),
            ).fetchall()
    except sqlite3.OperationalError:
        return {}
    return {_normalize_ea_key(r[0]): (int(r[1] or 0), int(r[2] or 0)) for r in rows}


def _build_ea_matrix(
    *,
    year: int,
    entries: list[dict[str, Any]],
    targets: dict[str, int],
    start_q: dict[str, int],
    known_companies: list[str],
) -> dict[str, Any]:
    metadata = _load_ea_metadata()
    special_rules = _load_sondervorgaben(CONFIG_XLSX, known_companies)
    special_order = {
        key: idx
        for idx, rule in enumerate(special_rules)
        for key in rule.get("ea_keys", [])
        if key
    }
    el_aggregates = _load_el_aggregates(year, 4)
    ek_totals = _load_ek_totals(year)
    _annual_targets, quarter_targets = _load_reporting_company_targets(year)
    groups = _build_matrix_groups(entries, quarter_targets)
    ea_totals: dict[str, dict[str, Any]] = defaultdict(lambda: {"quarters": defaultdict(int), "total": 0, "status": {}})

    for entry in entries:
        ea_key = _normalize_ea_key(entry.get("ea_number") or entry.get("ea"))
        quarter = entry.get("quarter")
        if not ea_key or quarter not in QUARTERS:
            continue
        group = _matrix_group_for_entry(entry, groups)
        value = int(entry.get("value") or 0)
        ea_totals[ea_key]["quarters"][(group, quarter)] += value
        key = (group, quarter)
        prev = ea_totals[ea_key]["status"].get(key)
        label = entry.get("status_label", "")
        if prev is None or _STATUS_PRIORITY.get(label, 99) < _STATUS_PRIORITY.get(prev, 99):
            ea_totals[ea_key]["status"][key] = label
        ea_totals[ea_key]["total"] += value
        if ea_key not in metadata:
            metadata[ea_key] = {
                "ea": ea_key,
                "title": entry.get("ea") or entry.get("title") or ea_key,
                "family": "",
                "sop": None,
                "hierarchy": "",
            }

    for ea_key in special_order:
        metadata.setdefault(ea_key, {"ea": ea_key, "title": ea_key, "family": "", "sop": None, "hierarchy": ""})

    rows: list[dict[str, Any]] = []
    grand_total = sum(int(data["total"]) for data in ea_totals.values())
    ek_plan_total = sum(v[0] for v in ek_totals.values()) or 1
    ek_gen_total = sum(v[1] for v in ek_totals.values()) or 1
    for ea_key in sorted(set(metadata) & (set(ea_totals) | set(special_order))):
        meta = metadata[ea_key]
        total = int(ea_totals[ea_key]["total"])
        rows.append(
            {
                "category": _ea_matrix_category(ea_key, meta, special_order),
                "special_order": special_order.get(ea_key, 9999),
                "ea": ea_key,
                "title": meta.get("title") or ea_key,
                "family": meta.get("family") or "",
                "sop": meta.get("sop"),
                "fl_te": round(total / 1000),
                "fl_pct": (total / grand_total) if grand_total else 0,
                "ek_plan": ek_totals[ea_key][0] / ek_plan_total if ea_key in ek_totals else None,
                "ek_genehmigt": ek_totals[ea_key][1] / ek_gen_total if ea_key in ek_totals else None,
                "el_te": round(float(el_aggregates.get(ea_key, {}).get("year_te", 0))) if ea_key in el_aggregates else None,
                "quarters": dict(ea_totals[ea_key]["quarters"]),
                "statuses": dict(ea_totals[ea_key]["status"]),
            }
        )
    rows.sort(key=_ea_matrix_sort_key)

    soll = {group: {q: 0 for q in QUARTERS} for group in groups}
    subgroup_targets = _subgroup_target_map(groups)
    for firm, values in quarter_targets.items():
        group = _matrix_group_for_entry({"reporting_company": firm}, groups)
        if group in subgroup_targets:
            continue
        for quarter in QUARTERS:
            soll[group][quarter] += int(values.get(quarter, 0))
    for group, area_key in subgroup_targets.items():
        soll[group] = _target_quarter_values_for_area(targets, start_q, area_key)

    ist = {group: {q: 0 for q in QUARTERS} for group in groups}
    for entry in entries:
        quarter = entry.get("quarter")
        if quarter in QUARTERS:
            grp = _matrix_group_for_entry(entry, groups)
            ist[grp][quarter] += int(entry.get("value") or 0)

    return {"groups": groups, "rows": rows, "soll": soll, "ist": ist}


def _copy_cell_style(cell) -> dict[str, Any]:
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "number_format": cell.number_format,
        "protection": copy(cell.protection),
    }


def _apply_cell_style(cell, style: dict[str, Any]) -> None:
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.number_format = style["number_format"]
    cell.protection = copy(style["protection"])


def _copy_row_style(ws, row_idx: int) -> tuple[float | None, list[dict[str, Any]]]:
    return (
        ws.row_dimensions[row_idx].height,
        [_copy_cell_style(ws.cell(row_idx, col_idx)) for col_idx in range(1, ws.max_column + 1)],
    )


def _apply_row_style(ws, row_idx: int, template: tuple[float | None, list[dict[str, Any]]], max_col: int | None = None) -> None:
    height, styles = template
    ws.row_dimensions[row_idx].height = height
    for col_idx, style in enumerate(styles[:max_col], start=1):
        _apply_cell_style(ws.cell(row_idx, col_idx), style)


def _copy_matrix_block_style(ws) -> dict[str, Any]:
    return {
        "widths": [
            ws.column_dimensions[get_column_letter(_MATRIX_START_COL + offset)].width
            for offset in range(len(QUARTERS))
        ],
        "rows": {
            row_idx: [
                _copy_cell_style(ws.cell(row_idx, _MATRIX_START_COL + offset))
                for offset in range(len(QUARTERS))
            ]
            for row_idx in range(1, 5)
        },
    }


def _reset_matrix_columns(ws, groups: dict[str, tuple[int, str]], block_style: dict[str, Any]) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_col >= _MATRIX_START_COL:
            ws.unmerge_cells(str(merged_range))
    if ws.max_column >= _MATRIX_START_COL:
        ws.delete_cols(_MATRIX_START_COL, ws.max_column - _MATRIX_START_COL + 1)

    for label, (start_col, _marker) in groups.items():
        for offset, quarter in enumerate(QUARTERS):
            col_idx = start_col + offset
            width = block_style["widths"][offset]
            if width is not None:
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            for row_idx, row_styles in block_style["rows"].items():
                _apply_cell_style(ws.cell(row_idx, col_idx), row_styles[offset])
            ws.cell(4, col_idx).value = quarter
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + len(QUARTERS) - 1)
        ws.cell(1, start_col).value = label


def _write_ea_matrix(ws, matrix: dict[str, Any] | None) -> None:
    if not matrix:
        return
    groups = matrix.get("groups") or {}
    max_matrix_col = _MATRIX_START_COL - 1 + len(groups) * len(QUARTERS)
    block_style = _copy_matrix_block_style(ws)
    data_style = _copy_row_style(ws, 5)
    category_style = _copy_row_style(ws, 7)
    if ws.max_row >= 5:
        ws.delete_rows(5, ws.max_row - 4)
    _reset_matrix_columns(ws, groups, block_style)
    ws.freeze_panes = "C5"

    ws["A2"] = f"Stand: {datetime.date.today().strftime('%d.%m.%Y')}"
    for group, (start_col, _marker) in groups.items():
        for offset, quarter in enumerate(QUARTERS):
            col_idx = start_col + offset
            ws.cell(2, col_idx).value = round(matrix["soll"].get(group, {}).get(quarter, 0) / 1000)
            ws.cell(3, col_idx).value = round(matrix["ist"].get(group, {}).get(quarter, 0) / 1000)

    current_row = 5
    for category in EA_MATRIX_CATEGORIES:
        category_rows = [row for row in matrix["rows"] if row["category"] == category]
        if not category_rows:
            continue
        _apply_row_style(ws, current_row, category_style, max_matrix_col)
        ws.cell(current_row, 2).value = category
        current_row += 1
        no_fill = PatternFill(fill_type=None)
        for data in category_rows:
            _apply_row_style(ws, current_row, data_style, max_matrix_col)
            values = [
                data["ea"],
                data["title"],
                data["family"],
                data["sop"],
                data.get("ek_plan"),
                data.get("ek_genehmigt"),
                data["fl_te"] or None,
                data["fl_pct"] or None,
                data["el_te"],
            ]
            for col_idx, value in enumerate(values, start=1):
                ws.cell(current_row, col_idx).value = value
            ws.cell(current_row, 5).number_format = "0.0%"
            ws.cell(current_row, 6).number_format = "0.0%"
            for col_idx in range(5, max_matrix_col + 1):
                ws.cell(current_row, col_idx).fill = no_fill
            for group, (start_col, _marker) in groups.items():
                for offset, quarter in enumerate(QUARTERS):
                    value = round(data["quarters"].get((group, quarter), 0) / 1000)
                    cell = ws.cell(current_row, start_col + offset)
                    cell.value = value or None
                    status = data.get("statuses", {}).get((group, quarter))
                    if value and status in _STATUS_FILL_MAP:
                        cell.fill = _STATUS_FILL_MAP[status]
            current_row += 1

    last = current_row - 1
    if last >= 5:
        pct_rule = ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF", end_type="num", end_value=0.05, end_color="00B050")
        num_rule = ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF", end_type="num", end_value=200, end_color="00B050")
        for col in ("E", "F", "H"):
            ws.conditional_formatting.add(f"{col}5:{col}{last}", pct_rule)
        ws.conditional_formatting.add(f"G5:G{last}", num_rule)


def _merge_quarter_todo_entries(
    base: dict[str, list[dict[str, Any]]],
    extra: dict[str, list[dict[str, Any]]],
) -> dict[str, list[dict[str, Any]]]:
    merged: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for quarter, entries in base.items():
        merged[quarter].extend(entries)
    for quarter, entries in extra.items():
        merged[quarter].extend(entries)
    return merged


def _build_quarter_delta_todos(
    actual_entries: list[dict[str, Any]],
    target_entries: list[dict[str, Any]],
    *,
    current_period_quarter: int,
) -> dict[str, list[dict[str, Any]]]:
    active_statuses = {"Konzept", "im Durchlauf", "bestellt"}
    target_totals: dict[tuple[str, str, str], int] = defaultdict(int)
    for entry in target_entries:
        if entry.get("quarter") not in QUARTERS:
            continue
        if entry.get("status_label") not in active_statuses:
            continue
        key = (
            str(entry.get("source_company") or ""),
            str(entry.get("quarter") or ""),
            str(entry.get("ea_number") or ""),
        )
        target_totals[key] += int(entry.get("value") or 0)

    actual_by_key: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    actual_totals: dict[tuple[str, str, str], int] = defaultdict(int)
    for entry in actual_entries:
        quarter = entry.get("quarter")
        if quarter not in QUARTERS or _quarter_index(str(quarter)) < current_period_quarter:
            continue
        if entry.get("status_label") not in active_statuses:
            continue
        key = (
            str(entry.get("source_company") or ""),
            str(quarter),
            str(entry.get("ea_number") or ""),
        )
        actual_by_key[key].append(entry)
        actual_totals[key] += int(entry.get("value") or 0)

    todo_entries: dict[str, list[dict[str, Any]]] = defaultdict(list)
    status_priority = {"Konzept": 0, "im Durchlauf": 1, "bestellt": 2}
    for key, current_total in actual_totals.items():
        target_total = target_totals.get(key, 0)
        if current_total <= target_total:
            continue
        remaining = current_total - target_total
        quarter = key[1]
        candidates = sorted(
            actual_by_key[key],
            key=lambda entry: (
                status_priority.get(str(entry.get("status_label") or ""), 9),
                -int(entry.get("value") or 0),
                str(entry.get("concept") or ""),
            ),
        )
        for entry in candidates:
            if remaining <= 0:
                break
            value = int(entry.get("value") or 0)
            reduce_value = min(value, remaining)
            if reduce_value <= 0:
                continue
            remaining_after = value - reduce_value
            is_delete = remaining_after <= 0
            todo_entries[quarter].append(
                {
                    **entry,
                    "value_ist": value,
                    "value_neu": remaining_after,
                    "sort_value": value,
                    "todo_type": "Löschen" if is_delete else "Reduzieren",
                    "action": "Abruf löschen / zurückziehen"
                    if is_delete
                    else f"Abruf auf {fmt(round(remaining_after / 1000))} reduzieren",
                }
            )
            remaining -= reduce_value
    return todo_entries


def _build_quarter_todo_sections(
    todo_bms_by_quarter: dict[str, list[dict[str, Any]]],
    *,
    current_period_quarter: int,
) -> list[TextSection | TableSection]:
    sections: list[TextSection | TableSection] = []
    for quarter in QUARTERS:
        entries = todo_bms_by_quarter.get(quarter, [])
        if not entries:
            continue
        quarter_num = _quarter_index(quarter)
        is_current = quarter_num == current_period_quarter
        if is_current:
            intro_title = f"{quarter}-Todo: Maßnahmen jetzt umsetzen"
            intro_lines = [
                "Diese Liste zeigt alle Maßnahmen, um den realen Ist-Stand auf das Optimierungssoll des aktuellen Quartals zu bringen.",
                f"Neue Abrufe stehen dafür bewusst auf dem frühen Durchlauf-Status **{CURRENT_QUARTER_TODO_STATUS}**.",
                "Aktion: neue Abrufe starten sowie bestehende Positionen bei Bedarf löschen oder reduzieren.",
            ]
        else:
            intro_title = f"{quarter}-Todo: künftige Maßnahmen vorbereiten"
            intro_lines = [
                f"Diese Liste zeigt alle Maßnahmen, um den realen Ist-Stand rechtzeitig auf das Optimierungssoll für {quarter} zu bringen.",
                "Neue Abrufe stehen dafür bewusst noch auf **01_In Erstellung**.",
                "Aktion: neue Abrufe vorbereiten sowie bestehende Positionen bei Bedarf löschen oder reduzieren.",
            ]

        sections.append(
            TextSection(
                title=intro_title,
                lines=intro_lines,
                sheet_name="Korrektur",
            )
        )
        todo_rows = [
            TableRow(
                [
                    entry["reporting_company"].split()[0] if str(entry["reporting_company"]).strip() else str(entry["reporting_company"]),
                    entry["todo_type"],
                    entry["concept"],
                    entry["ea_number"],
                    entry["ea"],
                    entry["title"],
                    "" if entry.get("value_ist") in (None, "") else fmt(round(int(entry["value_ist"]) / 1000)),
                    "" if entry.get("value_neu") in (None, "") else fmt(round(int(entry["value_neu"]) / 1000)),
                    entry["status_raw"],
                    entry["action"],
                ]
            )
            for entry in sorted(
                entries,
                key=lambda item: (
                    str(item["reporting_company"]),
                    {"Löschen": 0, "Reduzieren": 1, "Neu": 2}.get(str(item.get("todo_type") or ""), 9),
                    -int(item.get("sort_value", item.get("value_neu", 0)) or 0),
                    str(item["concept"]),
                ),
            )
        ]
        sections.append(
            TableSection(
                title=f"{quarter}-Todo-Liste",
                headers=["Firma", "Typ", "Konzept", "EA-Nr.", "EA", "BM-Titel", "Wert ist", "Wert neu", "Status", "Aktion"],
                rows=todo_rows,
                align_right={6, 7},
                level=3,
                separator_before=False,
                sheet_name="Korrektur",
                body_row_height=15,
            )
        )
    return sections


def _thiesen_manual_soll(company: str, targets: dict[str, int], start_q: dict[str, int], num_q: int) -> tuple[int, int] | None:
    if "THIESEN" not in company.upper():
        return None

    areas = [THIESEN_SUPPORT_KEY, THIESEN_SPEC_KEY]
    soll = 0
    soll_q = 0
    for area in areas:
        target_te = targets.get(area, 0)
        area_soll = target_te * 1000
        soll += area_soll
        sq = start_q.get(area, 1)
        if num_q < sq:
            a_q_ratio = 0.0
        else:
            a_q_ratio = (num_q - sq + 1) / (4 - sq + 1)
        soll_q += round(area_soll * a_q_ratio)
    return soll, soll_q


def _manual_company_soll(
    company: str,
    company_target_overrides: dict[str, dict[str, int | str]],
    num_q: int,
) -> tuple[str, int, int] | None:
    upper = company.upper()
    for company_key, config in company_target_overrides.items():
        if company_key not in upper:
            continue
        target_te = int(config["target_te"])
        start_q = int(config["start_q"])
        soll = target_te * 1000
        if num_q < start_q:
            q_ratio = 0.0
        else:
            q_ratio = (num_q - start_q + 1) / (4 - start_q + 1)
        soll_q = round(soll * q_ratio)
        return str(config["area"]), soll, soll_q
    return None


def _render_table_markdown(section: TableSection) -> list[str]:
    lines = [f"| {' | '.join(section.headers)} |"]
    separator = [
        "---:" if idx in section.align_right else "---"
        for idx in range(len(section.headers))
    ]
    lines.append(f"|{'|'.join(separator)}|")
    for row in section.rows:
        values = [_markdown_table_cell(value, row.style) for value in row.values]
        if (
            section.sheet_name == "Korrektur"
            and row.style == "body"
            and section.headers
            and "Konzept" in section.headers
            and values
        ):
            concept_idx = section.headers.index("Konzept")
            url = _bplus_vorgang_url(values[concept_idx])
            if url:
                values[concept_idx] = f"[{_strip_markdown(values[concept_idx])}]({url})"
        lines.append(f"| {' | '.join(values)} |")
    return lines


def render_markdown(report: ReportDocument) -> str:
    lines: list[str] = [f"# {report.title}", ""]
    lines.extend(_render_meta_line_markdown(line) for line in report.meta_lines)

    for section in report.sections:
        lines.append("")
        if section.separator_before:
            lines.extend(["---", ""])
        if isinstance(section, TextSection):
            if section.title:
                lines.extend([_heading(section.level, section.title), ""])
            lines.extend(section.lines)
        else:
            lines.extend([_heading(section.level, section.title), ""])
            lines.extend(section.intro_lines)
            if section.intro_lines:
                lines.append("")
            lines.extend(_render_table_markdown(section))
    lines.append("")
    return "\n".join(lines)


def write_xlsx_report(report: ReportDocument, xlsx_file: str, inherit_notes_from: str | None = None) -> str:
    workbook = load_workbook(EA_MATRIX_TEMPLATE) if os.path.isfile(EA_MATRIX_TEMPLATE) else Workbook()
    if "Übersicht" in workbook.sheetnames:
        del workbook["Übersicht"]
    if "Korrektur" in workbook.sheetnames:
        del workbook["Korrektur"]
    overview_ws = workbook.create_sheet("Übersicht", 0)
    correction_ws = workbook.create_sheet("Korrektur", 1)
    if "Sheet" in workbook.sheetnames and "EA-Matrix" not in workbook.sheetnames:
        del workbook["Sheet"]

    max_cols = max(1, report.max_columns)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_fill = PatternFill("solid", fgColor="EAF2F8")
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    summary_fill = PatternFill("solid", fgColor="F3F3F3")
    status_fill_bestellt = PatternFill("solid", fgColor="63BE7B")
    status_fill_storniert = PatternFill("solid", fgColor="D9D9D9")
    status_fill_abgelehnt = PatternFill("solid", fgColor="FFC7CE")
    status_fill_erstellung = PatternFill("solid", fgColor="8EA9DB")
    status_fill_durchlauf = PatternFill("solid", fgColor="FFC000")
    status_fill_ok = PatternFill("solid", fgColor="FFC6EFCE")
    status_fill_warn = PatternFill("solid", fgColor="FFFFEB9C")
    status_fill_x = PatternFill("solid", fgColor="FFFFC7CE")
    status_fill_na = PatternFill("solid", fgColor="FFD9D9D9")
    wrap = Alignment(vertical="top", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right = Alignment(horizontal="right", vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    def render_sheet(
        ws,
        sections: list[TextSection | TableSection],
        *,
        include_title: bool,
        include_meta: bool,
        empty_message: str | None = None,
        first_col_width: int = 32,
        first_col_min: int = 14,
        first_col_max: int = 42,
        fixed_widths: dict[int, float] | None = None,
    ) -> list[tuple[int, int, int]]:
        current_row = 1
        col_widths = [12] * (max_cols + 1)
        col_widths[1] = first_col_width
        table_ranges: list[tuple[int, int, int]] = []

        def estimate_lines(text: str, chars_per_line: int) -> int:
            clean = _strip_markdown(text)
            if not clean:
                return 1
            total = 0
            for part in clean.splitlines() or [""]:
                total += max(1, (len(part) + max(chars_per_line - 1, 1) - 1) // max(chars_per_line, 1))
            return max(total, 1)

        def set_row_height(row_idx: int, *, texts: list[str], chars_per_line: list[int], min_height: float = 15.0, line_height: float = 15.0) -> None:
            estimated = max(
                estimate_lines(text, chars)
                for text, chars in zip(texts, chars_per_line, strict=False)
            )
            ws.row_dimensions[row_idx].height = max(min_height, estimated * line_height)

        def merge_row(
            text: str,
            *,
            font: Font,
            fill: PatternFill | None = None,
            min_height: float | None = None,
            line_height: float | None = None,
        ) -> None:
            nonlocal current_row
            row_idx = current_row
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
            cell = ws.cell(current_row, 1, _strip_markdown(text))
            cell.font = font
            cell.alignment = wrap
            if fill is not None:
                cell.fill = fill
            set_row_height(
                row_idx,
                texts=[text],
                chars_per_line=[120],
                min_height=min_height or max(font.sz + 8 if font.sz else 18, 18),
                line_height=line_height or max((font.sz or 11) + 3, 15),
            )
            current_row += 1

        def update_width(col_idx: int, value: str) -> None:
            text = _strip_markdown(str(value))
            if col_idx == 1:
                width = min(max(len(text) + 2, first_col_min), first_col_max)
            else:
                width = min(max(len(text) + 2, 10), 32)
            col_widths[col_idx] = max(col_widths[col_idx], width)

        def effective_chars_per_line(col_idx: int) -> int:
            width = fixed_widths.get(col_idx, col_widths[col_idx]) if fixed_widths else col_widths[col_idx]
            return max(10, int(width) - 2)

        def status_fill_for(value: str) -> PatternFill | None:
            text = _strip_markdown(str(value)).upper()
            if not text:
                return None
            if "BESTELLT" in text:
                return status_fill_bestellt
            if "STORNIERT" in text:
                return status_fill_storniert
            if "ABGELEHNT" in text:
                return status_fill_abgelehnt
            if "ERSTELLUNG" in text:
                return status_fill_erstellung
            if (
                "IM DURCHLAUF" in text
                or
                "PLANEN-BM" in text
                or "FREIGABE KOSTENSTELLE" in text
                or "BEARBEITUNG BM-TEAM" in text
                or "IN DEN EINKAUF ÜBERTRAGEN" in text
            ):
                return status_fill_durchlauf
            return None

        def header_fill_for(value: str) -> PatternFill | None:
            return status_fill_for(value)

        def special_status_fill(value: str) -> PatternFill | None:
            text = _strip_markdown(str(value)).upper()
            if text in {"OK", "IO"}:
                return status_fill_ok
            if text == "WARN":
                return status_fill_warn
            if text in {"X", "NIO"}:
                return status_fill_x
            if text == "-":
                return status_fill_na
            return None

        if include_title:
            merge_row(report.title, font=Font(size=16, bold=True), min_height=30, line_height=18)
        if include_meta:
            for line in report.meta_lines:
                merge_row(_display_meta_line(line), font=_meta_line_font(line), min_height=20, line_height=15)
        if empty_message and not sections:
            merge_row(empty_message, font=Font(size=10, italic=True), min_height=20, line_height=15)

        for section in sections:
            if section.separator_before and current_row > 1:
                current_row += 1

            if isinstance(section, TextSection):
                if section.title:
                    merge_row(
                        section.title,
                        font=Font(size=13 if section.level == 2 else 11, bold=True),
                        fill=section_fill,
                        min_height=26 if section.level == 2 else 24,
                        line_height=16,
                    )
                for line in section.lines:
                    clean = _strip_markdown(line)
                    if clean:
                        heading_level = _markdown_heading_level(line)
                        if heading_level is not None:
                            merge_row(
                                clean,
                                font=Font(size=13 if heading_level <= 2 else 11, bold=True),
                                min_height=24 if heading_level <= 2 else 22,
                                line_height=16,
                            )
                        else:
                            merge_row(clean, font=Font(size=10), min_height=18, line_height=15)
                    else:
                        current_row += 1
                continue

            merge_row(
                section.title,
                font=Font(size=13 if section.level == 2 else 11, bold=True),
                fill=section_fill,
                min_height=26 if section.level == 2 else 24,
                line_height=16,
            )
            for line in section.intro_lines:
                clean = _strip_markdown(line)
                if clean:
                    merge_row(clean, font=Font(size=10), min_height=18, line_height=15)
                else:
                    current_row += 1

            header_row = current_row
            for col_idx, header in enumerate(section.headers, start=1):
                cell = ws.cell(current_row, col_idx, header)
                cell.font = Font(bold=True)
                custom_header_fill = header_fill_for(header) if ws.title == "Übersicht" else None
                cell.fill = custom_header_fill or header_fill
                cell.border = border
                cell.alignment = center if header in SPECIAL_STATUS_HEADERS else left
                update_width(col_idx, header)
            set_row_height(
                current_row,
                texts=section.headers,
                chars_per_line=[effective_chars_per_line(idx) for idx in range(1, len(section.headers) + 1)],
                min_height=20,
                line_height=15,
            )
            current_row += 1

            body_indexes = [idx for idx, row in enumerate(section.rows) if row.style == "body"]
            if ws.title == "Korrektur" and body_indexes:
                first_body = min(body_indexes)
                last_body = max(body_indexes)
                contiguous = body_indexes == list(range(first_body, last_body + 1))
                last_table_row = header_row + (last_body + 1 if contiguous else len(section.rows))
                table_ranges.append((header_row, last_table_row, len(section.headers)))

            is_sonder = section.title == "Sondervereinbarungen"

            for ridx, row in enumerate(section.rows):
                row_idx = current_row
                if is_sonder:
                    is_first_of_block = not (row.values and row.values[0] == "")
                    next_is_cont = (
                        ridx + 1 < len(section.rows)
                        and section.rows[ridx + 1].values
                        and section.rows[ridx + 1].values[0] == ""
                    )
                    row_border = Border(
                        left=thin,
                        right=thin,
                        top=thin if is_first_of_block else None,
                        bottom=thin if not next_is_cont else None,
                    )
                else:
                    row_border = border

                for col_idx, value in enumerate(row.values, start=1):
                    te_number = _parse_te_number(value)
                    cell_value = te_number if te_number is not None else _strip_markdown(value)
                    cell = ws.cell(current_row, col_idx, cell_value)
                    cell.border = row_border
                    header = section.headers[col_idx - 1] if col_idx - 1 < len(section.headers) else ""
                    cell_status = row.cell_statuses.get(header)
                    if header in SPECIAL_STATUS_HEADERS:
                        cell.alignment = center
                    else:
                        cell.alignment = right if (col_idx - 1) in section.align_right else left
                    if ws.title == "Korrektur" and row.style == "body" and header == "Konzept":
                        url = _bplus_vorgang_url(str(value))
                        if url:
                            cell.hyperlink = url
                            cell.font = Font(color="0563C1", underline="single")
                    if te_number is not None:
                        cell.number_format = _te_number_format(te_number)
                    if row.style == "summary":
                        cell.font = Font(bold=True)
                        cell.fill = summary_fill
                    elif row.style == "note":
                        cell.font = Font(italic=True)
                    elif cell_status is not None:
                        fill = special_status_fill(cell_status)
                        if fill is not None:
                            cell.fill = fill
                    elif header in SPECIAL_STATUS_HEADERS:
                        fill = special_status_fill(str(cell_value))
                        if fill is not None:
                            cell.fill = fill
                    elif ws.title == "Korrektur" and header == "Status":
                        fill = status_fill_for(str(cell_value))
                        if fill is not None:
                            cell.fill = fill
                    update_width(col_idx, value)
                set_row_height(
                    row_idx,
                    texts=row.values,
                    chars_per_line=[effective_chars_per_line(idx) for idx in range(1, len(row.values) + 1)],
                    min_height=section.body_row_height if section.body_row_height is not None else (18 if row.style == "summary" else 15),
                    line_height=0 if section.body_row_height is not None else 15,
                )
                current_row += 1

        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        for col_idx in range(1, max_cols + 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = fixed_widths.get(col_idx, col_widths[col_idx]) if fixed_widths else col_widths[col_idx]
        return table_ranges

    overview_sections = [section for section in report.sections if section.sheet_name == "Übersicht"]
    correction_sections = [section for section in report.sections if section.sheet_name == "Korrektur"]

    render_sheet(
        overview_ws,
        overview_sections,
        include_title=True,
        include_meta=True,
        first_col_width=34,
        first_col_min=18,
        first_col_max=44,
        fixed_widths={1: 44, 12: 65},
    )
    correction_table_ranges = render_sheet(
        correction_ws,
        correction_sections,
        include_title=False,
        include_meta=False,
        empty_message="Keine Korrekturabschnitte vorhanden.",
        first_col_width=16,
        first_col_min=10,
        first_col_max=20,
        fixed_widths={5: 48},
    )
    table_style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    for idx, (start_row, end_row, width) in enumerate(correction_table_ranges, start=1):
        ref = f"A{start_row}:{get_column_letter(width)}{end_row}"
        tab = Table(displayName=f"Korrektur_{idx}", ref=ref)
        tab.tableStyleInfo = table_style
        correction_ws.add_table(tab)
    _embed_target_image(overview_ws, TARGET_IMAGE)

    _, target_firm_rows = _collect_table_key_rows(overview_ws, "Firma")
    target_firm_header = _find_table_header_row(overview_ws, "Firma")
    if target_firm_header is not None:
        notes_header = overview_ws.cell(target_firm_header, 13)
        if not _cell_has_text(notes_header.value):
            template_header = overview_ws.cell(target_firm_header, 12)
            _copy_cell_display(template_header, notes_header, copy_value=False)
            notes_header.value = "Notizen"
        overview_ws.column_dimensions["M"].width = max(overview_ws.column_dimensions["M"].width or 0, 65)
        for row_idx in target_firm_rows.values():
            note_cell = overview_ws.cell(row_idx, 13)
            if not note_cell.has_style:
                template_cell = overview_ws.cell(row_idx, 12)
                _copy_cell_display(template_cell, note_cell, copy_value=False)
                note_cell.value = None
                note_cell.font = copy(template_cell.font)
                note_cell.fill = copy(template_cell.fill)
                note_cell.border = copy(template_cell.border)
                note_cell.alignment = copy(template_cell.alignment)

    _inherit_notes_from_workbook(workbook, inherit_notes_from)
    _write_overview_quarter_block(overview_ws, report.overview_quarter_block)
    if "EA-Matrix" in workbook.sheetnames:
        _write_ea_matrix(workbook["EA-Matrix"], report.ea_matrix)
    workbook.active = workbook.sheetnames.index("Übersicht")
    for ws in workbook.worksheets:
        ws.views.sheetView[0].tabSelected = (ws.title == "Übersicht")
    workbook.save(xlsx_file)
    return xlsx_file


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, default=datetime.date.today().year)
    parser.add_argument("--source-table", choices=sorted(SOURCE_TABLES), default="btl")
    parser.add_argument("--pdf", action="store_true", help="Erzeugt zusaetzlich eine PDF-Datei neben Markdown und XLSX.")
    parser.add_argument("--inherit-notes-from")
    return parser


def generate_report(
    *,
    year: int,
    source_table: str = "btl",
    pdf: bool = False,
    inherit_notes_from: str | None = None,
    open_excel: bool = False,
) -> dict[str, str]:
    source_table = _normalize_source_table(source_table)

    # ── Vorgaben + Prämissen laden ─────────────────────────────────────
    REF_2025, TARGET_2026, AREA_ORDER, START_Q, COMPANY_TARGET_OVERRIDES, QUARTER_SPLIT = load_targets()
    praemissen_text = load_praemissen()
    AUDI_FIXED = TARGET_2026.get(AUDI_KEY, 0)
    quarter_split = QUARTER_SPLIT if QUARTER_SPLIT and int(QUARTER_SPLIT.get("year", 0)) == year else None

    # ── Quartal-Periode berechnen ──────────────────────────────────────
    # Ab 1 Monat nach Quartalsanfang → nächstes Quartal einbeziehen
    month = datetime.date.today().month
    if month >= 8:
        num_q = 4
    elif month >= 5:
        num_q = 3
    elif month >= 2:
        num_q = 2
    else:
        num_q = 1
    q_label = f"Q1-{num_q}" if num_q > 1 else "Q1"

    # ── BMs laden (inkl. Status) ───────────────────────────────────────
    sql = _budget_source_sql(source_table)
    rows = run_sql(sql, year)
    ea_title_by_number = _load_ea_title_by_number(year)

    # ── Status-Mapping aus CSV laden ───────────────────────────────────
    status_map = load_status_mapping()

    # ── Sync-Zeitpunkt ermitteln ───────────────────────────────────────
    schema_cmd = [sys.executable, BUDGET_DB, "schema"]
    schema_out = subprocess.run(schema_cmd, capture_output=True, text=True, cwd=REPO_ROOT).stdout
    sync_match = re.search(r"btl\s+\|\s+\d+\s+\|\s+([\d\-T:]+)", schema_out)
    sync_time = sync_match.group(1) if sync_match else "unbekannt"

    # ── BMs klassifizieren ─────────────────────────────────────────────
    area_values: dict[str, int] = {a: 0 for a in AREA_ORDER}
    area_values["_SONSTIGE"] = 0
    firm_values: dict[str, dict[str, Any]] = {}
    firm_bms: dict[str, list[dict]] = {}  # firm → [{concept, ea, title, value, status_label}, ...]
    quarter_todo_bms: dict[str, list[dict[str, Any]]] = defaultdict(list)
    budget_entries: list[dict[str, Any]] = []
    # Für Soll-Berechnung: Firma→Area→Ist-Anteil
    firm_area_ist: dict[str, dict[str, int]] = {}
    actual_quarter_eur = {quarter: 0 for quarter in QUARTERS}

    # Pre-load btl keys so OPT-rows replacing existing orders are not shown as "Neu"
    _actual_btl_keys: set[tuple[str, str, str]] = set()
    _actual_entries_cache: list[dict[str, Any]] | None = None
    if _normalize_source_table(source_table) == "btl_opt":
        _actual_entries_cache = [
            e for e in (_budget_entry_from_row(r, status_map, ea_title_by_number) for r in run_sql(_budget_source_sql("btl"), year))
            if e is not None
        ]
        _actual_btl_keys = {
            (str(e["source_company"]), str(e["quarter"]), str(e["ea_number"]))
            for e in _actual_entries_cache
        }

    for row in rows:
        entry = _budget_entry_from_row(row, status_map, ea_title_by_number)
        if entry is None:
            continue
        budget_entries.append(entry)
        pv = int(entry["value"])
        company = str(entry["source_company"])
        ea = str(entry["ea"])
        title = str(entry["title"])
        status = str(entry["status_raw"])
        benennung = str(entry["status_label"])
        quarter = entry["quarter"]
        if quarter in actual_quarter_eur:
            actual_quarter_eur[quarter] += pv
        area = str(entry["area"])
        reporting_company = str(entry["reporting_company"])
        area_values[area] = area_values.get(area, 0) + pv

        # Firmen-Aggregation
        if reporting_company not in firm_values:
            firm_values[reporting_company] = _new_firm_totals()
        _accumulate_firm_totals(firm_values[reporting_company], benennung, pv, quarter)

        # BM-Einzeldaten für Korrektur-Abschnitt sammeln
        if benennung in ("Konzept", "im Durchlauf", "bestellt", "storniert"):
            if reporting_company not in firm_bms:
                firm_bms[reporting_company] = []
            bm_entry = {
                "concept": entry["concept"],
                "ea_number": entry["ea_number"],
                "ea": ea,
                "title": title,
                "value": pv,
                "status_label": benennung,
                "status_raw": status,
                "source_company": company,
            }
            firm_bms[reporting_company].append(bm_entry)
            if _normalize_source_table(source_table) == "btl_opt" and quarter in QUARTERS:
                include_current = quarter == f"Q{num_q}" and status == CURRENT_QUARTER_TODO_STATUS
                include_future = _quarter_index(quarter) > num_q and benennung == "Konzept" and (company, quarter, str(entry["ea_number"])) not in _actual_btl_keys
                if include_current or include_future:
                    quarter_todo_bms[quarter].append(
                        {
                            **entry,
                            "value_ist": "",
                            "value_neu": pv,
                            "sort_value": pv,
                            "todo_type": "Neu",
                            "action": "Abruf erstellen / in Durchlauf schicken"
                            if include_current
                            else f"Abruf vorbereiten / für {quarter} vormerken",
                        }
                    )

        # Firma→Area Zuordnung für Soll-Verteilung
        if area != "_SONSTIGE":
            if reporting_company not in firm_area_ist:
                firm_area_ist[reporting_company] = {}
            firm_area_ist[reporting_company][area] = firm_area_ist[reporting_company].get(area, 0) + pv

    if _normalize_source_table(source_table) == "btl_opt":
        actual_entries = _actual_entries_cache if _actual_entries_cache is not None else [
            entry
            for entry in (_budget_entry_from_row(row, status_map, ea_title_by_number) for row in run_sql(_budget_source_sql("btl"), year))
            if entry is not None
        ]
        quarter_todo_bms = _merge_quarter_todo_entries(
            quarter_todo_bms,
            _build_quarter_delta_todos(
                actual_entries,
                budget_entries,
                current_period_quarter=num_q,
            ),
        )

    # ── AUDI-Korrektur: von Systemschaltplänen abziehen ────────────────
    sysschalt_key = "Systemschaltpläne und Bibl. (EDAG, Bertrandt)"
    area_values[sysschalt_key] -= AUDI_FIXED * 1000
    area_values[AUDI_KEY] = AUDI_FIXED * 1000

    # ── In T€ umrechnen ───────────────────────────────────────────────
    area_te = {a: round(v / 1000) for a, v in area_values.items() if a != "_SONSTIGE"}

    # ── Soll pro Firma berechnen (proportional aus Area-Targets) ───────
    # Für jede Area: Firma-Anteil am Ist → gleicher Anteil am Target
    area_totals: dict[str, int] = {}
    for firm, areas in firm_area_ist.items():
        for area, val in areas.items():
            area_totals[area] = area_totals.get(area, 0) + val

    # AUDI-Target dem Systemschaltpläne-Target zuschlagen (AUDI-Ist steckt in EDAG-BMs)
    soll_targets: dict[str, int] = {a: v for a, v in TARGET_2026.items()}
    if AUDI_FIXED and AUDI_KEY in soll_targets:
        soll_targets[sysschalt_key] = soll_targets.get(sysschalt_key, 0) + AUDI_FIXED
        soll_targets.pop(AUDI_KEY, None)

    manual_company_soll: dict[str, tuple[str, int, int]] = {}
    override_target_by_area: dict[str, int] = {}
    area_totals_remaining: dict[str, int] = {}
    for firm, areas in firm_area_ist.items():
        manual_override = _manual_company_soll(firm, COMPANY_TARGET_OVERRIDES, num_q)
        if manual_override is not None:
            area_name, firm_soll_abs, firm_soll_q_abs = manual_override
            manual_company_soll[firm] = manual_override
            override_target_by_area[area_name] = override_target_by_area.get(area_name, 0) + firm_soll_abs
            continue
        for area, val in areas.items():
            area_totals_remaining[area] = area_totals_remaining.get(area, 0) + val

    firm_soll: dict[str, int] = {}
    firm_soll_q: dict[str, int] = {}
    for firm, areas in firm_area_ist.items():
        manual_override = manual_company_soll.get(firm)
        if manual_override is not None:
            _, firm_soll[firm], firm_soll_q[firm] = manual_override
            continue

        thiesen_override = _thiesen_manual_soll(firm, soll_targets, START_Q, num_q)
        if thiesen_override is not None:
            firm_soll[firm], firm_soll_q[firm] = thiesen_override
            continue

        soll = 0
        soll_q = 0
        for area, val in areas.items():
            total = area_totals_remaining.get(area, area_totals.get(area, 1))
            target = soll_targets.get(area, 0) * 1000 - override_target_by_area.get(area, 0)
            target = max(target, 0)
            area_soll = round(target * val / total) if total > 0 else 0
            soll += area_soll
            # Area-spezifische Q-Ratio (Start_Q berücksichtigen)
            sq = START_Q.get(area, 1)
            if num_q < sq:
                a_q_ratio = 0.0
            else:
                a_q_ratio = (num_q - sq + 1) / (4 - sq + 1)
            soll_q += round(area_soll * a_q_ratio)
        firm_soll[firm] = soll
        firm_soll_q[firm] = soll_q

    # --- Phase A2: COMPANY_AREA_MAP fallback for areas without BMs ---
    # The proportional loop above only distributes targets for areas that
    # already have BMs in firm_area_ist.  If a firm is mapped to additional
    # areas in COMPANY_AREA_MAP that have no BMs yet (e.g. RuleChecker before
    # Q3), those area-targets are lost.  Add them here as sole-firm targets.
    if COMPANY_AREA_MAP:
        for map_company, map_areas in COMPANY_AREA_MAP.items():
            if map_company in manual_company_soll:
                continue
            existing_areas = set(firm_area_ist.get(map_company, {}).keys())
            missing_areas = [a for a in map_areas if a not in existing_areas]
            if not missing_areas:
                continue
            extra_soll = 0
            extra_soll_q = 0
            for area in missing_areas:
                target = soll_targets.get(area, 0) * 1000 - override_target_by_area.get(area, 0)
                target = max(target, 0)
                extra_soll += target
                sq = START_Q.get(area, 1)
                if num_q < sq:
                    a_q_ratio = 0.0
                else:
                    a_q_ratio = (num_q - sq + 1) / (4 - sq + 1)
                extra_soll_q += round(target * a_q_ratio)
            firm_soll[map_company] = firm_soll.get(map_company, 0) + extra_soll
            firm_soll_q[map_company] = firm_soll_q.get(map_company, 0) + extra_soll_q

    firm_soll_q = _scale_distribution_to_target(firm_soll_q, _quarter_split_period_target_eur(quarter_split, num_q))
    _plan_firm_soll, plan_firm_targets_by_quarter = _load_reporting_company_targets(year)
    if plan_firm_targets_by_quarter:
        firm_soll_q.update(_cumulative_company_targets(plan_firm_targets_by_quarter, num_q))

    report_title = f"EKEK/1 Budget {year} — Target / Ist-Analyse"
    meta_lines = [
        f"**Stand BPLUS-NG:** {sync_time} | **Erstellt:** {datetime.date.today().strftime('%d.%m.%Y')}"
    ]
    source_notice = _source_notice_line(source_table)
    if source_notice:
        meta_lines.append(source_notice)
    sections: list[TextSection | TableSection] = []

    # ── Gesamtübersicht (oben) ────────────────────────────────────────
    sum_25 = sum_t = sum_i = 0
    for area in AREA_ORDER:
        sum_25 += REF_2025[area]
        sum_t += TARGET_2026[area]
        sum_i += area_te.get(area, 0)

    sections.append(
        TableSection(
            title="Gesamtübersicht",
            headers=["", "Wert"],
            rows=[
                TableRow(["IST 2025 (Referenz)", fmt(sum_25)]),
                TableRow(["Ist (BPLUS)", fmt(sum_i)]),
                TableRow(["Target", fmt(sum_t)]),
                TableRow(["Delta Ist vs. Target", delta_fmt(sum_i - sum_t)]),
            ],
            align_right={1},
            separator_before=False,
        )
    )
    overview_quarter_block = None
    if quarter_split:
        target_quarter_te = {quarter: float(quarter_split["quarters_te"][quarter]) for quarter in QUARTERS}
        annual_target_te = float(quarter_split["annual_te"])
        ref_quarter_te = {
            quarter: (sum_25 * target_quarter_te[quarter] / annual_target_te) if annual_target_te else 0.0
            for quarter in QUARTERS
        }
        actual_quarter_te = {quarter: actual_quarter_eur[quarter] / 1000 for quarter in QUARTERS}
        delta_quarter_te = {
            quarter: actual_quarter_te[quarter] - target_quarter_te[quarter]
            for quarter in QUARTERS
        }
        overview_quarter_block = {
            "headers": list(QUARTERS),
            "rows": {
                "IST 2025 (Referenz)": [_round_te_display(ref_quarter_te[quarter]) for quarter in QUARTERS],
                "Ist (BPLUS)": [_round_te_display(actual_quarter_te[quarter]) for quarter in QUARTERS],
                "Target": [_round_te_display(target_quarter_te[quarter]) for quarter in QUARTERS],
                "Delta Ist vs. Target": [_round_te_display(delta_quarter_te[quarter]) for quarter in QUARTERS],
            },
        }

    # ── Tabelle 1: Target - Ist - Vergleich ───────────────────────────
    area_rows: list[TableRow] = []
    s25 = s_t = s_i = 0
    for area in AREA_ORDER:
        if area == AUDI_KEY:
            continue
        v25 = REF_2025[area]
        vt = TARGET_2026[area]
        vi = area_te.get(area, 0)
        delta = vi - vt
        area_rows.append(TableRow([area, fmt(v25), fmt(vt), fmt(vi), delta_fmt(delta), ""]))
        s25 += v25
        s_t += vt
        s_i += vi

    area_rows.append(TableRow(["Summe", fmt(s25), fmt(s_t), fmt(s_i), delta_fmt(s_i - s_t), ""], style="summary"))

    v25_a = REF_2025[AUDI_KEY]
    vt_a = TARGET_2026[AUDI_KEY]
    vi_a = area_te.get(AUDI_KEY, 0)
    area_rows.append(TableRow([AUDI_KEY, fmt(v25_a), fmt(vt_a), fmt(vi_a), delta_fmt(vi_a - vt_a), ""]))
    s25 += v25_a
    s_t += vt_a
    s_i += vi_a
    area_rows.append(
        TableRow(
            ["Summe inkl. Audi", fmt(s25), fmt(s_t), fmt(s_i), delta_fmt(s_i - s_t), ""],
            style="summary",
        )
    )
    sections.append(
        TableSection(
            title="Target - Ist - Vergleich",
            headers=["Aufgabenbereich", "2025", "Target", "Ist", "Delta", "Maßnahmen"],
            rows=area_rows,
            align_right={1, 2, 3, 4},
            body_row_height=15,
        )
    )

    # ── Tabelle 2: Firmen ─────────────────────────────────────────────
    firm_rows: list[TableRow] = []
    firm_total = 0
    firm_total_soll = 0
    firm_total_bestellt = 0
    firm_total_durchlauf = 0
    firm_total_konzept = 0
    firm_total_storniert = 0
    firm_count_total = 0
    for firm, data in sorted(firm_values.items(), key=lambda x: -x[1]["sum"]):
        if _hide_from_firm_overview(firm):
            continue

        ist = round(data["sum"] / 1000)
        soll = round(firm_soll.get(firm, 0) / 1000)
        bestellt = round(data["bestellt"] / 1000)
        durchlauf = round(data["durchlauf"] / 1000)
        konzept = round(data["konzept"] / 1000)
        storniert = round(data["storniert"] / 1000)
        soll_q = round(firm_soll_q.get(firm, 0) / 1000)
        diff_planung = ist - soll
        period_ist = round(_firm_period_value_eur(data, num_q) / 1000)
        diff_q = soll_q - period_ist
        firm_short = firm.split()[0] if firm.strip() else firm

        massnahmen_parts: list[str] = []
        if diff_q < 0:
            massnahmen_parts.append(f"Quartal überplant: {fmt(abs(diff_q))} zurückholen")
        elif diff_q > 0:
            massnahmen_parts.append(f"Quartalswert zu gering: {fmt(diff_q)} beauftragen")
        if diff_planung > 0:
            if storniert > 0:
                massnahmen_parts.append(f"Stornierte Positionen zuerst löschen: {fmt(min(diff_planung, storniert))}")
            massnahmen_parts.append(f"Jahr überplant: {fmt(diff_planung)} streichen")
        elif diff_planung < 0:
            massnahmen_parts.append(f"Jahreswert zu gering: {fmt(abs(diff_planung))} einplanen")
        massnahmen = "; ".join(massnahmen_parts)

        year_band = _diff_band_status(ist, soll) if soll else "-"
        q_band = _diff_band_status(period_ist, soll_q) if soll_q else "-"

        firm_rows.append(
            TableRow(
                [
                    firm_short,
                    str(data["count"]),
                    fmt(soll),
                    fmt(ist),
                    delta_fmt(diff_planung),
                    fmt(soll_q),
                    fmt(bestellt),
                    fmt(durchlauf),
                    fmt(konzept),
                    fmt(storniert),
                    delta_fmt(diff_q),
                    massnahmen,
                ],
                cell_statuses={
                    "DIFF Ges.": year_band,
                    f"DIFF {q_label}": q_band,
                },
            )
        )
        firm_total += ist
        firm_total_soll += soll
        firm_total_bestellt += bestellt
        firm_total_durchlauf += durchlauf
        firm_total_konzept += konzept
        firm_total_storniert += storniert
        firm_count_total += data["count"]

    firm_total_soll_q = sum(
        round(firm_soll_q.get(f, 0) / 1000)
        for f in firm_values
        if not _hide_from_firm_overview(f)
    )
    firm_total_period_ist = sum(
        round(_firm_period_value_eur(data, num_q) / 1000)
        for firm, data in firm_values.items()
        if not _hide_from_firm_overview(firm)
    )
    firm_total_diff_q = firm_total_soll_q - firm_total_period_ist
    firm_total_diff_planung = firm_total - firm_total_soll
    firm_rows.append(
        TableRow(
            [
                "Gesamt",
                str(firm_count_total),
                fmt(firm_total_soll),
                fmt(firm_total),
                delta_fmt(firm_total_diff_planung),
                fmt(firm_total_soll_q),
                fmt(firm_total_bestellt),
                fmt(firm_total_durchlauf),
                fmt(firm_total_konzept),
                fmt(firm_total_storniert),
                delta_fmt(firm_total_diff_q),
                "",
            ],
            style="summary",
        )
    )
    sections.append(
        TableSection(
            title="Firmen-Übersicht",
            headers=[
                "Firma",
                "BMs",
                "Soll",
                "Ist",
                "DIFF Ges.",
                f"Soll {q_label}",
                "bestellt",
                "im Durchlauf",
                "01 Erstellung",
                "storniert",
                f"DIFF {q_label}",
                "Maßnahmen",
            ],
            rows=firm_rows,
            align_right={1, 2, 3, 4, 5, 6, 7, 8, 9, 10},
            body_row_height=15,
        )
    )

    special_section = build_special_rule_section(
        year=year,
        num_q=num_q,
        q_label=q_label,
        rows=rows,
        status_map=status_map,
    )
    if special_section is not None:
        sections.append(special_section)

    sections.extend(
        _build_quarter_todo_sections(
            quarter_todo_bms,
            current_period_quarter=num_q,
        )
    )

    # ── Tabelle 3: Korrektur Überplanung ──────────────────────────────
    korrektur_firmen: list[dict] = []
    for firm, data in sorted(firm_values.items(), key=lambda x: -x[1]["sum"]):
        ist = round(data["sum"] / 1000)
        soll = round(firm_soll.get(firm, 0) / 1000)
        bestellt_te = round(data["bestellt"] / 1000)
        durchlauf_te = round(data["durchlauf"] / 1000)
        soll_q_te = round(firm_soll_q.get(firm, 0) / 1000)
        diff_jahr = ist - soll
        diff_q = soll_q_te - round(_firm_period_value_eur(data, num_q) / 1000)
        if diff_jahr > 0 or diff_q < 0:
            korrektur_firmen.append({"firm": firm, "diff_jahr": diff_jahr, "diff_q": diff_q})

    if korrektur_firmen:
        sections.append(
            TextSection(
                title="Korrektur Überplanung",
                lines=[
                    "Überplante Lieferanten mit einzelnen Vorgängen zum Rückzug oder zur Reduzierung.",
                    "Stornierte Vorgänge werden pro Firma zuerst aufgeführt und in der **Aktion-Spalte** mit _löschen_ vorbelegt.",
                    "Alle übrigen Einträge in der **Aktion-Spalte** manuell befüllen (z.B. _zurückziehen_, _reduzieren auf X T€_, _verschieben Q3_).",
                ],
                sheet_name="Korrektur",
            )
        )

        korrektur_firmen.sort(key=lambda x: -(x["diff_jahr"] + abs(min(x["diff_q"], 0))))
        korrektur_headers = ["Firma", "Typ", "Konzept", "EA-Nr.", "EA", "BM-Titel", "Wert ist", "Wert neu", "Status", "Aktion"]

        for kf in korrektur_firmen:
            firm = kf["firm"]
            firm_short = firm.split()[0] if firm.strip() else firm
            bms = firm_bms.get(firm, [])
            diff_jahr = kf["diff_jahr"]
            diff_q = kf["diff_q"]
            storno_bms = sorted([b for b in bms if b["status_label"] == "storniert"], key=lambda x: -x["value"])

            if storno_bms:
                rows_storno: list[TableRow] = []
                sum_storno = 0
                for b in storno_bms:
                    v = round(b["value"] / 1000)
                    rows_storno.append(
                        TableRow([firm_short, "Löschen", b["concept"], b["ea_number"], b["ea"], b["title"], fmt(v), fmt(0), b["status_raw"], "löschen"])
                    )
                    sum_storno += v
                rows_storno.append(TableRow(["", "", "", "", "", "Summe storniert", fmt(sum_storno), fmt(0), "", ""], style="summary"))

                sections.append(
                    TableSection(
                        title=f"{firm_short} — Stornierte Vorgänge: löschen",
                        headers=korrektur_headers,
                        rows=rows_storno,
                        align_right={6, 7},
                        level=3,
                        separator_before=False,
                        sheet_name="Korrektur",
                        body_row_height=15,
                    )
                )

            if diff_q < 0:
                reduce_q = abs(diff_q)
                rows_q: list[TableRow] = []
                prio1 = sorted([b for b in bms if b["status_label"] == "im Durchlauf"], key=lambda x: -x["value"])
                sum_p1 = 0
                for b in prio1:
                    v = round(b["value"] / 1000)
                    rows_q.append(TableRow([firm_short, "Reduzieren", b["concept"], b["ea_number"], b["ea"], b["title"], fmt(v), "", b["status_raw"], ""]))
                    sum_p1 += v
                if prio1:
                    rows_q.append(TableRow(["", "", "", "", "", "Summe im Durchlauf", fmt(sum_p1), "", "", ""], style="summary"))

                if sum_p1 < reduce_q:
                    prio2 = sorted([b for b in bms if b["status_label"] == "bestellt"], key=lambda x: -x["value"])
                    sum_p2 = 0
                    for b in prio2:
                        v = round(b["value"] / 1000)
                        rows_q.append(TableRow([firm_short, "Reduzieren", b["concept"], b["ea_number"], b["ea"], b["title"], fmt(v), "", b["status_raw"], ""]))
                        sum_p2 += v
                    if prio2:
                        rows_q.append(TableRow(["", "", "", "", "", "Summe bestellt", fmt(sum_p2), "", "", ""], style="summary"))

                if not prio1 and not [b for b in bms if b["status_label"] == "bestellt"]:
                    rows_q.append(TableRow(["", "", "", "", "", "Keine rückziehbaren Vorgänge vorhanden", "", "", "", ""], style="note"))

                sections.append(
                    TableSection(
                        title=f"{firm_short} — Quartals-Korrektur ({q_label}): {fmt(reduce_q)} zu reduzieren",
                        headers=korrektur_headers,
                        rows=rows_q,
                        align_right={6, 7},
                        level=3,
                        separator_before=False,
                        sheet_name="Korrektur",
                        body_row_height=15,
                    )
                )

            if diff_jahr > 0:
                rows_jahr: list[TableRow] = []
                prio1 = sorted([b for b in bms if b["status_label"] == "Konzept"], key=lambda x: -x["value"])
                sum_p1 = 0
                for b in prio1:
                    v = round(b["value"] / 1000)
                    rows_jahr.append(TableRow([firm_short, "Reduzieren", b["concept"], b["ea_number"], b["ea"], b["title"], fmt(v), "", b["status_raw"], ""]))
                    sum_p1 += v
                if prio1:
                    rows_jahr.append(TableRow(["", "", "", "", "", "Summe 01 Erstellung", fmt(sum_p1), "", "", ""], style="summary"))

                if sum_p1 < diff_jahr:
                    prio2 = sorted([b for b in bms if b["status_label"] == "im Durchlauf"], key=lambda x: -x["value"])
                    sum_p2 = 0
                    for b in prio2:
                        v = round(b["value"] / 1000)
                        rows_jahr.append(TableRow([firm_short, "Reduzieren", b["concept"], b["ea_number"], b["ea"], b["title"], fmt(v), "", b["status_raw"], ""]))
                        sum_p2 += v
                    if prio2:
                        rows_jahr.append(TableRow(["", "", "", "", "", "Summe im Durchlauf", fmt(sum_p2), "", "", ""], style="summary"))

                if not prio1 and not [b for b in bms if b["status_label"] == "im Durchlauf"]:
                    rows_jahr.append(TableRow(["", "", "", "", "", "Keine rückziehbaren Vorgänge vorhanden", "", "", "", ""], style="note"))

                sections.append(
                    TableSection(
                        title=f"{firm_short} — Jahres-Korrektur: {fmt(diff_jahr)} zu reduzieren",
                        headers=korrektur_headers,
                        rows=rows_jahr,
                        align_right={6, 7},
                        level=3,
                        separator_before=False,
                        sheet_name="Korrektur",
                        body_row_height=15,
                    )
                )

    sections.append(
        TextSection(
            lines=praemissen_text.splitlines() + [f"- 2026-Ist: BPLUS-NG Sync vom {sync_time}"],
        )
    )

    max_columns = 1
    for section in sections:
        if isinstance(section, TableSection):
            max_columns = max(max_columns, len(section.headers), *(len(row.values) for row in section.rows or [TableRow([""])]))

    report = ReportDocument(
        title=report_title,
        meta_lines=meta_lines,
        sections=sections,
        max_columns=max_columns,
        overview_quarter_block=overview_quarter_block,
        ea_matrix=_build_ea_matrix(
            year=year,
            entries=budget_entries,
            targets=TARGET_2026,
            start_q=START_Q,
            known_companies=sorted({str(entry["reporting_company"]) for entry in budget_entries}),
        ),
    )

    md_text = render_markdown(report)

    # ── Datei schreiben ───────────────────────────────────────────────
    os.makedirs(OUT_DIR, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = os.path.join(OUT_DIR, f"{ts}_budget_massnahmenplan_ekek1.md")
    with open(out_file, "w", encoding="utf-8") as f:
        f.write(md_text + "\n")

    xlsx_file = os.path.join(OUT_DIR, f"{ts}_budget_massnahmenplan_ekek1.xlsx")
    inherit_notes_from = inherit_notes_from or _find_previous_xlsx(OUT_DIR, xlsx_file)
    write_xlsx_report(report, xlsx_file, inherit_notes_from=inherit_notes_from)

    pdf_file = write_pdf_beside_markdown(out_file) if pdf else None

    result = {
        "markdown": os.path.relpath(out_file, REPO_ROOT),
        "xlsx": os.path.relpath(xlsx_file, REPO_ROOT),
    }
    if pdf_file:
        result["pdf"] = os.path.relpath(pdf_file, REPO_ROOT)
    if open_excel:
        open_excel_file(xlsx_file)
    return result


def main():
    args = build_arg_parser().parse_args()
    result = generate_report(
        year=args.year,
        source_table=args.source_table,
        pdf=args.pdf,
        inherit_notes_from=args.inherit_notes_from,
        open_excel=True,
    )
    print(result["markdown"])
    print(result["xlsx"])
    if "pdf" in result:
        print(result["pdf"])


if __name__ == "__main__":
    main()
