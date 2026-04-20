#!/usr/bin/env python3
"""Excel I/O CLI — read, write, edit .xlsx files."""
from __future__ import annotations

import argparse
import csv
import io
import json
import sys
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import range_boundaries


BORDER_STYLES = {"thin", "medium", "thick", "none"}
FONT_KEYS = {"bold", "italic", "size", "color"}
ALIGN_KEYS = {"align", "valign", "wrap"}


def autocast(v):
    if v is None or isinstance(v, (int, float, bool)):
        return v
    s = str(v)
    if s.startswith("="):
        return s
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


def cells_in_range(ws, spec: str):
    if ":" in spec:
        min_col, min_row, max_col, max_row = range_boundaries(spec)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
            for c in row:
                yield c
    else:
        yield ws[spec]


def normalize_style(s: dict) -> dict:
    out = dict(s)
    if "bg" in out:
        out["fill"] = out.pop("bg")
    if "number-format" in out:
        out["number_format"] = out.pop("number-format")
    if "border-color" in out:
        out["border_color"] = out.pop("border-color")
    font = out.pop("font", None)
    if isinstance(font, dict):
        out.update({k: v for k, v in font.items() if k in FONT_KEYS})
    return out


def apply_style(cell, style: dict) -> None:
    if any(k in style for k in FONT_KEYS):
        f = copy(cell.font)
        cell.font = Font(
            name=f.name,
            size=style.get("size", f.size),
            bold=style.get("bold", f.bold),
            italic=style.get("italic", f.italic),
            color=style.get("color", f.color),
        )
    if style.get("fill"):
        cell.fill = PatternFill("solid", fgColor=style["fill"])
    if any(k in style for k in ALIGN_KEYS):
        a = copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal=style.get("align", a.horizontal),
            vertical=style.get("valign", a.vertical),
            wrap_text=style.get("wrap", a.wrap_text),
        )
    if "border" in style:
        b = style["border"]
        if b == "none":
            cell.border = Border()
        elif b in BORDER_STYLES:
            side = Side(style=b, color=style.get("border_color", "000000"))
            cell.border = Border(left=side, right=side, top=side, bottom=side)
    if style.get("number_format"):
        cell.number_format = style["number_format"]


def flags_to_style(args) -> dict:
    s = {}
    if args.style_json:
        s.update(normalize_style(json.loads(args.style_json)))
    if args.bold:
        s["bold"] = True
    if args.italic:
        s["italic"] = True
    if args.font_size is not None:
        s["size"] = args.font_size
    if args.color:
        s["color"] = args.color
    if args.bg:
        s["fill"] = args.bg
    if args.align:
        s["align"] = args.align
    if args.valign:
        s["valign"] = args.valign
    if args.wrap:
        s["wrap"] = True
    if args.border:
        s["border"] = args.border
    if args.border_color:
        s["border_color"] = args.border_color
    if args.number_format:
        s["number_format"] = args.number_format
    return s


def rows_to_md(rows):
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    rows = [list(r) + [None] * (width - len(r)) for r in rows]
    header = rows[0]
    lines = [
        "| " + " | ".join("" if v is None else str(v) for v in header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for r in rows[1:]:
        lines.append("| " + " | ".join("" if v is None else str(v) for v in r) + " |")
    return "\n".join(lines)


def read_rows(ws, cell_range=None):
    if cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        return [list(row) for row in ws.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col, values_only=True)]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def cmd_info(args):
    wb = load_workbook(args.file, read_only=True, data_only=True)
    lines = [f"**{args.file}**", ""]
    for name in wb.sheetnames:
        ws = wb[name]
        lines.append(f"- `{name}` — {ws.max_row} rows × {ws.max_column} cols")
    print("\n".join(lines))
    wb.close()


def cmd_read(args):
    wb = load_workbook(args.file, read_only=True, data_only=True)
    sheets = wb.sheetnames if args.sheet == "all" else [args.sheet or wb.sheetnames[0]]
    chunks = []
    multi = args.sheet == "all"
    for name in sheets:
        rows = read_rows(wb[name], args.range)
        if args.format == "md":
            chunks.append((f"## {name}\n\n" if multi else "") + rows_to_md(rows))
        elif args.format == "json":
            chunks.append(json.dumps({name: rows} if multi else rows, ensure_ascii=False))
        else:
            buf = io.StringIO()
            csv.writer(buf).writerows(rows)
            chunks.append((f"# {name}\n" if multi else "") + buf.getvalue().rstrip("\n"))
    print("\n\n".join(chunks))
    wb.close()


def cmd_edit(args):
    wb = load_workbook(args.file)
    ws = wb[args.sheet] if args.sheet else wb.active
    changed = 0
    if args.batch:
        ops = json.loads(Path(args.batch).read_text(encoding="utf-8"))
        for op in ops:
            style = normalize_style(op.get("style") or {})
            for cell in cells_in_range(ws, op["cell"]):
                if op.get("value") is not None:
                    cell.value = autocast(op["value"])
                if style:
                    apply_style(cell, style)
                changed += 1
    else:
        if not args.cell:
            sys.exit("error: --cell required (or use --batch)")
        style = flags_to_style(args)
        for cell in cells_in_range(ws, args.cell):
            if args.value is not None:
                cell.value = autocast(args.value)
            if style:
                apply_style(cell, style)
            changed += 1
    out = args.output or args.file
    wb.save(out)
    print(f"OK: {changed} cell(s) updated -> {out}")


def cmd_write(args):
    src = Path(args.src)
    ext = src.suffix.lower()
    if ext == ".csv":
        with src.open(encoding="utf-8", newline="") as f:
            rows = list(csv.reader(f))
    elif ext == ".json":
        data = json.loads(src.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            key = args.sheet or next(iter(data))
            data = data[key]
        if data and isinstance(data[0], dict):
            keys = list(data[0].keys())
            rows = [keys] + [[r.get(k) for k in keys] for r in data]
        else:
            rows = data
    else:
        sys.exit(f"error: unsupported input {ext}")
    file = Path(args.file)
    if file.exists():
        wb = load_workbook(file)
        name = args.sheet or wb.active.title
        if args.append and name in wb.sheetnames:
            ws = wb[name]
        else:
            if name in wb.sheetnames:
                del wb[name]
            ws = wb.create_sheet(name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = args.sheet or "Sheet1"
    for row in rows:
        ws.append([autocast(v) for v in row])
    out = args.output or str(file)
    wb.save(out)
    print(f"OK: {len(rows)} row(s) written -> {out}")


def add_style_flags(p: argparse.ArgumentParser) -> None:
    p.add_argument("--bold", action="store_true")
    p.add_argument("--italic", action="store_true")
    p.add_argument("--font-size", type=int, dest="font_size")
    p.add_argument("--color", help="hex RRGGBB")
    p.add_argument("--bg", help="hex RRGGBB fill")
    p.add_argument("--align", choices=["left", "center", "right"])
    p.add_argument("--valign", choices=["top", "center", "bottom"])
    p.add_argument("--wrap", action="store_true")
    p.add_argument("--border", choices=["thin", "medium", "thick", "none"])
    p.add_argument("--border-color", dest="border_color", help="hex RRGGBB")
    p.add_argument("--number-format", dest="number_format")
    p.add_argument("--style-json", dest="style_json", help="JSON dict (merged with flags)")


def main():
    ap = argparse.ArgumentParser(description="Excel I/O — read, write, edit .xlsx")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("info")
    p.add_argument("file")
    p.set_defaults(func=cmd_info)

    p = sub.add_parser("read")
    p.add_argument("file")
    p.add_argument("--sheet", help="sheet name or 'all'")
    p.add_argument("--range", help="e.g. A1:D10")
    p.add_argument("--as", dest="format", choices=["md", "json", "csv"], default="md")
    p.set_defaults(func=cmd_read)

    p = sub.add_parser("edit")
    p.add_argument("file")
    p.add_argument("--sheet")
    p.add_argument("--cell", help="A1 or A1:C3")
    p.add_argument("--value")
    p.add_argument("--batch", help="ops.json")
    p.add_argument("--output", help="write to new file")
    add_style_flags(p)
    p.set_defaults(func=cmd_edit)

    p = sub.add_parser("write")
    p.add_argument("file")
    p.add_argument("--from", dest="src", required=True)
    p.add_argument("--sheet")
    p.add_argument("--append", action="store_true")
    p.add_argument("--output")
    p.set_defaults(func=cmd_write)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
