"""Check / add EA entries in the EA-Blacklist sheet of beauftragungsplanung_config.xlsx."""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, Side

CONFIG_DEFAULT = Path(__file__).resolve().parents[2] / "userdata" / "budget" / "planning" / "beauftragungsplanung_config.xlsx"
SHEET = "EA-Blacklist"
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _norm(val: str) -> str:
    digits = "".join(ch for ch in val if ch.isdigit())
    return (digits.lstrip("0") or "0") if digits else ""


def _read(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows: list[dict[str, str]] = []
    for r in range(2, (ws.max_row or 1) + 1):
        ea = str(ws.cell(r, 1).value or "").strip()
        if not ea:
            continue
        rows.append({"ea": ea, "ea_title": str(ws.cell(r, 2).value or "").strip(), "reason": str(ws.cell(r, 3).value or "").strip()})
    wb.close()
    return rows


def check(ea: str, path: Path) -> dict:
    norm = _norm(ea)
    for entry in _read(path):
        if _norm(entry["ea"]) == norm:
            return {"found": True, "entry": entry}
    return {"found": False, "ea": ea}


def add(ea: str, title: str, reason: str, path: Path) -> dict:
    if check(ea, path)["found"]:
        return {"added": False, "ea": ea, "message": "EA bereits auf Blacklist"}
    wb = load_workbook(path)
    ws = wb[SHEET]
    r = (ws.max_row or 1) + 1
    for c, v in enumerate((ea, title, reason), 1):
        cell = ws.cell(r, c, v)
        cell.border = _BORDER
    wb.save(path)
    wb.close()
    return {"added": True, "ea": ea}


def main() -> None:
    ap = argparse.ArgumentParser(description="EA-Blacklist check/add")
    ap.add_argument("--ea", required=True)
    ap.add_argument("--add", action="store_true")
    ap.add_argument("--title", default="")
    ap.add_argument("--reason", default="")
    ap.add_argument("--config", type=Path, default=CONFIG_DEFAULT)
    args = ap.parse_args()
    if not args.config.is_file():
        print(json.dumps({"error": f"Config nicht gefunden: {args.config}"}))
        sys.exit(1)
    result = add(args.ea, args.title, args.reason, args.config) if args.add else check(args.ea, args.config)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
