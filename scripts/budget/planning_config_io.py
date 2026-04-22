"""Single-Excel config for Beauftragungsplanung: read + create."""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

log = logging.getLogger(__name__)

SHEET_SOLVER = "Solver-Parameter"
SHEET_TARGETS = "Firmenziele"
SHEET_SONDER = "Sondervorgaben"
SHEET_PRAEMISSEN = "Praemissen"

# ── Default solver rules ──────────────────────────────────────────────

DEFAULT_RULES: list[tuple[str, str, str]] = [
    ("stage2_source", "plan_stage2_results", "Quelltabelle fuer Stage-2-Ergebnisse in der SQLite-DB. Nicht aendern, es sei denn eine alternative Ergebnis-Tabelle verwendet werden soll."),
    ("stage2_solver", "highs", "Solver-Engine fuer das MIP-Optimierungsproblem. Aktuell wird ausschliesslich 'highs' (HiGHS) unterstuetzt."),
    ("stage2_activation_penalty", "100", "Jaehrliche Strafkosten (EUR) pro genutztem Entwicklungsauftrag (EA). Hoehere Werte fuehren dazu, dass der Solver weniger verschiedene EAs aktiviert und das Budget auf weniger EAs konzentriert."),
    ("stage2_quarter_activation_penalty", "50", "Zusaetzliche Strafkosten (EUR) fuer jedes Quartal, in dem ein EA aktiv ist. Hoehere Werte reduzieren die Anzahl der Quartale pro EA und fuehren zu gebuendelteren Beauftragungen."),
    ("stage2_min_new_order_amount", "10000", "Mindestbetrag in EUR fuer neue Beauftragungen. Liegt eine geplante Beauftragung unter diesem Wert, wird sie vom Solver nicht angelegt. Verhindert Kleinstbeauftragungen."),
    ("stage2_repeat_quarter_penalty", "200", "Strafkosten (EUR) wenn dieselbe EA in mehreren aufeinanderfolgenden Quartalen beauftragt wird. Hoehere Werte foerdern eine Rotation der EAs ueber die Quartale."),
    ("stage2_stop_penalty", "500", "Strafkosten (EUR) fuer das Stoppen/Nicht-Fortfuehren eines bestehenden Konzepts (laufende Beauftragung). Schuetzt laufende Beauftragungen vor unnoetiger Unterbrechung."),
    ("stage2_large_order_penalty", "700", "Strafkosten (EUR) fuer Einzelbeauftragungen die den large_order_threshold ueberschreiten. Wirkt gegen zu grosse Einzelpositionen und foerdert eine gleichmaessigere Verteilung."),
    ("stage2_large_order_threshold", "40000", "Schwellenwert in EUR ab dem die large_order_penalty greift. Beauftragungen ueber diesem Betrag erhalten den Strafaufschlag."),
    ("stage2_existing_small_amount_penalty", "10000", "Reserviert fuer kuenftige Nutzung, derzeit nicht aktiv im Solver. Vorgesehen fuer Strafkosten bei zu kleinen Betraegen auf bestehenden Konzepten."),
    ("stage2_soft_target_penalty", "100", "Strafkosten pro EUR Abweichung wenn eine Firma unter ihrem Jahresziel (Firmenziel) bleibt. Hoehere Werte erzwingen eine genauere Einhaltung der Firmenziele."),
    ("stage2_active_ea_cap_per_quarter", "0", "Maximale Anzahl aktiver EAs pro Firma und Quartal. 0 = unbegrenzt. Bei Werten > 0 darf eine Firma in einem Quartal hoechstens diese Anzahl verschiedener EAs bedienen."),
    ("stage2_hard_need_bonus", "50", "Bonus (negative Strafkosten, EUR) fuer EAs die als 'hart benoetigt' markiert sind (is_hard=1 in Stage-1). Bevorzugt die Zuweisung auf dringend benoetigte EAs."),
    ("stage2_throughlauf_change_penalty", "2500", "Strafkosten pro EUR Reduktion bei laufenden Positionen (Durchlaeufer). Sehr hoher Wert schuetzt bestehende Beauftragungen stark vor Kuerzungen."),
    ("stage2_special_rule_priority_penalty_step", "25", "Strafkosten-Stufe (EUR) fuer Firmenprioritaet bei Sondervorgaben. Firma auf Prio 1 bekommt 0 Aufschlag, Prio 2 bekommt 1x diesen Wert, Prio 3 bekommt 2x usw."),
    ("stage2_special_rule_annual_penalty", "500", "Strafkosten pro EUR Abweichung vom Sondervorgaben-Zielbetrag. Steuert wie streng der Solver die in den Sondervorgaben definierten FL-Betraege einhält."),
    ("stage2_global_quarter_undershoot_penalty", "500", "Strafkosten pro EUR wenn ein Quartal insgesamt unter dem anteiligen Gesamtplan liegt. Verhindert, dass einzelne Quartale zu wenig Budget erhalten."),
    ("stage2_global_quarter_overshoot_penalty", "5", "Strafkosten pro EUR wenn ein Quartal insgesamt ueber dem anteiligen Gesamtplan liegt. Bewusst niedrig, da leichte Ueberplanung weniger kritisch ist."),
    ("stage2_soft_target_overshoot_penalty", "10", "Strafkosten pro EUR wenn eine Firma ueber ihrem Jahresziel liegt. Niedriger als soft_target_penalty, da Ueberplanung weniger kritisch ist als Unterplanung."),
    ("stage2_time_limit_seconds", "120", "Maximale Rechenzeit des Solvers in Sekunden. Nach Ablauf wird die beste bisher gefundene Loesung verwendet (auch wenn nicht bewiesen optimal)."),
    ("enforce_company_annual_target_consistency", "true", "Konsistenzpruefung: Abbruch wenn die Summe der Quartalsziele einer Firma nicht dem Jahresziel entspricht. Sollte normalerweise auf 'true' stehen."),
    ("btl_opt_refresh", "replace", "Aktualisierungsmodus fuer die btl_opt-Tabelle. 'replace' = alle vorhandenen Zeilen loeschen und komplett neu schreiben. Einziger unterstuetzter Modus."),
]
_KNOWN_RULES = {r[0] for r in DEFAULT_RULES}
_DEFAULT_MAP = {r[0]: r[1] for r in DEFAULT_RULES}

# ── Styles ────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT_W = Font(bold=True, size=11, color="FFFFFF")
_LOCKED_FILL = PatternFill("solid", fgColor="F2F2F2")
_EDIT_FILL = PatternFill("solid", fgColor="FFFFFF")
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _auto_width(ws, *, min_w: int = 12, max_w: int = 50) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value or "")) for c in col_cells), default=min_w)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, min_w), max_w)


def _header_row(ws, headers: list[str], row: int = 1) -> None:
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row, col, text)
        cell.font = _HEADER_FONT_W
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _WRAP


# ── Find header row by scanning ───────────────────────────────────────

def _find_header_row(ws, required_cols: set[str]) -> tuple[int, dict[str, int]]:
    """Scan sheet for row containing all required columns. Returns (row_num, {col_name: col_idx})."""
    for row_idx in range(1, min(ws.max_row or 1, 20) + 1):
        mapping = {
            str(ws.cell(row_idx, c).value or "").strip(): c
            for c in range(1, (ws.max_column or 1) + 1)
            if ws.cell(row_idx, c).value is not None
        }
        if required_cols.issubset(mapping):
            return row_idx, mapping
    raise ValueError(f"Header mit {required_cols} nicht gefunden in Sheet '{ws.title}'.")


# ── READ ──────────────────────────────────────────────────────────────

@dataclass(slots=True)
class PlanningConfig:
    rules: dict[str, str]
    company_targets: list[dict[str, Any]]
    sondervorgaben: list[dict[str, Any]]
    praemissen: str


def read_config(path: Path) -> PlanningConfig:
    """Read planning config from Excel workbook."""
    wb = load_workbook(path, data_only=True)
    rules = _read_solver_sheet(wb)
    targets = _read_targets_sheet(wb)
    sonder = _read_sonder_sheet(wb) if SHEET_SONDER in wb.sheetnames else []
    praemissen = _read_praemissen_sheet(wb) if SHEET_PRAEMISSEN in wb.sheetnames else ""
    wb.close()
    return PlanningConfig(rules=rules, company_targets=targets, sondervorgaben=sonder, praemissen=praemissen)


def _read_solver_sheet(wb: Workbook) -> dict[str, str]:
    ws = wb[SHEET_SOLVER]
    header_row, cols = _find_header_row(ws, {"Parameter", "Wert"})
    rules: dict[str, str] = {}
    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        key = str(ws.cell(row_idx, cols["Parameter"]).value or "").strip()
        if not key or key.startswith("#"):
            continue
        val = str(ws.cell(row_idx, cols["Wert"]).value or "").strip()
        if key not in _KNOWN_RULES:
            log.warning("Solver-Parameter ignoriert (Zeile %d): '%s'", row_idx, key)
            continue
        rules[key] = val
    # Fill missing with defaults
    for k, v in _DEFAULT_MAP.items():
        if k not in rules:
            log.info("Solver-Parameter '%s' nicht in Excel, verwende Standard: %s", k, v)
            rules[k] = v
    return rules


def _read_targets_sheet(wb: Workbook) -> list[dict[str, Any]]:
    ws = wb[SHEET_TARGETS]
    header_row, cols = _find_header_row(ws, {"Firma", "Jahresziel_TE"})
    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        company = str(ws.cell(row_idx, cols["Firma"]).value or "").strip()
        if not company or company.startswith("#"):
            continue
        annual_te = ws.cell(row_idx, cols["Jahresziel_TE"]).value
        if annual_te is None:
            log.warning("Firmenziel ohne Jahresziel (Zeile %d): '%s'", row_idx, company)
            continue
        entry: dict[str, Any] = {"company": company, "annual_te": int(float(str(annual_te)))}
        for q in ("Q1", "Q2", "Q3", "Q4"):
            if q in cols:
                raw = ws.cell(row_idx, cols[q]).value
                entry[q] = float(str(raw).replace("%", "").replace(",", ".")) if raw is not None else 0.25
            else:
                entry[q] = 0.25
        if "Schrittweite" in cols:
            raw = ws.cell(row_idx, cols["Schrittweite"]).value
            entry["step"] = int(float(str(raw))) if raw is not None else 1
        else:
            entry["step"] = 1
        rows.append(entry)
    return rows


_CADENCE_ALIASES: dict[str, str] = {
    "jaehrlich": "annual_exact",
    "jährlich": "annual_exact",
    "1. halbjahr": "first_half_exact",
    "pro halbjahr": "semiannual_tranche_exact",
    "pro quartal": "quarterly_tranche_exact",
    "quartalsweise": "quarterly_split_annual",
    "keine fl": "fl_forbidden",
}


def _read_sonder_sheet(wb: Workbook) -> list[dict[str, Any]]:
    ws = wb[SHEET_SONDER]
    required = {"Thema", "EA", "FL", "Kadenz", "Bemerkung"}
    try:
        header_row, cols = _find_header_row(ws, required)
    except ValueError:
        return []
    rules: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        topic = str(ws.cell(row_idx, cols["Thema"]).value or "").strip()
        if not topic:
            continue
        raw_ea = str(ws.cell(row_idx, cols["EA"]).value or "").strip()
        ea_keys = [p.strip() for p in re.split(r"[\n,;]+", raw_ea) if p.strip()]
        fl_raw = ws.cell(row_idx, cols["FL"]).value
        fl_te = float(str(fl_raw).replace(",", ".")) if fl_raw is not None and str(fl_raw).strip() else None
        el_raw = ws.cell(row_idx, cols.get("EL", 0)).value if "EL" in cols else None
        raw_kadenz = str(ws.cell(row_idx, cols["Kadenz"]).value or "").strip()
        kadenz = _CADENCE_ALIASES.get(raw_kadenz.lower(), raw_kadenz)
        remark = str(ws.cell(row_idx, cols["Bemerkung"]).value or "").strip()
        notes_col = cols.get("Hinweise", cols.get("Hinweise für autom. Auslegung", 0))
        notes = str(ws.cell(row_idx, notes_col).value or "").strip() if notes_col else ""
        allowed_col = cols.get("Erlaubte_Firmen", 0)
        allowed = str(ws.cell(row_idx, allowed_col).value or "").strip() if allowed_col else ""
        prio_col = cols.get("Prioritaet_Firmen", 0)
        prio = str(ws.cell(row_idx, prio_col).value or "").strip() if prio_col else ""
        rules.append({
            "topic": topic,
            "ea_keys": ea_keys,
            "ea_display": ", ".join(ea_keys),
            "fl_target_te": fl_te,
            "el_target": str(el_raw or "").strip() if el_raw else None,
            "remark": remark,
            "notes": notes,
            "cadence_type": kadenz or "annual_exact",
            "allowed_companies_raw": allowed,
            "priority_companies_raw": prio,
        })
    return rules


def _read_praemissen_sheet(wb: Workbook) -> str:
    ws = wb[SHEET_PRAEMISSEN]
    return str(ws.cell(1, 1).value or "").strip()


# ── Transform raw Excel → solver format ───────────────────────────────

def _norm_ea(v: Any) -> str:
    digits = "".join(ch for ch in str(v or "") if ch.isdigit())
    return (digits.lstrip("0") or "0") if digits else ""


def _canon_ea(v: Any) -> str:
    digits = "".join(ch for ch in str(v or "") if ch.isdigit())
    return digits.zfill(7) if digits else str(v or "").strip()


def _period_target(base: float | None, cadence: str, num_q: int) -> float | None:
    if base is None:
        return None
    if cadence == "first_half_exact":
        return 0.0 if num_q < 2 else base
    if cadence == "quarterly_tranche_exact":
        return base * num_q / 4 if num_q >= 1 else 0.0
    if cadence == "semiannual_tranche_exact":
        return 0.0 if num_q < 2 else base / 2
    return base * num_q / 4


def _resolve_companies(raw: str, known: list[str]) -> list[str]:
    if not raw:
        return []
    resolved: list[str] = []
    for token in re.split(r"[,;]+", raw):
        token = token.strip()
        if not token:
            continue
        upper = token.upper()
        match = next((c for c in known if upper in c.upper()), None)
        if match and match not in resolved:
            resolved.append(match)
        elif token not in resolved:
            log.warning("Sondervorgaben-Firma nicht aufgeloest: '%s'", token)
    return resolved


def transform_sondervorgaben(
    raw_rules: list[dict[str, Any]],
    known_companies: list[str],
    current_period_quarter: int,
) -> list[dict[str, Any]]:
    """Transform raw Excel sondervorgaben dicts into solver-ready format."""
    result: list[dict[str, Any]] = []
    for rule in raw_rules:
        fl_te = rule.get("fl_target_te")
        if fl_te is None:
            continue
        cadence = rule.get("cadence_type", "annual_exact")
        period_te = _period_target(fl_te, cadence, current_period_quarter)
        ea_norms = {_norm_ea(ea) for ea in rule.get("ea_keys", []) if _norm_ea(ea)}
        if not ea_norms:
            continue
        result.append({
            "topic": rule["topic"],
            "ea_keys": ea_norms,
            "candidate_eas": {_canon_ea(ea) for ea in rule.get("ea_keys", []) if _norm_ea(ea)},
            "allowed_companies": set(_resolve_companies(rule.get("allowed_companies_raw", ""), known_companies)),
            "priority_companies": _resolve_companies(rule.get("priority_companies_raw", ""), known_companies),
            "target_amount": int(round(float(fl_te) * 1000)),
            "period_target_amount": int(round(float(period_te) * 1000)) if period_te is not None else None,
            "enforce_period_exact": period_te is not None and float(period_te) > 0,
        })
    return result


# ── WRITE (create default template) ──────────────────────────────────

def create_default_config(path: Path, *, company_targets: list[dict[str, Any]] | None = None) -> Path:
    """Create a formatted Excel config template."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Sheet 1: Solver-Parameter
    ws = wb.active
    ws.title = SHEET_SOLVER
    _header_row(ws, ["Parameter", "Wert", "Standard", "Beschreibung"])
    for i, (key, default, desc) in enumerate(DEFAULT_RULES, 2):
        ws.cell(i, 1, key).fill = _LOCKED_FILL
        ws.cell(i, 1, key).border = _BORDER
        ws.cell(i, 2, default).fill = _EDIT_FILL
        ws.cell(i, 2, default).border = _BORDER
        ws.cell(i, 3, default).fill = _LOCKED_FILL
        ws.cell(i, 3, default).border = _BORDER
        ws.cell(i, 3, default).font = Font(italic=True, color="808080")
        ws.cell(i, 4, desc).fill = _LOCKED_FILL
        ws.cell(i, 4, desc).border = _BORDER
        ws.cell(i, 4, desc).alignment = _WRAP
    # Bool validation
    bool_dv = DataValidation(type="list", formula1='"true,false"', showErrorMessage=True)
    bool_dv.error = "Nur true oder false"
    ws.add_data_validation(bool_dv)
    for i, (key, _, _) in enumerate(DEFAULT_RULES, 2):
        if key == "enforce_company_annual_target_consistency":
            bool_dv.add(ws.cell(i, 2))
    _auto_width(ws)
    ws.column_dimensions["D"].width = 60

    # Sheet 2: Firmenziele
    ws2 = wb.create_sheet(SHEET_TARGETS)
    _header_row(ws2, ["Firma", "Jahresziel_TE", "Q1", "Q2", "Q3", "Q4", "Schrittweite"])
    if company_targets:
        for i, ct in enumerate(company_targets, 2):
            ws2.cell(i, 1, ct["company"]).border = _BORDER
            ws2.cell(i, 2, ct.get("annual_te", 0)).border = _BORDER
            for qi, q in enumerate(("Q1", "Q2", "Q3", "Q4"), 3):
                ws2.cell(i, qi, ct.get(q, 0.25)).border = _BORDER
                ws2.cell(i, qi).number_format = "0%"
            ws2.cell(i, 7, ct.get("step", 1)).border = _BORDER
    _auto_width(ws2)

    # Sheet 3: Sondervorgaben
    ws3 = wb.create_sheet(SHEET_SONDER)
    _header_row(ws3, ["Thema", "EA", "EL", "FL", "Kadenz", "Erlaubte_Firmen", "Prioritaet_Firmen", "Bemerkung", "Hinweise"])
    _auto_width(ws3)

    # Sheet 4: Praemissen
    ws4 = wb.create_sheet(SHEET_PRAEMISSEN)
    ws4.cell(1, 1, "").alignment = _WRAP
    ws4.column_dimensions["A"].width = 120
    ws4.row_dimensions[1].height = 200

    wb.save(path)
    wb.close()
    return path
