from __future__ import annotations

import importlib
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import pytest


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts" / "budget"
SKILL = ROOT / ".agents" / "skills" / "skill-budget-beauftragungsplanung"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import beauftragungsplanung_core as core  # noqa: E402
import planning_config_io  # noqa: E402


def _reload_wrapper():
    spec = importlib.util.spec_from_file_location(
        "report_beauftragungsplanung",
        SKILL / "report_beauftragungsplanung.py",
    )
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _write_config_excel(path: Path, *, cap: int = 0, company_targets: list | None = None) -> None:
    """Create a minimal config Excel for tests."""
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = planning_config_io.SHEET_SOLVER
    ws.cell(1, 1, "Parameter")
    ws.cell(1, 2, "Wert")
    overrides = {"stage2_active_ea_cap_per_quarter": str(cap), "stage2_large_order_penalty": "0", "stage2_soft_target_penalty": "10"}
    for i, (key, default, _) in enumerate(planning_config_io.DEFAULT_RULES, 2):
        ws.cell(i, 1, key)
        ws.cell(i, 2, overrides.get(key, default))
    ws2 = wb.create_sheet(planning_config_io.SHEET_TARGETS)
    ws2.cell(1, 1, "Firma")
    ws2.cell(1, 2, "Jahresziel_TE")
    ws2.cell(1, 3, "Schrittweite")
    for i, ct in enumerate(company_targets or [], 2):
        ws2.cell(i, 1, ct["company"])
        ws2.cell(i, 2, ct["annual_te"])
        ws2.cell(i, 3, ct.get("step", 1))
    wb.create_sheet(planning_config_io.SHEET_SONDER)
    wb.save(path)
    wb.close()


def _seed_basic_case(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        INSERT INTO plan_company_targets (year, company, quarter, target_value, annual_target, step_value)
        VALUES
            (2026, 'Firma A', 'Q1', 100, 200, 50),
            (2026, 'Firma A', 'Q2', 100, 200, 50)
        """
    )
    conn.execute(
        """
        INSERT INTO plan_stage1_results (run_id, year, ea_number, target_value, reference_score, is_hard, note)
        VALUES
            ('stage1', 2026, 'EA1', 100, 1.0, 0, NULL),
            ('stage1', 2026, 'EA2', 100, 1.0, 0, NULL)
        """
    )
    conn.commit()


def test_execute_planning_runs_solver_and_materializes_btl_opt(tmp_path, monkeypatch):
    db_path = tmp_path / "budget.db"
    config_path = tmp_path / "config.xlsx"
    report_path = tmp_path / "planung.md"
    _write_config_excel(config_path)

    monkeypatch.setattr(core, "DB_PATH", db_path)

    with core.connect() as conn:
        core.init_planning_schema(conn)
        conn.execute(f"CREATE TABLE btl ({core.BTL_COLUMNS_SQL})")
        _seed_basic_case(conn)
        conn.commit()

    result, report = core.execute_planning(
        year=2026,
        config_xlsx=str(config_path),
        output=str(report_path),
        planning_start_quarter=1,
    )

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT concept, ea, planned_value, target_date FROM btl_opt ORDER BY concept"
        ).fetchall()

    assert result["solver_status"] == "optimal"
    assert result["btl_opt_rows"] == 2
    assert Path(report).is_file()
    assert sorted(row["ea"] for row in rows) == ["EA1", "EA2"]
    assert sum(row["planned_value"] for row in rows) == 200
    assert {row["target_date"] for row in rows} == {"2026-03-31", "2026-06-30"}


def test_bootstrap_existing_orders_populates_live_btl_rows(tmp_path, monkeypatch):
    db_path = tmp_path / "budget.db"
    monkeypatch.setattr(core, "DB_PATH", db_path)

    with core.connect() as conn:
        core.init_planning_schema(conn)
        conn.execute(f"CREATE TABLE btl ({core.BTL_COLUMNS_SQL})")
        conn.execute(
            "INSERT INTO plan_company_targets (year, company, quarter, target_value, annual_target, step_value) "
            "VALUES (2026, 'Firma A', 'Q2', 100, 100, 1)"
        )
        conn.execute(
            """
            INSERT INTO btl (
                concept, ea, title, status, planned_value, org_unit, company, creator,
                bm_number, az_number, projektfamilie, dev_order, bm_text, last_updated,
                category, cost_type, quantity, unit, supplier_number,
                first_signature, second_signature, target_date, invoices
            )
            VALUES (
                'C1','EA1','EA1','07_In Planen-BM: Bestellt',50,'EKEK/1','Firma A','T',
                NULL,NULL,NULL,'EA1','BM','2026-01-01','TEST',NULL,NULL,NULL,NULL,
                NULL,NULL,'2026-06-30',NULL
            )
            """
        )
        conn.commit()
        core._bootstrap_existing_orders_from_btl(conn, 2026)
        rows = conn.execute(
            "SELECT company, quarter, ea_number, amount, note FROM plan_existing_orders"
        ).fetchall()

    assert len(rows) == 1
    assert rows[0]["company"] == "Firma A"
    assert rows[0]["quarter"] == "Q2"
    assert rows[0]["amount"] == 50
    assert rows[0]["note"] == "auto:07_In Planen-BM: Bestellt"


def test_ensure_special_company_targets_adds_voitas_zero_scope(tmp_path, monkeypatch):
    db_path = tmp_path / "budget.db"
    monkeypatch.setattr(core, "DB_PATH", db_path)

    with core.connect() as conn:
        core.init_planning_schema(conn)
        conn.execute(f"CREATE TABLE btl ({core.BTL_COLUMNS_SQL})")
        conn.executemany(
            """
            INSERT INTO plan_company_targets (year, company, quarter, target_value, annual_target, step_value)
            VALUES (2026, '4SOFT GMBH MUENCHEN', ?, 60000, 120000, 10000)
            """,
            [("Q3",), ("Q4",)],
        )
        conn.execute(
            """
            INSERT INTO btl (
                concept, ea, title, status, planned_value, org_unit, company, creator, bm_number, az_number,
                projektfamilie, dev_order, bm_text, last_updated, category, cost_type, quantity, unit,
                supplier_number, first_signature, second_signature, target_date, invoices
            )
            VALUES (
                'C1', 'EA1', 'EA1', '01_In Erstellung', 1, 'EKEK/1', ?, 'Test', NULL, NULL,
                NULL, 'EA1', 'RuleChecker', '2026-01-01', 'TEST', NULL, NULL, NULL,
                NULL, NULL, NULL, '2026-12-31', NULL
            )
            """,
            (core.VOITAS_RULECHECKER_COMPANY,),
        )
        conn.commit()
        core._ensure_special_company_targets(conn, 2026)
        rows = conn.execute(
            """
            SELECT company, quarter, target_value, annual_target, step_value
            FROM plan_company_targets
            WHERE year = 2026 AND company = ?
            ORDER BY quarter
            """,
            (core.VOITAS_RULECHECKER_COMPANY,),
        ).fetchall()

    assert [(row["quarter"], row["target_value"], row["annual_target"], row["step_value"]) for row in rows] == [
        ("Q1", 0, 0, 10000),
        ("Q2", 0, 0, 10000),
        ("Q3", 0, 0, 10000),
        ("Q4", 0, 0, 10000),
    ]


def test_run_target_ist_follow_up_uses_btl_opt(monkeypatch):
    wrapper = _reload_wrapper()
    captured: dict[str, object] = {}

    class DummyTargetModule:
        @staticmethod
        def generate_report(**kwargs):
            captured.update(kwargs)
            return {"markdown": "md", "xlsx": "xlsx"}

    monkeypatch.setattr(wrapper, "_load_target_ist_module", lambda: DummyTargetModule())
    result = wrapper._run_target_ist_follow_up(2026)

    assert result == {"markdown": "md", "xlsx": "xlsx"}
    assert captured == {"year": 2026, "source_table": "btl_opt", "open_excel": False}


def test_materialize_btl_opt_preserves_live_statuses_in_priority(tmp_path, monkeypatch):
    db_path = tmp_path / "budget.db"
    monkeypatch.setattr(core, "DB_PATH", db_path)

    with core.connect() as conn:
        core.init_planning_schema(conn)
        conn.execute(f"CREATE TABLE btl ({core.BTL_COLUMNS_SQL})")
        conn.execute(
            """
            INSERT INTO plan_stage2_results
                (run_id, year, company, quarter, ea_number, amount, source, is_locked, note)
            VALUES
                ('run1', 2026, 'Firma A', 'Q1', 'EA1', 150, 'highs', 0, NULL)
            """
        )
        conn.executemany(
            """
            INSERT INTO plan_existing_orders
                (year, company, quarter, ea_number, amount, is_fixed, can_stop, note)
            VALUES (2026, 'Firma A', 'Q1', 'EA1', ?, 0, ?, ?)
            """,
            [
                (50, 0, "auto:07_In Planen-BM: Bestellt"),
                (50, 1, "auto:06_In Bearbeitung BM-Team"),
            ],
        )
        conn.commit()
        core._materialize_btl_opt(conn, 2026)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT status, planned_value FROM btl_opt ORDER BY planned_value, status"
        ).fetchall()

    assert [(row["status"], row["planned_value"]) for row in rows] == [
        (core.CURRENT_QUARTER_TODO_STATUS, 50),
        ("06_In Bearbeitung BM-Team", 50),
        ("07_In Planen-BM: Bestellt", 50),
    ]


def test_materialize_btl_opt_suppresses_passthrough_for_explicit_zero_stage2_rows(tmp_path, monkeypatch):
    db_path = tmp_path / "budget.db"
    monkeypatch.setattr(core, "DB_PATH", db_path)

    with core.connect() as conn:
        core.init_planning_schema(conn)
        conn.execute(f"CREATE TABLE btl ({core.BTL_COLUMNS_SQL})")
        conn.execute(
            """
            INSERT INTO plan_stage2_results
                (run_id, year, company, quarter, ea_number, amount, source, is_locked, note)
            VALUES
                ('run1', 2026, 'Firma A', 'Q1', 'EA_DROP', 0, 'highs', 0, NULL),
                ('run1', 2026, 'Firma A', 'Q1', 'EA_KEEP', 100, 'highs', 0, NULL)
            """
        )
        conn.execute(
            """
            INSERT INTO btl (
                concept, ea, title, status, planned_value, org_unit, company, creator, bm_number, az_number,
                projektfamilie, dev_order, bm_text, last_updated, category, cost_type, quantity, unit,
                supplier_number, first_signature, second_signature, target_date, invoices
            )
            VALUES (
                'C1', 'Legacy EA', 'Legacy EA', '01_In Erstellung', 50, 'EKEK/1', 'Firma A', 'Test', NULL, NULL,
                NULL, 'EA_DROP', 'GW', '2026-01-01', 'TEST', NULL, NULL, NULL,
                NULL, NULL, NULL, '2026-03-31', NULL
            )
            """
        )
        conn.commit()
        core._materialize_btl_opt(conn, 2026)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT ea, planned_value FROM btl_opt ORDER BY ea"
        ).fetchall()

    assert [(row["ea"], row["planned_value"]) for row in rows] == [("EA_KEEP", 100)]


def test_materialize_btl_opt_marks_new_current_quarter_rows_as_early_durchlauf(tmp_path, monkeypatch):
    db_path = tmp_path / "budget.db"
    monkeypatch.setattr(core, "DB_PATH", db_path)

    class FixedDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 4, 18, 12, 0, 0)

    monkeypatch.setattr(core, "datetime", FixedDateTime)

    with core.connect() as conn:
        core.init_planning_schema(conn)
        conn.execute(f"CREATE TABLE btl ({core.BTL_COLUMNS_SQL})")
        conn.execute(
            """
            INSERT INTO plan_stage2_results
                (run_id, year, company, quarter, ea_number, amount, source, is_locked, note)
            VALUES
                ('run1', 2026, 'Firma A', 'Q2', 'EA_Q2', 100, 'highs', 0, NULL)
            """
        )
        conn.commit()
        core._materialize_btl_opt(conn, 2026)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT ea, planned_value, status FROM btl_opt"
        ).fetchall()

    assert [(row["ea"], row["planned_value"], row["status"]) for row in rows] == [
        ("EA_Q2", 100, core.CURRENT_QUARTER_TODO_STATUS)
    ]


def test_materialize_btl_opt_merges_existing_and_new_concept_amounts_with_same_status(tmp_path, monkeypatch):
    db_path = tmp_path / "budget.db"
    monkeypatch.setattr(core, "DB_PATH", db_path)

    class FixedDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 4, 1, 12, 0, 0)

    monkeypatch.setattr(core, "datetime", FixedDateTime)

    with core.connect() as conn:
        core.init_planning_schema(conn)
        conn.execute(f"CREATE TABLE btl ({core.BTL_COLUMNS_SQL})")
        conn.execute(
            """
            INSERT INTO plan_stage2_results
                (run_id, year, company, quarter, ea_number, amount, source, is_locked, note)
            VALUES
                ('run1', 2026, 'Firma A', 'Q4', 'EA1', 10000, 'highs', 0, NULL)
            """
        )
        conn.execute(
            """
            INSERT INTO plan_existing_orders
                (year, company, quarter, ea_number, amount, is_fixed, can_stop, note)
            VALUES (2026, 'Firma A', 'Q4', 'EA1', 1, 0, 1, 'auto:01_In Erstellung')
            """
        )
        conn.execute(
            """
            INSERT INTO btl (
                concept, ea, title, status, planned_value, org_unit, company, creator, bm_number, az_number,
                projektfamilie, dev_order, bm_text, last_updated, category, cost_type, quantity, unit,
                supplier_number, first_signature, second_signature, target_date, invoices
            )
            VALUES (
                'C1', 'EA1', 'EA1', '01_In Erstellung', 1, 'EKEK/1', 'Firma A', 'Test', NULL, NULL,
                NULL, 'EA1', 'GW', '2026-01-01', 'TEST', NULL, NULL, NULL,
                NULL, NULL, NULL, '2026-12-31', NULL
            )
            """
        )
        conn.commit()
        core._materialize_btl_opt(conn, 2026)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT status, planned_value FROM btl_opt ORDER BY status, planned_value"
        ).fetchall()

    assert [(row["status"], row["planned_value"]) for row in rows] == [
        ("01_In Erstellung", 10000)
    ]


def test_materialize_btl_opt_past_quarter_rows_get_durchlauf_status(tmp_path, monkeypatch):
    """Solver rows for past quarters (Q1 when current=Q2) must get Durchlauf status."""
    db_path = tmp_path / "budget.db"
    monkeypatch.setattr(core, "DB_PATH", db_path)

    class FixedDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 4, 18, 12, 0, 0)

    monkeypatch.setattr(core, "datetime", FixedDateTime)

    with core.connect() as conn:
        core.init_planning_schema(conn)
        conn.execute(f"CREATE TABLE btl ({core.BTL_COLUMNS_SQL})")
        conn.execute(
            """
            INSERT INTO plan_stage2_results
                (run_id, year, company, quarter, ea_number, amount, source, is_locked, note)
            VALUES
                ('run1', 2026, 'Firma A', 'Q1', 'EA_Q1', 200, 'highs', 0, NULL)
            """
        )
        conn.commit()
        core._materialize_btl_opt(conn, 2026)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT ea, planned_value, status FROM btl_opt").fetchall()

    assert [(row["ea"], row["planned_value"], row["status"]) for row in rows] == [
        ("EA_Q1", 200, core.CURRENT_QUARTER_TODO_STATUS)
    ]


def test_bootstrap_sets_can_stop_zero_for_bestellt_and_durchlauf(tmp_path, monkeypatch):
    """bestellt and im Durchlauf rows must get can_stop=0; Konzept gets can_stop=1."""
    db_path = tmp_path / "budget.db"
    monkeypatch.setattr(core, "DB_PATH", db_path)

    with core.connect() as conn:
        core.init_planning_schema(conn)
        conn.execute(f"CREATE TABLE btl ({core.BTL_COLUMNS_SQL})")
        conn.execute(
            "INSERT INTO plan_company_targets (year, company, quarter, target_value, annual_target, step_value) "
            "VALUES (2026, 'Firma A', 'Q2', 100, 100, 1)"
        )
        for status, ea in [
            ("07_In Planen-BM: Bestellt", "EA1"),
            ("06_In Bearbeitung BM-Team", "EA2"),
            ("01_In Erstellung", "EA3"),
        ]:
            conn.execute(
                f"""
                INSERT INTO btl (
                    concept, ea, title, status, planned_value, org_unit, company, creator,
                    bm_number, az_number, projektfamilie, dev_order, bm_text, last_updated,
                    category, cost_type, quantity, unit, supplier_number,
                    first_signature, second_signature, target_date, invoices
                )
                VALUES (
                    'C1','{ea}','{ea}','{status}',50,'EKEK/1','Firma A','T',
                    NULL,NULL,NULL,'{ea}','BM','2026-01-01','TEST',NULL,NULL,NULL,NULL,
                    NULL,NULL,'2026-06-30',NULL
                )
                """
            )
        conn.commit()
        core._bootstrap_existing_orders_from_btl(conn, 2026)
        rows = {
            row["ea_number"]: row["can_stop"]
            for row in conn.execute("SELECT ea_number, can_stop FROM plan_existing_orders").fetchall()
        }
    assert rows["EA1"] == 0  # bestellt
    assert rows["EA2"] == 0  # im Durchlauf
    assert rows["EA3"] == 1  # Konzept


def test_materialize_btl_opt_excludes_storniert_passthrough(tmp_path, monkeypatch):
    """storniert/abgelehnt BTL rows must NOT pass through to btl_opt."""
    db_path = tmp_path / "budget.db"
    monkeypatch.setattr(core, "DB_PATH", db_path)

    with core.connect() as conn:
        core.init_planning_schema(conn)
        conn.execute(f"CREATE TABLE btl ({core.BTL_COLUMNS_SQL})")
        conn.execute(
            """
            INSERT INTO plan_stage2_results
                (run_id, year, company, quarter, ea_number, amount, source, is_locked, note)
            VALUES ('run1', 2026, 'Firma A', 'Q1', 'EA_KEEP', 100, 'highs', 0, NULL)
            """
        )
        for status, ea in [
            ("09_Storniert", "EA_STORNO"),
            ("10_Abgelehnt", "EA_REJECT"),
            ("01_In Erstellung", "EA_OK"),
        ]:
            conn.execute(
                f"""
                INSERT INTO btl (
                    concept, ea, title, status, planned_value, org_unit, company, creator,
                    bm_number, az_number, projektfamilie, dev_order, bm_text, last_updated,
                    category, cost_type, quantity, unit, supplier_number,
                    first_signature, second_signature, target_date, invoices
                )
                VALUES (
                    'C1','{ea}','{ea}','{status}',50,'EKEK/1','Firma A','T',
                    NULL,NULL,NULL,'{ea}','BM','2026-01-01','TEST',NULL,NULL,NULL,NULL,
                    NULL,NULL,'2026-03-31',NULL
                )
                """
            )
        conn.commit()
        core._materialize_btl_opt(conn, 2026)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        eas = sorted(row["ea"] for row in conn.execute("SELECT ea FROM btl_opt").fetchall())

    assert "EA_STORNO" not in eas
    assert "EA_REJECT" not in eas
    assert "EA_OK" in eas
    assert "EA_KEEP" in eas


def test_materialize_btl_opt_keeps_dev_order_active_from_passthrough_rows(tmp_path, monkeypatch):
    db_path = tmp_path / "budget.db"
    monkeypatch.setattr(core, "DB_PATH", db_path)

    with core.connect() as conn:
        core.init_planning_schema(conn)
        conn.execute(f"CREATE TABLE btl ({core.BTL_COLUMNS_SQL})")
        conn.execute(
            """
            INSERT INTO plan_stage2_results
                (run_id, year, company, quarter, ea_number, amount, source, is_locked, note)
            VALUES ('run1', 2026, 'Firma A', 'Q1', 'EA_PLAN', 100, 'highs', 0, NULL)
            """
        )
        conn.execute(
            """
            INSERT INTO btl (
                concept, ea, title, status, planned_value, org_unit, company, creator,
                bm_number, az_number, projektfamilie, dev_order, bm_text, last_updated,
                category, cost_type, quantity, unit, supplier_number,
                first_signature, second_signature, target_date, invoices, dev_order_active
            )
            VALUES (
                'C1','EA_PASS','EA_PASS','01_In Erstellung',50,'EKEK/1','Firma A','T',
                NULL,NULL,NULL,'EA_PASS','BM','2026-01-01','TEST',NULL,NULL,NULL,NULL,
                NULL,NULL,'2026-03-31',NULL,0
            )
            """
        )
        conn.commit()
        core._materialize_btl_opt(conn, 2026)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = {
            row["ea"]: row["dev_order_active"]
            for row in conn.execute("SELECT ea, dev_order_active FROM btl_opt").fetchall()
        }

    assert rows["EA_PASS"] == 0
    assert rows["EA_PLAN"] is None
