from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import Workbook, load_workbook


MODULE_PATH = (
    Path(__file__).resolve().parents[1]
    / ".agents"
    / "skills"
    / "skill-budget-target-ist-analyse"
    / "report_massnahmenplan.py"
)


def load_module():
    spec = importlib.util.spec_from_file_location("report_massnahmenplan", MODULE_PATH)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def sample_report(mod):
    return mod.ReportDocument(
        title="Testreport",
        meta_lines=["**Stand BPLUS-NG:** 2026-04-09T12:00:00 | **Erstellt:** 09.04.2026"],
        sections=[
            mod.TableSection(
                title="Gesamtübersicht",
                headers=["", "Wert"],
                rows=[
                    mod.TableRow(["Ist (BPLUS)", "100 T€"]),
                    mod.TableRow(["Summe", "100 T€"], style="summary"),
                ],
                align_right={1},
                separator_before=False,
            ),
            mod.TextSection(
                title="Korrektur Überplanung",
                lines=["**Aktion-Spalte** manuell befüllen"],
                sheet_name="Korrektur",
            ),
            mod.TableSection(
                title="Lieferant A — Jahres-Korrektur: 50 T€ zu reduzieren",
                headers=["Firma", "Typ", "Konzept", "EA-Nr.", "EA", "BM-Titel", "Wert ist", "Wert neu", "Status", "Aktion"],
                rows=[mod.TableRow(["Lieferant A", "Reduzieren", "K1", "0001", "EA1", "BM 1", "50 T€", "", "01 Erstellung", ""])],
                align_right={6, 7},
                level=3,
                separator_before=False,
                sheet_name="Korrektur",
                body_row_height=15,
            ),
            mod.TextSection(
                title="Hinweise",
                lines=["### Zwischenüberschrift", "- Erster Punkt", "", "**Wichtig**"],
            ),
        ],
        max_columns=11,
    )


def test_render_markdown_formats_summary_and_sections():
    mod = load_module()
    report = sample_report(mod)

    content = mod.render_markdown(report)

    assert "# Testreport" in content
    assert "## Gesamtübersicht" in content
    assert "| **Summe** | **100 T€** |" in content
    assert "## Hinweise" in content
    assert "**Wichtig**" in content


def test_render_markdown_marks_btl_opt_as_optimization_proposal():
    mod = load_module()
    report = mod.ReportDocument(
        title="Testreport",
        meta_lines=[mod._source_notice_line("btl_opt")],
        sections=[],
        max_columns=1,
    )

    content = mod.render_markdown(report)

    assert "OPTIMIERUNGSVORSCHLAG AUS btl_opt" in content
    assert "<span style=\"color:#c00000\"><strong>" in content


def test_exclude_from_budget_filters_storniert():
    mod = load_module()

    assert mod._exclude_from_budget("storniert") is False
    assert mod._exclude_from_budget("bestellt") is False
    assert mod._exclude_from_budget("im Durchlauf") is False


def test_hide_from_firm_overview_filters_voitas():
    mod = load_module()

    assert mod._hide_from_firm_overview("Voitas GmbH") is True
    assert mod._hide_from_firm_overview("4Soft GmbH") is False


def test_thiesen_manual_soll_includes_both_thiesen_areas():
    mod = load_module()

    result = mod._thiesen_manual_soll(
        "Thiesen GmbH",
        {
            "Bordnetz Support, RollOut (Thiesen)": 288,
            "Spez. und Test VOBES2025 (Thiesen)": 263,
        },
        {
            "Bordnetz Support, RollOut (Thiesen)": 1,
            "Spez. und Test VOBES2025 (Thiesen)": 1,
        },
        2,
    )

    assert result == (551000, 275500)


def test_write_xlsx_report_writes_titles_and_summary_style(tmp_path):
    mod = load_module()
    report = sample_report(mod)
    output = tmp_path / "report.xlsx"

    mod.write_xlsx_report(report, str(output))

    workbook = load_workbook(output)
    overview = workbook["Übersicht"]
    correction = workbook["Korrektur"]

    positions = {}
    for row in overview.iter_rows():
        for cell in row:
            if cell.value in {"Gesamtübersicht", "Summe", "Hinweise", "Wichtig", "Zwischenüberschrift"}:
                positions[cell.value] = cell.coordinate

    correction_values = {
        cell.value
        for row in correction.iter_rows()
        for cell in row
        if cell.value is not None
    }

    assert workbook.sheetnames == ["Übersicht", "Korrektur"]
    assert overview["A1"].value == "Testreport"
    assert overview.row_dimensions[1].height and overview.row_dimensions[1].height >= 30
    assert overview["A1"].fill.fill_type is None
    assert correction.row_dimensions[1].height and correction.row_dimensions[1].height >= 24
    assert overview.column_dimensions["A"].width == 44
    assert correction.column_dimensions["A"].width <= 20
    assert correction.column_dimensions["E"].width == 48
    assert correction.row_dimensions[5].height == 15
    assert positions["Gesamtübersicht"] == "A3"
    assert positions["Summe"] == "A6"
    assert overview[positions["Summe"]].font.bold is True
    assert overview["B5"].value == 100
    assert 'T€' in overview["B5"].number_format
    assert positions["Hinweise"] == "A8"
    assert positions["Zwischenüberschrift"] == "A9"
    assert overview["A9"].font.bold is True
    assert positions["Wichtig"] == "A12"
    assert correction["A1"].value == "Korrektur Überplanung"
    assert "Korrektur Überplanung" in correction_values
    assert "Lieferant A — Jahres-Korrektur: 50 T€ zu reduzieren" in correction_values


def test_write_xlsx_report_creates_neutral_tables_only_on_korrektur(tmp_path):
    mod = load_module()
    report = sample_report(mod)
    output = tmp_path / "report.xlsx"

    mod.write_xlsx_report(report, str(output))

    workbook = load_workbook(output)

    assert len(workbook["Übersicht"].tables) == 0
    assert len(workbook["Korrektur"].tables) == 1
    assert workbook["Korrektur"].tables["Korrektur_1"].ref == "A4:J5"


def test_write_xlsx_report_korrektur_table_falls_back_to_full_block_for_interleaved_summary(tmp_path):
    mod = load_module()
    report = mod.ReportDocument(
        title="Testreport",
        meta_lines=["Meta"],
        sections=[
            mod.TableSection(
                title="Block",
                headers=["Firma", "Typ", "Konzept", "EA-Nr.", "EA", "BM-Titel", "Wert ist", "Wert neu", "Status", "Aktion"],
                rows=[
                    mod.TableRow(["4SOFT", "Reduzieren", "831058", "0043516", "EA1", "Titel 1", "2 T€", "", "01_In Erstellung", ""]),
                    mod.TableRow(["", "", "", "", "", "Summe", "2 T€", "", "", ""], style="summary"),
                    mod.TableRow(["EDAG", "Reduzieren", "831059", "0043517", "EA2", "Titel 2", "3 T€", "", "im Durchlauf", ""]),
                ],
                align_right={6, 7},
                sheet_name="Korrektur",
                separator_before=False,
            ),
        ],
        max_columns=10,
    )
    output = tmp_path / "report.xlsx"

    mod.write_xlsx_report(report, str(output))

    workbook = load_workbook(output)

    assert workbook["Korrektur"].tables["Korrektur_1"].ref == "A2:J5"


def test_write_xlsx_report_marks_btl_opt_notice_red_and_bold(tmp_path):
    mod = load_module()
    report = mod.ReportDocument(
        title="Testreport",
        meta_lines=[mod._source_notice_line("btl_opt"), "Meta"],
        sections=[],
        max_columns=1,
    )
    output = tmp_path / "report.xlsx"

    mod.write_xlsx_report(report, str(output))

    workbook = load_workbook(output)
    overview = workbook["Übersicht"]

    assert overview["A2"].value == "OPTIMIERUNGSVORSCHLAG AUS btl_opt - keine Ist-Analyse aus btl"
    assert overview["A2"].font.bold is True
    assert overview["A2"].font.color.type == "rgb"
    assert overview["A2"].font.color.rgb == "FFC00000"


def test_reporting_company_maps_voitas_to_4soft():
    mod = load_module()

    assert mod._reporting_company("VOITAS ENGINEERING GMBH GEISENFELD") == "4SOFT GMBH MUENCHEN"
    assert mod._reporting_company("4SOFT GMBH MUENCHEN") == "4SOFT GMBH MUENCHEN"


def test_firm_period_value_eur_uses_global_ordered_and_future_quarter_concept_only():
    mod = load_module()
    totals = mod._new_firm_totals()
    mod._accumulate_firm_totals(totals, "bestellt", 470_000, "Q1")
    mod._accumulate_firm_totals(totals, "im Durchlauf", 78_000, "Q2")
    mod._accumulate_firm_totals(totals, "Konzept", 168_000, "Q3")
    mod._accumulate_firm_totals(totals, "Konzept", 113_000, "Q4")

    assert mod._firm_period_value_eur(totals, 2) == 548_000
    assert mod._firm_period_value_eur(totals, 3) == 716_000
    assert mod._firm_period_value_eur(totals, 4) == 661_000


def test_load_reporting_company_targets_prefers_plan_company_targets(tmp_path, monkeypatch):
    mod = load_module()
    db_path = tmp_path / "budget.db"
    monkeypatch.setattr(mod, "DB_PATH", str(db_path))

    import sqlite3

    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE plan_company_targets (
                year INTEGER NOT NULL,
                company TEXT NOT NULL,
                quarter TEXT NOT NULL,
                target_value INTEGER NOT NULL,
                annual_target INTEGER,
                quarter_cap INTEGER,
                step_value INTEGER NOT NULL DEFAULT 1
            )
            """
        )
        conn.executemany(
            """
            INSERT INTO plan_company_targets
                (year, company, quarter, target_value, annual_target, step_value)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            [
                (2026, "4SOFT GMBH MUENCHEN", "Q1", 100_000, 400_000, 100_000),
                (2026, "4SOFT GMBH MUENCHEN", "Q2", 120_000, 400_000, 100_000),
                (2026, "4SOFT GMBH MUENCHEN", "Q3", 80_000, 400_000, 100_000),
                (2026, "4SOFT GMBH MUENCHEN", "Q4", 100_000, 400_000, 100_000),
                (2026, "EDAG ENGINEERING GMBH WOLFSBURG", "Q1", 200_000, 1_088_000, 100_000),
                (2026, "EDAG ENGINEERING GMBH WOLFSBURG", "Q2", 300_000, 1_088_000, 100_000),
            ],
        )
        conn.commit()

    annual_targets, quarter_targets = mod._load_reporting_company_targets(2026)

    assert annual_targets["4SOFT GMBH MUENCHEN"] == 400_000
    assert annual_targets["EDAG ENGINEERING GMBH WOLFSBURG"] == 500_000
    assert quarter_targets["4SOFT GMBH MUENCHEN"] == {
        "Q1": 100_000,
        "Q2": 120_000,
        "Q3": 80_000,
        "Q4": 100_000,
    }
    assert mod._cumulative_company_targets(quarter_targets, 2)["4SOFT GMBH MUENCHEN"] == 220_000


def test_build_quarter_todo_sections_creates_current_and_future_lists():
    mod = load_module()

    sections = mod._build_quarter_todo_sections(
        {
            "Q2": [
                {
                    "reporting_company": "4SOFT GMBH MUENCHEN",
                    "todo_type": "Neu",
                    "concept": "OPT-2026-Q2-0001",
                    "ea_number": "0048040",
                    "ea": "0048040",
                    "title": "Vorentwicklung (4soft) 0048040",
                    "value_ist": "",
                    "value_neu": 75_000,
                    "sort_value": 75_000,
                    "status_raw": "02_Freigabe Kostenstelle",
                    "action": "Abruf erstellen / in Durchlauf schicken",
                }
            ],
            "Q3": [
                {
                    "reporting_company": "EDAG ENGINEERING GMBH WOLFSBURG",
                    "todo_type": "Neu",
                    "concept": "OPT-2026-Q3-0001",
                    "ea_number": "0093037",
                    "ea": "0093037",
                    "title": "Systemschaltpläne 0093037",
                    "value_ist": "",
                    "value_neu": 125_000,
                    "sort_value": 125_000,
                    "status_raw": "01_In Erstellung",
                    "action": "Abruf vorbereiten / für Q3 vormerken",
                }
            ],
        },
        current_period_quarter=2,
    )

    assert [section.title for section in sections] == [
        "Q2-Todo: Maßnahmen jetzt umsetzen",
        "Q2-Todo-Liste",
        "Q3-Todo: künftige Maßnahmen vorbereiten",
        "Q3-Todo-Liste",
    ]
    assert sections[1].headers == ["Firma", "Typ", "Konzept", "EA-Nr.", "EA", "BM-Titel", "Wert ist", "Wert neu", "Status", "Aktion"]
    assert sections[1].rows[0].values[1] == "Neu"
    assert sections[1].rows[0].values[5] == "Vorentwicklung (4soft) 0048040"
    assert sections[1].rows[0].values[6] == ""
    assert sections[1].rows[0].values[7] == "75 T€"
    assert sections[1].rows[0].values[9] == "Abruf erstellen / in Durchlauf schicken"
    assert sections[3].rows[0].values[6] == ""
    assert sections[3].rows[0].values[7] == "125 T€"
    assert sections[3].rows[0].values[9] == "Abruf vorbereiten / für Q3 vormerken"


def test_build_quarter_delta_todos_creates_delete_and_reduce_entries():
    mod = load_module()

    actual_entries = [
        {
            "concept": "A1",
            "ea": "EA1",
            "ea_number": "EA1",
            "title": "Konzept A",
            "value": 20_000,
            "status_label": "Konzept",
            "status_raw": "01_In Erstellung",
            "source_company": "Firma A",
            "reporting_company": "Firma A",
            "quarter": "Q3",
            "area": "Area",
        },
        {
            "concept": "A2",
            "ea": "EA1",
            "ea_number": "EA1",
            "title": "Durchlauf A",
            "value": 30_000,
            "status_label": "im Durchlauf",
            "status_raw": "02_Freigabe Kostenstelle",
            "source_company": "Firma A",
            "reporting_company": "Firma A",
            "quarter": "Q3",
            "area": "Area",
        },
        {
            "concept": "B1",
            "ea": "EA2",
            "ea_number": "EA2",
            "title": "Konzept B",
            "value": 50_000,
            "status_label": "Konzept",
            "status_raw": "01_In Erstellung",
            "source_company": "Firma B",
            "reporting_company": "Firma B",
            "quarter": "Q4",
            "area": "Area",
        },
        {
            "concept": "OLD",
            "ea": "EA0",
            "ea_number": "EA0",
            "title": "Vergangen",
            "value": 40_000,
            "status_label": "Konzept",
            "status_raw": "01_In Erstellung",
            "source_company": "Firma Alt",
            "reporting_company": "Firma Alt",
            "quarter": "Q1",
            "area": "Area",
        },
    ]
    target_entries = [
        {
            "source_company": "Firma A",
            "quarter": "Q3",
            "ea_number": "EA1",
            "value": 30_000,
            "status_label": "Konzept",
        },
        {
            "source_company": "Firma B",
            "quarter": "Q4",
            "ea_number": "EA2",
            "value": 20_000,
            "status_label": "Konzept",
        },
    ]

    todos = mod._build_quarter_delta_todos(
        actual_entries,
        target_entries,
        current_period_quarter=2,
    )

    assert [entry["todo_type"] for entry in todos["Q3"]] == ["Löschen"]
    assert todos["Q3"][0]["concept"] == "A1"
    assert todos["Q3"][0]["value_ist"] == 20_000
    assert todos["Q3"][0]["value_neu"] == 0
    assert todos["Q3"][0]["action"] == "Abruf löschen / zurückziehen"

    assert [entry["todo_type"] for entry in todos["Q4"]] == ["Reduzieren"]
    assert todos["Q4"][0]["concept"] == "B1"
    assert todos["Q4"][0]["value_ist"] == 50_000
    assert todos["Q4"][0]["value_neu"] == 20_000
    assert todos["Q4"][0]["action"] == "Abruf auf 20 T€ reduzieren"
    assert "Q1" not in todos


def test_budget_entry_from_row_resolves_numeric_ea_to_title():
    mod = load_module()

    entry = mod._budget_entry_from_row(
        {
            "concept": "OPT-1",
            "ea": "0093037",
            "dev_order": "0093037",
            "title": "Dummy BM",
            "planned_value": 10000,
            "company": "EDAG ENGINEERING GMBH WOLFSBURG",
            "status": "01_In Erstellung",
            "bm_text": "",
            "target_date": "2026-09-30",
        },
        {"01_In Erstellung": "Konzept"},
        {"0093037": "Arbeiten EK für China nicht projektbez."},
    )

    assert entry is not None
    assert entry["ea"] == "Arbeiten EK für China nicht projektbez."
    assert entry["ea_number"] == "0093037"


def test_load_targets_reads_company_target_overrides():
    mod = load_module()

    _, _, _, _, overrides, _quarter_split = mod.load_targets()

    assert "BERTRANDT" in overrides
    assert overrides["BERTRANDT"]["area"] == "Systemschaltpläne und Bibl. (EDAG, Bertrandt)"
    assert overrides["BERTRANDT"]["target_te"] == 905


def test_write_xlsx_report_inherits_notes_and_column_widths(tmp_path):
    mod = load_module()
    report = mod.ReportDocument(
        title="Testreport",
        meta_lines=["Meta"],
        sections=[
            mod.TableSection(
                title="Target - Ist - Vergleich",
                headers=["Aufgabenbereich", "2025", "Target", "Ist", "Delta", "Maßnahmen"],
                rows=[mod.TableRow(["Area A", "1 T€", "2 T€", "3 T€", "+1 T€", ""])],
                align_right={1, 2, 3, 4},
                separator_before=False,
            ),
            mod.TableSection(
                title="Firmen-Übersicht",
                headers=["Firma", "BMs", "Soll", "Ist", "DIFF Ges.", "Soll Q1-2", "bestellt", "im Durchlauf", "01 Erstellung", "storniert", "DIFF Q1-2", "Maßnahmen"],
                rows=[mod.TableRow(["4SOFT", "2", "10 T€", "12 T€", "+2 T€", "5 T€", "2 T€", "1 T€", "3 T€", "0 T€", "-1 T€", "auto"])],
                align_right={1, 2, 3, 4, 5, 6, 7, 8, 9, 10},
            ),
            mod.TextSection(title="Korrektur Überplanung", lines=["Hinweis"], sheet_name="Korrektur"),
            mod.TableSection(
                title="4SOFT — Jahres-Korrektur: 2 T€ zu reduzieren",
                headers=["Firma", "Typ", "Konzept", "EA-Nr.", "EA", "BM-Titel", "Wert ist", "Wert neu", "Status", "Aktion"],
                rows=[mod.TableRow(["4SOFT", "Reduzieren", "831058", "0043516", "EA1", "Titel", "2 T€", "", "01_In Erstellung", ""])],
                align_right={6, 7},
                level=3,
                separator_before=False,
                sheet_name="Korrektur",
                body_row_height=15,
            ),
        ],
        max_columns=13,
    )

    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsx"

    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment

    wb = Workbook()
    overview = wb.active
    overview.title = "Übersicht"
    correction = wb.create_sheet("Korrektur")
    overview["A1"] = "Aufgabenbereich"
    overview["F1"] = "Maßnahmen"
    overview["A2"] = "Area A"
    overview["F2"] = "Bereichsnotiz"
    overview["A4"] = "Firma"
    overview["M4"] = "Notizen"
    overview["A5"] = "4SOFT"
    overview["M5"] = "Firmennotiz"
    overview.column_dimensions["F"].width = 33
    overview.column_dimensions["M"].width = 77
    overview["M5"].font = Font(color="FFFF0000", bold=False)
    overview["M5"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    overview["M5"].border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    correction["A1"] = "4SOFT — Jahres-Korrektur: 2 T€ zu reduzieren"
    correction["A2"] = "Firma"
    correction["A3"] = "4SOFT"
    correction["B3"] = "Reduzieren"
    correction["C3"] = "831058"
    correction["D3"] = "0043516"
    correction["E3"] = "EA1"
    correction["F3"] = "Titel"
    correction["G3"] = "2 T€"
    correction["I3"] = "01_In Erstellung"
    correction["J3"] = "manuell"
    correction.column_dimensions["J"].width = 88
    wb.save(source)

    mod.write_xlsx_report(report, str(target), inherit_notes_from=str(source))

    out = load_workbook(target)
    out_overview = out["Übersicht"]
    out_correction = out["Korrektur"]

    area_row = next(row for row in range(1, out_overview.max_row + 1) if out_overview[f"A{row}"].value == "Area A")
    firm_header_row = next(row for row in range(1, out_overview.max_row + 1) if out_overview[f"A{row}"].value == "Firma")
    firm_row = next(row for row in range(firm_header_row + 1, out_overview.max_row + 1) if out_overview[f"A{row}"].value == "4SOFT")
    correction_row = next(row for row in range(1, out_correction.max_row + 1) if out_correction[f"C{row}"].value == "831058")

    assert out_overview.cell(area_row, 6).value == "Bereichsnotiz"
    assert out_overview.cell(firm_header_row, 13).value == "Notizen"
    assert out_overview.cell(firm_row, 13).value == "Firmennotiz"
    assert out_overview.cell(firm_row, 13).font.color.type == "rgb"
    assert out_overview.cell(firm_row, 13).font.color.rgb == "FFFF0000"
    assert out_overview.column_dimensions["F"].width == 33
    assert out_overview.column_dimensions["M"].width == 77
    assert out_correction.cell(correction_row, 10).value == "manuell"
    assert out_correction.column_dimensions["J"].width == 88


def test_load_special_rule_rows_reads_xlsx_and_resolves_companies(tmp_path):
    mod = load_module()
    workbook = Workbook()
    ws = workbook.active
    ws["A1"] = "Sondervereinbarungen"
    ws.append([])
    ws.append(["Thema", "EA", "EL", "FL", "Bemerkung", "Link", "Hinweise für autom. Auslegung"])
    ws.append([
        "PMT",
        "43932",
        "8.449h",
        1048,
        "Hinweis",
        "",
        "EL und FL müssen exakt passen!\n"
        "EL und FL dürfen nur Quartalsweise beauftragen und werden sehr genau beobachtet.\n"
        "Folgende Firmen können auf PMT buchen: 4soft, Thiesen, FES, Sumitomo in genau dieser Priorität.",
    ])
    path = tmp_path / "sondervereinbarungen.xlsx"
    workbook.save(path)

    rows = mod.load_special_rule_rows(
        str(path),
        [
            "4SOFT GMBH MUENCHEN",
            "THIESEN HARDWARE SOFTW. GMBH WARTENBERG",
            "FES GMBH ZWICKAU/00",
            "SUMITOMO ELECTRIC BORDNETZE SE WOLFSBURG",
        ],
    )

    assert len(rows) == 1
    row = rows[0]
    assert row["topic"] == "PMT"
    assert row["ea_keys"] == ["43932"]
    assert row["fl_target_te"] == 1048.0
    assert row["el_target"]["mode"] == "hours"
    assert row["el_target"]["value"] == 8449
    assert row["cadence_type"] == "quarterly_split_annual"
    assert row["allowed_companies"] == [
        "4SOFT GMBH MUENCHEN",
        "THIESEN HARDWARE SOFTW. GMBH WARTENBERG",
        "FES GMBH ZWICKAU/00",
        "SUMITOMO ELECTRIC BORDNETZE SE WOLFSBURG",
    ]
    assert row["priority_companies"] == row["allowed_companies"]


def test_build_special_rule_section_computes_status_columns(tmp_path, monkeypatch):
    mod = load_module()
    workbook = Workbook()
    ws = workbook.active
    ws.append(["Sondervereinbarungen"])
    ws.append([])
    ws.append(["Thema", "EA", "EL", "FL", "Bemerkung", "Link", "Hinweise für autom. Auslegung"])
    ws.append([
        "PMT",
        "43932",
        "8.449h",
        1048,
        "Hinweis",
        "",
        "EL und FL müssen exakt passen!\n"
        "EL und FL dürfen nur Quartalsweise beauftragen und werden sehr genau beobachtet.\n"
        "Folgende Firmen können auf PMT buchen: 4soft, Thiesen, FES, Sumitomo in genau dieser Priorität.",
    ])
    special_path = tmp_path / "sondervereinbarungen.xlsx"
    workbook.save(special_path)
    monkeypatch.setattr(mod, "SPECIAL_RULES_XLSX", str(special_path))
    monkeypatch.setattr(
        mod,
        "_load_el_aggregates",
        lambda year, num_q: {
            "43932": {
                "year_hours": 8449.0,
                "period_hours": 4224.5,
                "year_te": 0.0,
                "period_te": 0.0,
            }
        },
    )

    section = mod.build_special_rule_section(
        year=2026,
        num_q=2,
        q_label="Q1-2",
        rows=[
            {"ea": "43932", "planned_value": "300000", "company": "4SOFT GMBH MUENCHEN", "title": "BM A", "bm_text": "", "status": "best"},
            {"ea": "43932", "planned_value": "224000", "company": "4SOFT GMBH MUENCHEN", "title": "BM B", "bm_text": "", "status": "flow"},
            {"ea": "43932", "planned_value": "524000", "company": "4SOFT GMBH MUENCHEN", "title": "BM C", "bm_text": "", "status": "concept"},
        ],
        status_map={"best": "bestellt", "flow": "im Durchlauf", "concept": "Konzept"},
    )

    assert section is not None
    assert section.headers == [
        "Thema",
        "EA",
        "Soll",
        "Ist",
        "DIFF Ges.",
        "Soll Q1-2",
        "bestellt",
        "im Durchlauf",
        "01 Erstellung",
        "EL DIFF",
        "DIFF Q1-2",
        "Bemerkung",
        "Hinweise",
        "IO/nIO",
    ]
    assert len(section.rows) == 3
    assert section.rows[0].values[0] == "PMT"
    assert section.rows[0].values[1] == "43932"
    assert section.rows[0].values[9] == "0 h"
    assert section.rows[0].values[12] == "EL und FL müssen exakt passen!"
    assert section.rows[0].values[13] == "IO"
    assert section.rows[0].cell_statuses["DIFF Ges."] == "OK"
    assert section.rows[0].cell_statuses["DIFF Q1-2"] == "OK"
    assert section.rows[0].cell_statuses["EL DIFF"] == "OK"
    assert section.rows[0].cell_statuses["IO/nIO"] == "IO"
    assert section.rows[1].values[0] == ""
    assert section.rows[1].values[12].startswith("EL und FL dürfen nur Quartalsweise")
    assert section.rows[1].values[13] == "IO"
    assert section.rows[1].cell_statuses["IO/nIO"] == "IO"
    assert section.rows[2].values[12].startswith("Folgende Firmen")
    assert section.rows[2].values[13] == "IO"
    assert section.rows[2].cell_statuses["IO/nIO"] == "IO"


def test_build_special_rule_section_applies_diff_bands_to_el_and_quarter(tmp_path, monkeypatch):
    mod = load_module()
    workbook = Workbook()
    ws = workbook.active
    ws.append(["Sondervereinbarungen"])
    ws.append([])
    ws.append(["Thema", "EA", "EL", "FL", "Bemerkung", "Link", "Hinweise für autom. Auslegung"])
    ws.append([
        "PMT",
        "43932",
        "8.449h",
        1048,
        "Hinweis",
        "",
        "EL und FL müssen exakt passen!\n"
        "EL und FL dürfen nur Quartalsweise beauftragen und werden sehr genau beobachtet.\n"
        "Folgende Firmen können auf PMT buchen: 4soft, Thiesen, FES, Sumitomo in genau dieser Priorität.",
    ])
    special_path = tmp_path / "sondervereinbarungen_warn.xlsx"
    workbook.save(special_path)
    monkeypatch.setattr(mod, "SPECIAL_RULES_XLSX", str(special_path))
    monkeypatch.setattr(
        mod,
        "_load_el_aggregates",
        lambda year, num_q: {
            "43932": {
                "year_hours": 8871.45,
                "period_hours": 4224.5,
                "year_te": 0.0,
                "period_te": 0.0,
            }
        },
    )

    section = mod.build_special_rule_section(
        year=2026,
        num_q=2,
        q_label="Q1-2",
        rows=[
            {"ea": "43932", "planned_value": "275000", "company": "4SOFT GMBH MUENCHEN", "title": "BM A", "bm_text": "", "status": "best"},
            {"ea": "43932", "planned_value": "275000", "company": "4SOFT GMBH MUENCHEN", "title": "BM B", "bm_text": "", "status": "flow"},
            {"ea": "43932", "planned_value": "498000", "company": "4SOFT GMBH MUENCHEN", "title": "BM C", "bm_text": "", "status": "concept"},
        ],
        status_map={"best": "bestellt", "flow": "im Durchlauf", "concept": "Konzept"},
    )

    assert section is not None
    assert len(section.rows) == 3
    # Row 0: "EL und FL müssen exakt passen!" -> FL annual band = OK -> IO
    assert section.rows[0].cell_statuses["DIFF Ges."] == "OK"
    assert section.rows[0].cell_statuses["DIFF Q1-2"] == "WARN"
    assert section.rows[0].cell_statuses["EL DIFF"] == "WARN"
    assert section.rows[0].cell_statuses["IO/nIO"] == "IO"
    # Row 1: "Quartalsweise" -> worst(quarter_band=WARN, el_period_band=OK) -> nIO
    assert section.rows[1].values[13] == "nIO"
    assert section.rows[1].cell_statuses["IO/nIO"] == "nIO"
    # Row 2: "Folgende Firmen ... Priorität" -> IO
    assert section.rows[2].values[13] == "IO"
    assert section.rows[2].cell_statuses["IO/nIO"] == "IO"


def test_write_xlsx_report_styles_special_status_cells(tmp_path):
    mod = load_module()
    report = mod.ReportDocument(
        title="Testreport",
        meta_lines=["Meta"],
        sections=[
            mod.TableSection(
                title="Sondervereinbarungen",
                headers=[
                    "Thema",
                    "EA",
                    "Soll",
                    "Ist",
                    "DIFF Ges.",
                    "Soll Q1-2",
                    "bestellt",
                    "im Durchlauf",
                    "01 Erstellung",
                    "EL DIFF",
                    "DIFF Q1-2",
                    "Bemerkung",
                    "Hinweise",
                    "IO/nIO",
                ],
                rows=[
                    mod.TableRow([
                        "PMT",
                        "43932",
                        "1.048 T€",
                        "1.049,7 T€",
                        "+1,7 T€",
                        "524 T€",
                        "300 T€",
                        "224 T€",
                        "524 T€",
                        "0 T€",
                        "-5 T€",
                        "Hinweis",
                        "Notiz",
                        "nIO",
                    ], cell_statuses={"DIFF Ges.": "OK", "DIFF Q1-2": "X", "EL DIFF": "OK", "IO/nIO": "nIO"}),
                    mod.TableRow(
                        [""] * 12 + ["Folgezeile", "IO"],
                        cell_statuses={"IO/nIO": "IO"},
                    ),
                ],
                align_right={1, 2, 3, 4, 5, 6, 7, 8, 10},
            )
        ],
        max_columns=14,
    )
    output = tmp_path / "special.xlsx"

    mod.write_xlsx_report(report, str(output))

    workbook = load_workbook(output)
    overview = workbook["Übersicht"]
    row_idx = next(row for row in range(1, overview.max_row + 1) if overview[f"A{row}"].value == "PMT")

    assert overview.cell(row_idx, 5).value == 1.7
    assert overview.cell(row_idx, 5).fill.fgColor.rgb == "FFC6EFCE"
    assert overview.cell(row_idx, 10).value == 0
    assert overview.cell(row_idx, 10).fill.fgColor.rgb == "FFC6EFCE"
    assert overview.cell(row_idx, 11).value == -5
    assert overview.cell(row_idx, 11).fill.fgColor.rgb == "FFFFC7CE"
    assert overview.cell(row_idx, 14).value == "nIO"
    assert overview.cell(row_idx, 14).fill.fgColor.rgb == "FFFFC7CE"
    # Continuation row: only IO/nIO column colored
    cont_row = row_idx + 1
    assert overview.cell(cont_row, 1).value is None
    assert overview.cell(cont_row, 13).value == "Folgezeile"
    assert overview.cell(cont_row, 14).value == "IO"
    assert overview.cell(cont_row, 14).fill.fgColor.rgb == "FFC6EFCE"


def test_write_xlsx_report_styles_diff_ges_warn_yellow(tmp_path):
    mod = load_module()
    report = mod.ReportDocument(
        title="Testreport",
        meta_lines=["Meta"],
        sections=[
            mod.TableSection(
                title="Sondervereinbarungen",
                headers=[
                    "Thema",
                    "EA",
                    "Soll",
                    "Ist",
                    "DIFF Ges.",
                    "Soll Q1-2",
                    "bestellt",
                    "im Durchlauf",
                    "01 Erstellung",
                    "EL DIFF",
                    "DIFF Q1-2",
                    "Bemerkung",
                    "Hinweise",
                    "IO/nIO",
                ],
                rows=[
                    mod.TableRow(
                        [
                            "Warnfall",
                            "EA_WARN",
                            "100 T€",
                            "106 T€",
                            "+6 T€",
                            "50 T€",
                            "40 T€",
                            "10 T€",
                            "56 T€",
                            "+5 h",
                            "+3 T€",
                            "Hinweis",
                            "Notiz",
                            "nIO",
                        ],
                        cell_statuses={"DIFF Ges.": "WARN", "DIFF Q1-2": "WARN", "EL DIFF": "WARN", "IO/nIO": "nIO"},
                    )
                ],
                align_right={1, 2, 3, 4, 5, 6, 7, 8, 10},
            )
        ],
        max_columns=14,
    )
    output = tmp_path / "special_warn.xlsx"

    mod.write_xlsx_report(report, str(output))

    workbook = load_workbook(output)
    overview = workbook["Übersicht"]
    row_idx = next(row for row in range(1, overview.max_row + 1) if overview[f"A{row}"].value == "Warnfall")

    assert overview.cell(row_idx, 5).value == 6
    assert overview.cell(row_idx, 5).fill.fgColor.rgb == "FFFFEB9C"
    assert overview.cell(row_idx, 10).value == "+5 h"
    assert overview.cell(row_idx, 10).fill.fgColor.rgb == "FFFFEB9C"
    assert overview.cell(row_idx, 11).value == 3
    assert overview.cell(row_idx, 11).fill.fgColor.rgb == "FFFFEB9C"


def test_build_special_rule_section_digi_budget_el_kann_niedriger(tmp_path, monkeypatch):
    """Digi-budget: 'EL kann niedriger sein' checks only EL overshoot with IO/WARN/nIO."""
    mod = load_module()
    workbook = Workbook()
    ws = workbook.active
    ws.append(["Sondervereinbarungen"])
    ws.append([])
    ws.append(["Thema", "EA", "EL", "FL", "Bemerkung", "Link", "Hinweise für autom. Auslegung"])
    ws.append([
        "Digi-budget",
        "48040",
        "-",
        290,
        "Bemerkung",
        "",
        "FL muss exakt passen!\n"
        "EL kann niedriger sein, darf aber auf keinen Fall höher sein.\n"
        "Werte für 1. HJ müssen auch wirklich im 1.HJ beauftragt werden.",
    ])
    special_path = tmp_path / "sondervereinbarungen_digi.xlsx"
    workbook.save(special_path)
    monkeypatch.setattr(mod, "SPECIAL_RULES_XLSX", str(special_path))
    monkeypatch.setattr(
        mod,
        "_load_el_aggregates",
        lambda year, num_q: {
            "48040": {
                "year_hours": 0.0,
                "period_hours": 0.0,
                "year_te": 0.0,
                "period_te": 0.0,
            }
        },
    )

    section = mod.build_special_rule_section(
        year=2026,
        num_q=2,
        q_label="Q1-2",
        rows=[
            {
                "ea": "48040",
                "planned_value": "290000",
                "company": "4SOFT GMBH MUENCHEN",
                "title": "Entwicklungsleistung VW-EK-26-0001",
                "bm_text": "Projekt Digitalisierung DMU",
                "status": "best",
            },
        ],
        status_map={"best": "bestellt"},
    )

    assert section is not None
    assert len(section.rows) == 3
    assert section.rows[0].values[5] == "290 T€"
    assert section.rows[0].values[10] == "0 T€"


def test_build_special_rule_section_digi_budget_filters_budget_rows_by_title_and_bm_text(tmp_path, monkeypatch):
    mod = load_module()
    workbook = Workbook()
    ws = workbook.active
    ws.append(["Sondervereinbarungen"])
    ws.append([])
    ws.append(["Thema", "EA", "EL", "FL", "Bemerkung", "Link", "Hinweise für autom. Auslegung"])
    ws.append([
        "Digi-budget",
        "48040",
        "-",
        290,
        "1. HJ DMU (170T€) + RuleChecker (120T€), aber mind. je 10T€ EL, 2. HJ noch kein Budget",
        "",
        "FL muss exakt passen!\nEL kann niedriger sein, darf aber auf keinen Fall höher sein.\nWerte für 1. HJ müssen auch wirklich im 1.HJ beauftragt werden.",
    ])
    special_path = tmp_path / "sondervereinbarungen_digi_filter.xlsx"
    workbook.save(special_path)
    monkeypatch.setattr(mod, "SPECIAL_RULES_XLSX", str(special_path))
    monkeypatch.setattr(
        mod,
        "_load_el_aggregates",
        lambda year, num_q: {
            "48040": {
                "year_hours": 0.0,
                "period_hours": 0.0,
                "year_te": 0.0,
                "period_te": 0.0,
            }
        },
    )

    section = mod.build_special_rule_section(
        year=2026,
        num_q=2,
        q_label="Q1-2",
        rows=[
            {
                "ea": "48040",
                "planned_value": "170000",
                "company": "4SOFT GMBH MUENCHEN",
                "title": "Entwicklungsleistung VW-EK-26-0001",
                "bm_text": "Projekt Digitalisierung DMU",
                "status": "best",
            },
            {
                "ea": "48040",
                "planned_value": "120000",
                "company": "4SOFT GMBH MUENCHEN",
                "title": "Entwicklungsleistung VW-EK-26-0002",
                "bm_text": "Digitalisierung RuleChecker",
                "status": "flow",
            },
            {
                "ea": "48040",
                "planned_value": "999000",
                "company": "4SOFT GMBH MUENCHEN",
                "title": "Falscher Abruf ohne Kennung",
                "bm_text": "Digitalisierung RuleChecker",
                "status": "flow",
            },
            {
                "ea": "48040",
                "planned_value": "888000",
                "company": "4SOFT GMBH MUENCHEN",
                "title": "Entwicklungsleistung VW-EK-26-0003",
                "bm_text": "Andere Maßnahme ohne Schlüsselwort",
                "status": "flow",
            },
        ],
        status_map={"best": "bestellt", "flow": "im Durchlauf"},
    )

    assert section is not None
    first_row = section.rows[0]
    assert first_row.values[0] == "Digi-budget"
    assert first_row.values[3] == "290 T€"
    assert first_row.values[5] == "290 T€"
    assert first_row.values[6] == "170 T€"
    assert first_row.values[7] == "120 T€"
    assert first_row.values[4] == "0 T€"
    assert first_row.values[10] == "0 T€"
    # Hint 0: "FL muss exakt passen!" -> year_band OK -> IO
    assert section.rows[0].values[13] == "IO"
    # Hint 1: "EL kann niedriger sein..." -> EL mode=manual, no target -> "-"
    assert section.rows[1].values[13] == "-"
    # Hint 2: "Werte für 1. HJ..." -> period_te (290) vs annual target (290) -> IO
    assert section.rows[2].values[13] == "IO"


def test_detect_special_cadence_and_period_target_for_first_half_exact():
    mod = load_module()

    cadence = mod._detect_special_cadence(
        "Digi-budget",
        "1. HJ DMU (170T€) + RuleChecker (120T€), 2. HJ noch kein Budget",
        "Werte für 1. HJ müssen auch wirklich im 1.HJ beauftragt werden.",
    )

    assert cadence == "first_half_exact"
    assert mod._period_target_for_cadence(290.0, cadence, 1) == 0.0
    assert mod._period_target_for_cadence(290.0, cadence, 2) == 290.0
    assert mod._period_target_for_cadence(290.0, cadence, 3) == 290.0


def test_period_target_for_semiannual_tranche_exact_is_half_year_share():
    mod = load_module()

    assert mod._period_target_for_cadence(150.0, "semiannual_tranche_exact", 1) == 0.0
    assert mod._period_target_for_cadence(150.0, "semiannual_tranche_exact", 2) == 75.0
    assert mod._period_target_for_cadence(150.0, "semiannual_tranche_exact", 3) == 75.0


def test_period_target_for_quarterly_tranche_exact_is_year_share():
    mod = load_module()

    assert mod._period_target_for_cadence(150.0, "quarterly_tranche_exact", 1) == 37.5
    assert mod._period_target_for_cadence(150.0, "quarterly_tranche_exact", 2) == 75.0
    assert mod._period_target_for_cadence(150.0, "quarterly_tranche_exact", 3) == 112.5
    assert mod._period_target_for_cadence(150.0, "quarterly_tranche_exact", 4) == 150.0


def test_build_special_rule_section_pro_halbjahr_uses_half_year_target(tmp_path, monkeypatch):
    mod = load_module()
    workbook = Workbook()
    ws = workbook.active
    ws.append(["Sondervereinbarungen"])
    ws.append([])
    ws.append(["Thema", "EA", "EL", "FL", "Bemerkung", "Link", "Hinweise für autom. Auslegung"])
    ws.append([
        "NE Space Crafter",
        "22020",
        "1,9",
        150,
        "Pro Halbjahr, EL: Armin, Donato, Afshin, TCM",
        "",
        "EL und FL müssen exakt passen!\nEL und FL müssen hier pro Halbjahr beauftragt werden!",
    ])
    special_path = tmp_path / "sondervereinbarungen_halbjahr.xlsx"
    workbook.save(special_path)
    monkeypatch.setattr(mod, "SPECIAL_RULES_XLSX", str(special_path))
    monkeypatch.setattr(
        mod,
        "_load_el_aggregates",
        lambda year, num_q: {
            "22020": {
                "year_hours": 0.0,
                "period_hours": 0.0,
                "year_te": 1.9,
                "period_te": 0.0,
            }
        },
    )

    section = mod.build_special_rule_section(
        year=2026,
        num_q=2,
        q_label="Q1-2",
        rows=[
            {
                "ea": "22020",
                "planned_value": "75100",
                "company": "Firma A",
                "title": "BM A",
                "bm_text": "",
                "status": "best",
            }
        ],
        status_map={"best": "bestellt"},
    )

    assert section is not None
    first_row = section.rows[0]
    assert first_row.values[5] == "75 T€"
    assert first_row.values[6] == "75,1 T€"
    assert first_row.values[10] == "+0,1 T€"


def test_build_special_rule_section_pro_quartal_uses_year_share_target(tmp_path, monkeypatch):
    mod = load_module()
    workbook = Workbook()
    ws = workbook.active
    ws.append(["Sondervereinbarungen"])
    ws.append([])
    ws.append(["Thema", "EA", "EL", "FL", "Bemerkung", "Link", "Hinweise für autom. Auslegung"])
    ws.append([
        "NE ID Buzz AD",
        "35566",
        "2,9",
        150,
        "Pro Quartal, EL: Armin, Donato, Afshin, TCM",
        "",
        "EL und FL müssen exakt passen!\nEL und FL müssen hier pro por Quartal beauftragt werden!",
    ])
    special_path = tmp_path / "sondervereinbarungen_quartal.xlsx"
    workbook.save(special_path)
    monkeypatch.setattr(mod, "SPECIAL_RULES_XLSX", str(special_path))
    monkeypatch.setattr(
        mod,
        "_load_el_aggregates",
        lambda year, num_q: {
            "35566": {
                "year_hours": 0.0,
                "period_hours": 0.0,
                "year_te": 2.9,
                "period_te": 0.0,
            }
        },
    )

    section = mod.build_special_rule_section(
        year=2026,
        num_q=2,
        q_label="Q1-2",
        rows=[
            {
                "ea": "35566",
                "planned_value": "76300",
                "company": "Firma A",
                "title": "BM A",
                "bm_text": "",
                "status": "best",
            }
        ],
        status_map={"best": "bestellt"},
    )

    assert section is not None
    first_row = section.rows[0]
    assert first_row.values[5] == "75 T€"
    assert first_row.values[6] == "76,3 T€"
    assert first_row.values[10] == "+1,3 T€"


def test_build_special_rule_section_nur_fl_und_nur_el(tmp_path, monkeypatch):
    """Test 'Hier nur FL buchen' and 'Nur EL. Keine FL.' rules."""
    mod = load_module()
    workbook = Workbook()
    ws = workbook.active
    ws.append(["Sondervereinbarungen"])
    ws.append([])
    ws.append(["Thema", "EA", "EL", "FL", "Bemerkung", "Link", "Hinweise für autom. Auslegung"])
    ws.append([
        "VCTC (KÜE)",
        "93037",
        "",
        45,
        "KÜE",
        "",
        "Hier nur FL buchen.\nFL muss exakt passen.",
    ])
    ws.append([
        "Scania",
        "536",
        "15000",
        "",
        "Nur EL",
        "",
        "Nur EL. Keine FL.",
    ])
    special_path = tmp_path / "sondervereinbarungen_fl_el.xlsx"
    workbook.save(special_path)
    monkeypatch.setattr(mod, "SPECIAL_RULES_XLSX", str(special_path))
    monkeypatch.setattr(
        mod,
        "_load_el_aggregates",
        lambda year, num_q: {
            "93037": {
                "year_hours": 0.0,
                "period_hours": 0.0,
                "year_te": 0.0,
                "period_te": 0.0,
            },
            "536": {
                "year_hours": 100.0,
                "period_hours": 50.0,
                "year_te": 5.0,
                "period_te": 2.5,
            },
        },
    )

    section = mod.build_special_rule_section(
        year=2026,
        num_q=2,
        q_label="Q1-2",
        rows=[
            {"ea": "93037", "planned_value": "45000", "company": "FES GMBH ZWICKAU/00", "title": "BM A", "bm_text": "", "status": "best"},
        ],
        status_map={"best": "bestellt"},
    )

    assert section is not None
    # VCTC: 2 hints
    # Hint 0: "Hier nur FL buchen." -> EL=0 -> IO
    assert section.rows[0].values[0] == "VCTC (KÜE)"
    assert section.rows[0].values[13] == "IO"
    # Hint 1: "FL muss exakt passen." -> year_band
    assert section.rows[1].values[13] == "IO"
    # Scania: 1 hint
    # "Nur EL. Keine FL." -> FL=0 -> IO
    scania_row = next(r for r in section.rows if r.values[0] == "Scania")
    assert scania_row.values[13] == "IO"


def test_build_special_rule_section_nur_el_with_fl_is_nio(tmp_path, monkeypatch):
    """'Nur EL. Keine FL.' should be nIO when FL exists."""
    mod = load_module()
    workbook = Workbook()
    ws = workbook.active
    ws.append(["Sondervereinbarungen"])
    ws.append([])
    ws.append(["Thema", "EA", "EL", "FL", "Bemerkung", "Link", "Hinweise für autom. Auslegung"])
    ws.append([
        "Scania",
        "536",
        "15000",
        "",
        "Nur EL",
        "",
        "Nur EL. Keine FL.",
    ])
    special_path = tmp_path / "sondervereinbarungen_nio.xlsx"
    workbook.save(special_path)
    monkeypatch.setattr(mod, "SPECIAL_RULES_XLSX", str(special_path))
    monkeypatch.setattr(
        mod,
        "_load_el_aggregates",
        lambda year, num_q: {
            "536": {
                "year_hours": 100.0,
                "period_hours": 50.0,
                "year_te": 5.0,
                "period_te": 2.5,
            },
        },
    )

    section = mod.build_special_rule_section(
        year=2026,
        num_q=2,
        q_label="Q1-2",
        rows=[
            {"ea": "536", "planned_value": "50000", "company": "EDAG GMBH", "title": "BM", "bm_text": "", "status": "best"},
        ],
        status_map={"best": "bestellt"},
    )

    assert section is not None
    assert section.rows[0].values[0] == "Scania"
    # FL actual = 50 T€ > 0 -> nIO
    assert section.rows[0].values[13] == "nIO"
